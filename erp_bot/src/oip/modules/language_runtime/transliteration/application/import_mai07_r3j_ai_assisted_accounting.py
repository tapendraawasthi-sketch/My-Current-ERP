"""MAI-07R3J-AI-ASSISTED-ACCOUNTING-IMPORT — fail-closed engineering evidence import.

Imports ACCOUNTING_DOMAIN Round A workbooks accepted by the user after AI autofill.
This is NOT independent blinded human review, NOT linguist approval, NOT Round A lock,
NOT Round B, NOT frozen V3 gold, and NOT runtime/training promotion.

Governance flags are hard-coded and cannot be upgraded by workbook content.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from .mai07_r3ja_v3_agreement import ROUND_A_DISPOSITIONS, ROUND_B_ACCEPTABILITY
from .mai07_r3ja_v3_firewall import REPO
from .validate_mai07_r3ja_round_a import CONFIDENCE, REQUIRED_SHEETS, ROUND_A_HEADERS

PHASE = "MAI-07R3J-AI-ASSISTED-ACCOUNTING-IMPORT"
SCHEMA_ID = "mai07_v3_ai_assisted_accounting_import_v1"
ROLE_ID = "ACCOUNTING_DOMAIN"

EXPECTED_PACKAGE_SHA256 = "f558fefdc186ba79bbe2a8757569204b88ce1aa1ed27400cda7705c1551cdb68"
EXPECTED_ROW_COUNT = 611
EXPECTED_WORKBOOK_COUNT = 6

YES_NO = frozenset({"YES", "NO"})

# Template-only declaration values permitted; all reviewer-entered fields must stay blank.
ALLOWED_DECLARATION_DEFAULTS: dict[str, str] = {
    "role_id": ROLE_ID,
    "credential_status": "CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION",
}

USER_DECLARATION_FIELDS = (
    "reviewer_full_name",
    "reviewer_email",
    "contact_reference_id",
    "qualification_summary",
    "professional_linguist_credentials",
    "relevant_experience",
    "declaration_independent",
    "declaration_no_conflict_of_interest",
    "declaration_answers_are_own",
    "declaration_did_not_see_other_reviewers",
    "declaration_no_ai_autofill",
    "declaration_no_runtime_predictions_used",
    "declaration_date_utc",
    "signature_or_typed_name",
)

# Hard governance — workbook content cannot flip these.
FIXED_PROVENANCE: dict[str, Any] = {
    "review_method": "AI_ASSISTED_HUMAN_VERIFIED",
    "status": "AI_ASSISTED_HUMAN_VERIFIED",
    "user_confirmation": "USER_ACCEPTED_AI_SUGGESTIONS_WITHOUT_CHANGES",
    "independent_human_review": False,
    "row_by_row_independent_review_performed": False,
    "ai_autofill_used": True,
    "user_accepted": True,
    "professional_linguist_adjudication": False,
    "linguist_approved": False,
    "production_approved": False,
    "official_round_a_lock_eligible": False,
    "round_b_authorized": False,
    "frozen_v3_quality_gate_authorized": False,
    "prohibited_for_training": True,
    "eligible_for_frozen_quality_gold": False,
    "declaration_completed": False,
    "declaration_no_ai_autofill": False,
}

OFFICIAL_OPS_INBOX = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3"
    / "review_operations"
    / "round_a_inbox"
    / ROLE_ID
)
OFFICIAL_SOURCE_BATCHES = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3"
    / "review_operations"
    / "reviewer_packages"
    / ROLE_ID
    / "round_a"
)
DEFAULT_EVIDENCE_ROOT = (
    REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3_ai_assisted" / "accounting_domain"
)

_ROUND_B_PLACEHOLDER = re.compile(r"^\(Round B opens", re.IGNORECASE)
_V3R_ID = re.compile(r"^V3R-[0-9a-f]{12}$", re.IGNORECASE)


class Mai07AiAssistedAccountingImportError(ValueError):
    """Fail-closed AI-assisted accounting import error."""


@dataclass(frozen=True)
class CanonicalReviewRecord:
    review_id: str
    input_text: str
    highlighted_span: str
    disposition: str
    confidence: str
    reason_category: str
    natural_context_ok: str
    suspected_ambiguity: str
    reviewer_notes: str
    source_workbook: str
    source_row: int
    role_id: str
    review_method: str
    independent_human_review: bool
    ai_autofill_used: bool
    user_accepted: bool
    professional_linguist_adjudication: bool
    prohibited_for_training: bool
    eligible_for_frozen_quality_gold: bool


@dataclass
class ImportResult:
    ok: bool
    records: list[CanonicalReviewRecord] = field(default_factory=list)
    semantic_hash: str = ""
    package_sha256: str = ""
    workbook_hashes: dict[str, str] = field(default_factory=dict)
    disposition_distribution: dict[str, int] = field(default_factory=dict)
    confidence_distribution: dict[str, int] = field(default_factory=dict)
    reason_category_distribution: dict[str, int] = field(default_factory=dict)
    natural_context_ok_distribution: dict[str, int] = field(default_factory=dict)
    suspected_ambiguity_distribution: dict[str, int] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    provenance: dict[str, Any] = field(default_factory=dict)
    evidence_root: str = ""
    canonical_jsonl_path: str = ""
    import_report_path: str = ""


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return sha256_file(path)


def _write_jsonl(path: Path, records: list[CanonicalReviewRecord]) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        json.dumps(asdict(r), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for r in sorted(records, key=lambda x: x.review_id)
    ]
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8", newline="\n")
    return sha256_file(path)


def compute_semantic_hash(records: list[CanonicalReviewRecord]) -> str:
    payload = {
        "schema": SCHEMA_ID,
        "phase": PHASE,
        "provenance": {k: FIXED_PROVENANCE[k] for k in sorted(FIXED_PROVENANCE)},
        "records": [asdict(r) for r in sorted(records, key=lambda x: x.review_id)],
    }
    blob = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
        "utf-8"
    )
    return sha256_bytes(blob)


def assert_official_inbox_untouched() -> None:
    if not OFFICIAL_OPS_INBOX.exists():
        return
    xlsx = list(OFFICIAL_OPS_INBOX.glob("*.xlsx")) + list(OFFICIAL_OPS_INBOX.glob("*.xlsm"))
    if xlsx:
        raise Mai07AiAssistedAccountingImportError(
            f"official_round_a_inbox_not_empty:{[p.name for p in xlsx]}"
        )


def assert_not_writing_official_inbox(dest: Path) -> None:
    try:
        dest.resolve().relative_to(OFFICIAL_OPS_INBOX.resolve())
    except ValueError:
        return
    raise Mai07AiAssistedAccountingImportError("refusing_write_into_official_round_a_inbox")


def extract_package(zip_path: Path, dest_dir: Path) -> Path:
    dest_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        # reject path traversal
        for info in zf.infolist():
            name = info.filename.replace("\\", "/")
            if name.startswith("/") or ".." in name.split("/"):
                raise Mai07AiAssistedAccountingImportError(f"unsafe_zip_entry:{info.filename}")
        zf.extractall(dest_dir)
    round_a = dest_dir / ROLE_ID / "round_a"
    if not round_a.is_dir():
        raise Mai07AiAssistedAccountingImportError("missing_ACCOUNTING_DOMAIN/round_a")
    return round_a


def verify_package_sha256(zip_path: Path, expected: str = EXPECTED_PACKAGE_SHA256) -> str:
    actual = sha256_file(zip_path)
    if actual.lower() != expected.lower():
        raise Mai07AiAssistedAccountingImportError(
            f"package_sha256_mismatch:expected={expected}:actual={actual}"
        )
    return actual.lower()


def load_and_validate_provenance(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    required = {
        "status": "AI_ASSISTED_HUMAN_VERIFIED",
        "user_confirmation": "USER_ACCEPTED_AI_SUGGESTIONS_WITHOUT_CHANGES",
        "independent_human_review": False,
        "row_by_row_independent_review_performed": False,
        "ai_autofill_used": True,
        "professional_linguist_adjudication": False,
        "linguist_approved": False,
        "production_approved": False,
        "official_round_a_lock_eligible": False,
        "round_b_authorized": False,
        "frozen_v3_quality_gate_authorized": False,
        "engineering_import_ready": True,
        "declaration_completed": False,
        "row_count": EXPECTED_ROW_COUNT,
        "workbook_count": EXPECTED_WORKBOOK_COUNT,
    }
    errors: list[str] = []
    for key, expected in required.items():
        if data.get(key) != expected:
            errors.append(f"provenance_mismatch:{key}:got={data.get(key)!r}:expected={expected!r}")
    if errors:
        raise Mai07AiAssistedAccountingImportError(";".join(errors))
    return data


def load_and_validate_manifest(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    files = data.get("files") or []
    if len(files) != EXPECTED_WORKBOOK_COUNT:
        raise Mai07AiAssistedAccountingImportError(
            f"manifest_workbook_count:{len(files)}!={EXPECTED_WORKBOOK_COUNT}"
        )
    total_rows = sum(int(f.get("rows") or 0) for f in files)
    if total_rows != EXPECTED_ROW_COUNT:
        raise Mai07AiAssistedAccountingImportError(
            f"manifest_row_count:{total_rows}!={EXPECTED_ROW_COUNT}"
        )
    if data.get("all_review_ids_unique") is not True:
        raise Mai07AiAssistedAccountingImportError("manifest_all_review_ids_unique_not_true")
    if data.get("all_rows_complete") is not True:
        raise Mai07AiAssistedAccountingImportError("manifest_all_rows_complete_not_true")
    prov = data.get("provenance") or {}
    expected_nested = {
        "status": "AI_ASSISTED_HUMAN_VERIFIED",
        "user_confirmation": "USER_ACCEPTED_AI_SUGGESTIONS_WITHOUT_CHANGES",
        "independent_human_review": False,
        "row_by_row_independent_review_performed": False,
        "ai_autofill_used": True,
        "professional_linguist_adjudication": False,
        "linguist_approved": False,
        "production_approved": False,
        "official_round_a_lock_eligible": False,
        "round_b_authorized": False,
        "frozen_v3_quality_gate_authorized": False,
        "engineering_import_ready": True,
        "declaration_completed": False,
    }
    for key, expected in expected_nested.items():
        if prov.get(key) != expected:
            raise Mai07AiAssistedAccountingImportError(
                f"manifest_provenance_mismatch:{key}:{prov.get(key)!r}"
            )
    return data


def load_source_batch_ids(path: Path) -> dict[str, list[str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, list[str]] = {}
    all_ids: list[str] = []
    for batch in data.get("batches") or []:
        fname = str(batch["filename"])
        ids = [str(x) for x in batch["review_ids"]]
        out[fname] = ids
        all_ids.extend(ids)
    if len(all_ids) != EXPECTED_ROW_COUNT:
        raise Mai07AiAssistedAccountingImportError(
            f"source_batch_row_count:{len(all_ids)}!={EXPECTED_ROW_COUNT}"
        )
    if len(set(all_ids)) != len(all_ids):
        raise Mai07AiAssistedAccountingImportError("source_batch_duplicate_review_ids")
    return out


def load_official_authority_map(source_batches_dir: Path = OFFICIAL_SOURCE_BATCHES) -> dict[str, tuple[str, str]]:
    """review_id → (input_text, highlighted_span) from official blank Round A batches."""
    authority: dict[str, tuple[str, str]] = {}
    if not source_batches_dir.is_dir():
        raise Mai07AiAssistedAccountingImportError(f"missing_official_source_batches:{source_batches_dir}")
    for path in sorted(source_batches_dir.glob("MokXya_MAI07_V3__ACCOUNTING_DOMAIN__ROUND_A__BATCH_*.xlsx")):
        if "__AI_ASSISTED" in path.name:
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "ROUND_A_CONTEXT" not in wb.sheetnames:
                raise Mai07AiAssistedAccountingImportError(f"official_missing_ROUND_A:{path.name}")
            for raw in wb["ROUND_A_CONTEXT"].iter_rows(min_row=2, values_only=True):
                if raw is None or not raw[0]:
                    continue
                rid = _cell_str(raw[0])
                authority[rid] = (_cell_str(raw[1] if len(raw) > 1 else ""), _cell_str(raw[2] if len(raw) > 2 else ""))
        finally:
            wb.close()
    if len(authority) != EXPECTED_ROW_COUNT:
        raise Mai07AiAssistedAccountingImportError(
            f"official_authority_count:{len(authority)}!={EXPECTED_ROW_COUNT}"
        )
    return authority


def _reject_macros(path: Path) -> None:
    if path.suffix.lower() == ".xlsm":
        raise Mai07AiAssistedAccountingImportError(f"macro_enabled_extension:{path.name}")
    # OOXML: presence of vbaProject.bin
    with zipfile.ZipFile(path, "r") as zf:
        names = {n.replace("\\", "/").lower() for n in zf.namelist()}
        if any("vbaproject.bin" in n for n in names):
            raise Mai07AiAssistedAccountingImportError(f"macro_vba_project:{path.name}")


def _validate_declaration(wb: Any, workbook_name: str) -> None:
    if "REVIEWER_DECLARATION" not in wb.sheetnames:
        raise Mai07AiAssistedAccountingImportError(f"missing_REVIEWER_DECLARATION:{workbook_name}")
    decl: dict[str, str] = {}
    for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        decl[_cell_str(row[0])] = _cell_str(row[1] if len(row) > 1 else "")
    for field_name in USER_DECLARATION_FIELDS:
        val = decl.get(field_name, "")
        if val:
            raise Mai07AiAssistedAccountingImportError(
                f"populated_reviewer_declaration:{workbook_name}:{field_name}={val!r}"
            )
    for key, allowed in ALLOWED_DECLARATION_DEFAULTS.items():
        got = decl.get(key, "")
        if got and got != allowed:
            raise Mai07AiAssistedAccountingImportError(
                f"declaration_template_mismatch:{workbook_name}:{key}={got!r}"
            )
    # Explicit: never claim no-AI autofill
    no_ai = decl.get("declaration_no_ai_autofill", "").upper()
    if no_ai in {"TRUE", "YES", "1", "Y"}:
        raise Mai07AiAssistedAccountingImportError(
            f"declaration_no_ai_autofill_claimed_true:{workbook_name}"
        )


def _validate_round_b_uncompleted(wb: Any, workbook_name: str) -> None:
    if "ROUND_B_CANDIDATES" not in wb.sheetnames:
        raise Mai07AiAssistedAccountingImportError(f"missing_ROUND_B_CANDIDATES:{workbook_name}")
    ws = wb["ROUND_B_CANDIDATES"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    for raw in rows[1:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        rid = _cell_str(raw[0] if len(raw) > 0 else "")
        acceptability = _cell_str(raw[5] if len(raw) > 5 else "")
        entered = _cell_str(raw[6] if len(raw) > 6 else "")
        notes = _cell_str(raw[7] if len(raw) > 7 else "")
        if _ROUND_B_PLACEHOLDER.match(rid):
            if acceptability or entered:
                raise Mai07AiAssistedAccountingImportError(
                    f"round_b_placeholder_polluted:{workbook_name}"
                )
            continue
        if acceptability in ROUND_B_ACCEPTABILITY or entered or (
            _V3R_ID.match(rid) and (acceptability or entered or notes)
        ):
            raise Mai07AiAssistedAccountingImportError(
                f"round_b_completed:{workbook_name}:review_id={rid!r}"
            )
        # Any non-placeholder content with filled acceptability columns is a fail
        if acceptability or entered:
            raise Mai07AiAssistedAccountingImportError(
                f"round_b_answer_present:{workbook_name}:review_id={rid!r}"
            )


def _validate_dropdowns(path: Path, expected_rows: int) -> None:
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ws = wb["ROUND_A_CONTEXT"]
        dvs = list(ws.data_validations.dataValidation)
        formulas = {str(dv.formula1 or "") for dv in dvs}
        if not any("ENGLISH_IDENTITY_REQUIRED" in f for f in formulas):
            raise Mai07AiAssistedAccountingImportError(f"missing_disposition_dropdown:{path.name}")
        if not any("HIGH,MEDIUM,LOW" in f.replace(" ", "") or '"HIGH,MEDIUM,LOW"' in f for f in formulas):
            raise Mai07AiAssistedAccountingImportError(f"missing_confidence_dropdown:{path.name}")
        # ensure D/E ranges cover data rows (soft: at least one D and one E validation)
        sqrefs = " ".join(str(dv.sqref) for dv in dvs)
        if "D" not in sqrefs or "E" not in sqrefs:
            raise Mai07AiAssistedAccountingImportError(f"dropdown_sqref_missing_D_or_E:{path.name}")
        _ = expected_rows  # documented expectation; range size varies by batch
    finally:
        wb.close()


def parse_workbook(
    path: Path,
    *,
    expected_ids: list[str],
    authority: Mapping[str, tuple[str, str]],
    verified_sha256: str,
) -> list[CanonicalReviewRecord]:
    actual_sha = sha256_file(path)
    if actual_sha.lower() != verified_sha256.lower():
        raise Mai07AiAssistedAccountingImportError(
            f"workbook_sha256_mismatch:{path.name}:expected={verified_sha256}:actual={actual_sha}"
        )
    _reject_macros(path)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise Mai07AiAssistedAccountingImportError(
            f"cannot_open_workbook:{path.name}:{type(exc).__name__}:{exc}"
        ) from exc

    try:
        missing_sheets = set(REQUIRED_SHEETS) - set(wb.sheetnames)
        if missing_sheets:
            raise Mai07AiAssistedAccountingImportError(
                f"missing_sheets:{path.name}:{sorted(missing_sheets)}"
            )
        _validate_declaration(wb, path.name)
        _validate_round_b_uncompleted(wb, path.name)

        if "ROUND_A_CONTEXT" not in wb.sheetnames:
            raise Mai07AiAssistedAccountingImportError(f"missing_ROUND_A_CONTEXT:{path.name}")
        rows_iter = list(wb["ROUND_A_CONTEXT"].iter_rows(values_only=True))
        if not rows_iter:
            raise Mai07AiAssistedAccountingImportError(f"empty_ROUND_A_CONTEXT:{path.name}")
        header = tuple(_cell_str(c) for c in rows_iter[0])
        if header != ROUND_A_HEADERS:
            raise Mai07AiAssistedAccountingImportError(f"header_mismatch:{path.name}:{header}")

        records: list[CanonicalReviewRecord] = []
        seen_ids: set[str] = set()
        for excel_row_idx, raw in enumerate(rows_iter[1:], start=2):
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            cells = {
                ROUND_A_HEADERS[i]: _cell_str(raw[i] if i < len(raw) else "")
                for i in range(len(ROUND_A_HEADERS))
            }
            rid = cells["review_id"]
            if not rid:
                raise Mai07AiAssistedAccountingImportError(f"blank_review_id:{path.name}:row={excel_row_idx}")
            if rid in seen_ids:
                raise Mai07AiAssistedAccountingImportError(
                    f"duplicate_review_ids:{path.name}:{rid}"
                )
            seen_ids.add(rid)

            for field_name in (
                "disposition",
                "confidence",
                "reason_category",
                "natural_context_ok",
                "suspected_ambiguity",
                "reviewer_notes",
                "input_text",
                "highlighted_span",
            ):
                if not cells.get(field_name):
                    raise Mai07AiAssistedAccountingImportError(
                        f"missing_supporting_field:{path.name}:{rid}:{field_name}"
                    )

            disp = cells["disposition"]
            conf = cells["confidence"]
            if disp not in ROUND_A_DISPOSITIONS:
                raise Mai07AiAssistedAccountingImportError(
                    f"unknown_disposition:{path.name}:{rid}:{disp!r}"
                )
            if conf not in CONFIDENCE:
                raise Mai07AiAssistedAccountingImportError(
                    f"unknown_confidence:{path.name}:{rid}:{conf!r}"
                )
            if cells["natural_context_ok"] not in YES_NO:
                raise Mai07AiAssistedAccountingImportError(
                    f"invalid_natural_context_ok:{path.name}:{rid}:{cells['natural_context_ok']!r}"
                )
            if cells["suspected_ambiguity"] not in YES_NO:
                raise Mai07AiAssistedAccountingImportError(
                    f"invalid_suspected_ambiguity:{path.name}:{rid}:{cells['suspected_ambiguity']!r}"
                )

            if rid not in authority:
                raise Mai07AiAssistedAccountingImportError(f"unknown_review_id_vs_official:{rid}")
            auth_text, auth_span = authority[rid]
            if cells["input_text"] != auth_text:
                raise Mai07AiAssistedAccountingImportError(
                    f"changed_input_text:{rid}"
                )
            if cells["highlighted_span"] != auth_span:
                raise Mai07AiAssistedAccountingImportError(
                    f"changed_highlighted_span:{rid}"
                )

            records.append(
                CanonicalReviewRecord(
                    review_id=rid,
                    input_text=cells["input_text"],
                    highlighted_span=cells["highlighted_span"],
                    disposition=disp,
                    confidence=conf,
                    reason_category=cells["reason_category"],
                    natural_context_ok=cells["natural_context_ok"],
                    suspected_ambiguity=cells["suspected_ambiguity"],
                    reviewer_notes=cells["reviewer_notes"],
                    source_workbook=path.name,
                    source_row=excel_row_idx,
                    role_id=ROLE_ID,
                    review_method=FIXED_PROVENANCE["review_method"],
                    independent_human_review=False,
                    ai_autofill_used=True,
                    user_accepted=True,
                    professional_linguist_adjudication=False,
                    prohibited_for_training=True,
                    eligible_for_frozen_quality_gold=False,
                )
            )
    finally:
        wb.close()

    _validate_dropdowns(path, len(expected_ids))

    got_ids = [r.review_id for r in records]
    if len(got_ids) != len(set(got_ids)):
        raise Mai07AiAssistedAccountingImportError(f"duplicate_review_ids_in_workbook:{path.name}")
    if set(got_ids) != set(expected_ids):
        missing = sorted(set(expected_ids) - set(got_ids))
        extra = sorted(set(got_ids) - set(expected_ids))
        raise Mai07AiAssistedAccountingImportError(
            f"batch_id_mismatch:{path.name}:missing={len(missing)}:extra={len(extra)}"
        )
    if len(records) != len(expected_ids):
        raise Mai07AiAssistedAccountingImportError(
            f"batch_row_count:{path.name}:{len(records)}!={len(expected_ids)}"
        )
    return records


def _verified_filename_to_source_filename(verified_name: str) -> str:
    # MokXya_...BATCH_01_of_06__AI_ASSISTED_HUMAN_VERIFIED.xlsx → ...BATCH_01_of_06.xlsx
    return verified_name.replace("__AI_ASSISTED_HUMAN_VERIFIED", "")


def import_from_round_a_dir(
    round_a_dir: Path,
    *,
    package_sha256: str = "",
    evidence_root: Path | None = None,
    write_evidence: bool = True,
    official_authority: Mapping[str, tuple[str, str]] | None = None,
) -> ImportResult:
    """Validate and optionally materialize canonical evidence from an extracted round_a dir."""
    result = ImportResult(ok=False, package_sha256=package_sha256)
    try:
        assert_official_inbox_untouched()

        provenance_path = round_a_dir / "AI_ASSISTED_HUMAN_VERIFICATION_PROVENANCE.json"
        manifest_path = round_a_dir / "AI_ASSISTED_HUMAN_VERIFIED_MANIFEST.json"
        source_batch_path = round_a_dir / "SOURCE_BATCH_MANIFEST.json"
        for required in (provenance_path, manifest_path, source_batch_path):
            if not required.is_file():
                raise Mai07AiAssistedAccountingImportError(f"missing_package_file:{required.name}")

        provenance = load_and_validate_provenance(provenance_path)
        manifest = load_and_validate_manifest(manifest_path)
        batch_ids = load_source_batch_ids(source_batch_path)
        authority = official_authority or load_official_authority_map()

        all_records: list[CanonicalReviewRecord] = []
        workbook_hashes: dict[str, str] = {}

        for file_meta in manifest["files"]:
            fname = str(file_meta["filename"])
            path = round_a_dir / fname
            if not path.is_file():
                raise Mai07AiAssistedAccountingImportError(f"missing_workbook:{fname}")
            source_fname = _verified_filename_to_source_filename(fname)
            if source_fname not in batch_ids:
                raise Mai07AiAssistedAccountingImportError(
                    f"workbook_not_in_source_batch_manifest:{fname}"
                )
            expected_ids = batch_ids[source_fname]
            if int(file_meta["rows"]) != len(expected_ids):
                raise Mai07AiAssistedAccountingImportError(
                    f"manifest_vs_source_rows:{fname}:{file_meta['rows']}!={len(expected_ids)}"
                )
            verified = str(file_meta["verified_sha256"])
            if not file_meta.get("declaration_blank", False):
                raise Mai07AiAssistedAccountingImportError(f"manifest_declaration_not_blank:{fname}")
            if not file_meta.get("round_b_not_completed", False):
                raise Mai07AiAssistedAccountingImportError(f"manifest_round_b_completed:{fname}")

            records = parse_workbook(
                path,
                expected_ids=expected_ids,
                authority=authority,
                verified_sha256=verified,
            )
            workbook_hashes[fname] = verified.lower()
            all_records.extend(records)

        if len(all_records) != EXPECTED_ROW_COUNT:
            raise Mai07AiAssistedAccountingImportError(
                f"total_rows:{len(all_records)}!={EXPECTED_ROW_COUNT}"
            )
        ids = [r.review_id for r in all_records]
        if len(set(ids)) != EXPECTED_ROW_COUNT:
            raise Mai07AiAssistedAccountingImportError(
                f"unique_review_ids:{len(set(ids))}!={EXPECTED_ROW_COUNT}"
            )

        # Governance freeze: every record must carry fixed false/true flags
        for r in all_records:
            if r.independent_human_review or r.professional_linguist_adjudication:
                raise Mai07AiAssistedAccountingImportError("governance_flag_upgrade_attempt")
            if not r.ai_autofill_used or not r.prohibited_for_training:
                raise Mai07AiAssistedAccountingImportError("governance_flag_downgrade_attempt")
            if r.eligible_for_frozen_quality_gold:
                raise Mai07AiAssistedAccountingImportError("eligible_for_frozen_quality_gold_true")
            if r.review_method != "AI_ASSISTED_HUMAN_VERIFIED":
                raise Mai07AiAssistedAccountingImportError("review_method_mismatch")

        semantic = compute_semantic_hash(all_records)
        disp_c = Counter(r.disposition for r in all_records)
        conf_c = Counter(r.confidence for r in all_records)
        reason_c = Counter(r.reason_category for r in all_records)
        nat_c = Counter(r.natural_context_ok for r in all_records)
        amb_c = Counter(r.suspected_ambiguity for r in all_records)

        result.records = sorted(all_records, key=lambda x: x.review_id)
        result.semantic_hash = semantic
        result.workbook_hashes = workbook_hashes
        result.disposition_distribution = dict(sorted(disp_c.items()))
        result.confidence_distribution = dict(sorted(conf_c.items()))
        result.reason_category_distribution = dict(sorted(reason_c.items()))
        result.natural_context_ok_distribution = dict(sorted(nat_c.items()))
        result.suspected_ambiguity_distribution = dict(sorted(amb_c.items()))
        result.provenance = {**FIXED_PROVENANCE, "package_provenance": provenance}

        if write_evidence:
            root = evidence_root or DEFAULT_EVIDENCE_ROOT
            assert_not_writing_official_inbox(root)
            materialize_evidence(round_a_dir, root, result)
            result.evidence_root = str(root)
            result.canonical_jsonl_path = str(root / "canonical" / "ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl")
            result.import_report_path = str(root / "reports" / "IMPORT_REPORT.json")

        result.ok = True
        return result
    except Mai07AiAssistedAccountingImportError as exc:
        result.errors.append(str(exc))
        result.ok = False
        return result


def materialize_evidence(round_a_dir: Path, evidence_root: Path, result: ImportResult) -> None:
    """Append-only style evidence write: workbooks copied byte-for-byte; reports overwritten deterministically."""
    assert_not_writing_official_inbox(evidence_root)
    wb_dir = evidence_root / "evidence" / "workbooks"
    pkg_dir = evidence_root / "evidence" / "package_metadata"
    canon_dir = evidence_root / "canonical"
    reports_dir = evidence_root / "reports"
    wb_dir.mkdir(parents=True, exist_ok=True)
    pkg_dir.mkdir(parents=True, exist_ok=True)
    canon_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)

    for fname in result.workbook_hashes:
        src = round_a_dir / fname
        dest = wb_dir / fname
        data = src.read_bytes()
        if sha256_bytes(data).lower() != result.workbook_hashes[fname].lower():
            raise Mai07AiAssistedAccountingImportError(f"copy_hash_drift:{fname}")
        if dest.exists() and dest.read_bytes() != data:
            raise Mai07AiAssistedAccountingImportError(
                f"evidence_workbook_immutable_conflict:{fname}"
            )
        if not dest.exists():
            dest.write_bytes(data)

    for name in (
        "AI_ASSISTED_HUMAN_VERIFICATION_PROVENANCE.json",
        "AI_ASSISTED_HUMAN_VERIFIED_MANIFEST.json",
        "SOURCE_BATCH_MANIFEST.json",
        "README_FIRST.txt",
        "CURSOR_PROMPT_IMPORT_AI_ASSISTED_ACCOUNTING_REVIEW.txt",
        "SOURCE_INSTRUCTIONS.md",
    ):
        src = round_a_dir / name
        if src.is_file():
            shutil.copy2(src, pkg_dir / name)

    jsonl_path = canon_dir / "ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
    jsonl_sha = _write_jsonl(jsonl_path, result.records)

    report = {
        "phase": PHASE,
        "schema": SCHEMA_ID,
        "ok": True,
        "imported_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "package_sha256": result.package_sha256.lower() if result.package_sha256 else "",
        "expected_package_sha256": EXPECTED_PACKAGE_SHA256,
        "workbook_count": EXPECTED_WORKBOOK_COUNT,
        "row_count": EXPECTED_ROW_COUNT,
        "unique_review_ids": EXPECTED_ROW_COUNT,
        "semantic_hash": result.semantic_hash,
        "canonical_jsonl_sha256": jsonl_sha,
        "workbook_hashes": result.workbook_hashes,
        "disposition_distribution": result.disposition_distribution,
        "confidence_distribution": result.confidence_distribution,
        "reason_category_distribution": result.reason_category_distribution,
        "natural_context_ok_distribution": result.natural_context_ok_distribution,
        "suspected_ambiguity_distribution": result.suspected_ambiguity_distribution,
        "provenance": result.provenance,
        "governance": {
            "official_round_a_inbox_used": False,
            "ROUND_A_LOCKED": False,
            "ROUND_B_READY": False,
            "QUALITY_GATES_PASSED": False,
            "LINGUIST_APPROVED": False,
            "PRODUCTION_APPROVED": False,
            "MAI-07": "NEEDS_CORRECTIVE_WORK",
            "MAI-08": "NOT_STARTED",
            "independent_human_review": False,
            "professional_linguist_adjudication": False,
            "eligible_for_frozen_quality_gold": False,
            "prohibited_for_training": True,
        },
        "validation": {
            "workbook_hashes_verified": True,
            "source_batch_reconciled": True,
            "authority_fields_unchanged": True,
            "declarations_blank": True,
            "round_b_uncompleted": True,
            "macros_absent": True,
            "enums_valid": True,
        },
    }
    _write_json(reports_dir / "IMPORT_REPORT.json", report)
    _write_json(
        reports_dir / "SEMANTIC_HASH.json",
        {
            "phase": PHASE,
            "schema": SCHEMA_ID,
            "semantic_hash": result.semantic_hash,
            "row_count": EXPECTED_ROW_COUNT,
            "algorithm": "sha256(json.dumps({schema,phase,provenance,records_sorted}, sort_keys=True, separators=(',', ':')))",
        },
    )
    readme = evidence_root / "README.md"
    readme.write_text(
        "\n".join(
            [
                "# MAI-07 V3 ACCOUNTING_DOMAIN — AI-Assisted Human-Verified Engineering Evidence",
                "",
                "Status: `AI_ASSISTED_HUMAN_VERIFIED` (engineering import only).",
                "",
                "- `independent_human_review=false`",
                "- `professional_linguist_adjudication=false`",
                "- `eligible_for_frozen_quality_gold=false`",
                "- `prohibited_for_training=true`",
                "- Not placed in official `round_a_inbox`",
                "- Does not set `ROUND_A_LOCKED` / `ROUND_B_READY`",
                "- Does not authorize frozen V3 quality gates or MAI-08",
                "",
                f"Canonical JSONL: `canonical/ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl`",
                f"Semantic hash: `{result.semantic_hash}`",
                "",
            ]
        ),
        encoding="utf-8",
        newline="\n",
    )


def import_package(
    zip_path: Path,
    *,
    extract_dir: Path | None = None,
    evidence_root: Path | None = None,
    write_evidence: bool = True,
    reuse_extract: bool = False,
) -> ImportResult:
    package_sha = verify_package_sha256(zip_path)
    tmp = extract_dir or (REPO / "tmp_mai07_v3_ai_assisted_accounting")
    round_a_path = tmp / ROLE_ID / "round_a"
    if reuse_extract and round_a_path.is_dir():
        round_a = round_a_path
    else:
        if tmp.exists():
            shutil.rmtree(tmp)
        round_a = extract_package(zip_path, tmp)
    return import_from_round_a_dir(
        round_a,
        package_sha256=package_sha,
        evidence_root=evidence_root,
        write_evidence=write_evidence,
    )


def prove_deterministic_reimport(
    zip_path: Path,
    *,
    work_dir: Path,
) -> dict[str, Any]:
    """Run two isolated imports; require identical canonical JSONL bytes and semantic hash."""
    a_dir = work_dir / "run_a"
    b_dir = work_dir / "run_b"
    a_ev = work_dir / "ev_a"
    b_ev = work_dir / "ev_b"
    for d in (a_dir, b_dir, a_ev, b_ev):
        if d.exists():
            shutil.rmtree(d)
    r1 = import_package(zip_path, extract_dir=a_dir, evidence_root=a_ev, write_evidence=True)
    r2 = import_package(zip_path, extract_dir=b_dir, evidence_root=b_ev, write_evidence=True)
    if not r1.ok or not r2.ok:
        raise Mai07AiAssistedAccountingImportError(
            f"reimport_failed:r1={r1.errors}:r2={r2.errors}"
        )
    j1 = Path(r1.canonical_jsonl_path).read_bytes()
    j2 = Path(r2.canonical_jsonl_path).read_bytes()
    if j1 != j2:
        raise Mai07AiAssistedAccountingImportError("canonical_jsonl_not_deterministic")
    if r1.semantic_hash != r2.semantic_hash:
        raise Mai07AiAssistedAccountingImportError("semantic_hash_not_deterministic")
    return {
        "ok": True,
        "semantic_hash": r1.semantic_hash,
        "canonical_jsonl_sha256": sha256_bytes(j1),
        "row_count": len(r1.records),
    }


def assert_no_accounting_mutation_imports() -> None:
    """Static guard: this module must not import posting/accounting mutation surfaces."""
    src = Path(__file__).read_text(encoding="utf-8")
    import_lines = [
        ln.strip()
        for ln in src.splitlines()
        if ln.strip().startswith("import ") or ln.strip().startswith("from ")
    ]
    joined = "\n".join(import_lines)
    for bad in (
        "orbixPostingService",
        "post_voucher",
        "ledger_tools",
        "khataConfirm",
        "mutation_gateway",
        "execute_command",
    ):
        if bad in joined:
            raise Mai07AiAssistedAccountingImportError(f"forbidden_accounting_token:{bad}")


def main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(description=PHASE)
    parser.add_argument(
        "--zip",
        type=Path,
        default=REPO / "MokXya_MAI07_V3_ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED_READY.zip",
    )
    parser.add_argument("--extract-dir", type=Path, default=None)
    parser.add_argument("--evidence-root", type=Path, default=DEFAULT_EVIDENCE_ROOT)
    parser.add_argument("--no-write", action="store_true")
    parser.add_argument("--prove-deterministic", action="store_true")
    args = parser.parse_args(argv)

    assert_no_accounting_mutation_imports()
    if args.prove_deterministic:
        proof = prove_deterministic_reimport(
            args.zip, work_dir=REPO / "tmp_mai07_ai_assisted_det_proof"
        )
        print(json.dumps(proof, indent=2))
        return 0

    result = import_package(
        args.zip,
        extract_dir=args.extract_dir,
        evidence_root=args.evidence_root,
        write_evidence=not args.no_write,
    )
    if not result.ok:
        print(json.dumps({"ok": False, "errors": result.errors}, indent=2))
        return 1
    print(
        json.dumps(
            {
                "ok": True,
                "row_count": len(result.records),
                "semantic_hash": result.semantic_hash,
                "evidence_root": result.evidence_root,
                "canonical_jsonl_path": result.canonical_jsonl_path,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
