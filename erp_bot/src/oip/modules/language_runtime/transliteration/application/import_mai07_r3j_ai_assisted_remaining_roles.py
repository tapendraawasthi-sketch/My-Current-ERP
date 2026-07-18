"""MAI-07R3J-AI-ASSISTED-REMAINING-ROLES-IMPORT

Fail-closed engineering import of user-accepted AI-assisted Round A drafts for:
PRODUCT_POLICY, NEPALI_FLUENT_A, PROFESSIONAL_LINGUIST_B.

Governance (hard-coded; workbook content cannot upgrade):
- review_method = AI_ASSISTED_HUMAN_VERIFIED
- independent_human_review = false
- PROFESSIONAL_LINGUIST_B is an AI role simulation only
- professional_linguist_adjudication = false
- linguist_approved = false
- production_approved = false
- official_round_a_lock_eligible = false
- round_b_authorized = false
- frozen_v3_quality_gate_authorized = false
- prohibited_for_training = true

Does NOT write to official round_a_inbox, lock Round A, release Round B,
promote runtime, or claim frozen V3 gold.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from .mai07_r3ja_v3_agreement import ROUND_A_DISPOSITIONS, ROUND_B_ACCEPTABILITY
from .mai07_r3ja_v3_firewall import REPO
from .validate_mai07_r3ja_round_a import CONFIDENCE, REQUIRED_SHEETS, ROUND_A_HEADERS

PHASE = "MAI-07R3J-AI-ASSISTED-REMAINING-ROLES-IMPORT"
SCHEMA_ID = "mai07_v3_ai_assisted_remaining_roles_import_v1"

ROLES = (
    "PRODUCT_POLICY",
    "NEPALI_FLUENT_A",
    "PROFESSIONAL_LINGUIST_B",
)
EXPECTED_ROWS_PER_ROLE = 1111
EXPECTED_WORKBOOKS_PER_ROLE = 10
EXPECTED_TOTAL_ROWS = 3333

YES_NO = frozenset({"YES", "NO"})

USER_DECLARATION_FIELDS = (
    "reviewer_full_name",
    "reviewer_email",
    "contact_reference_id",
    "qualification_summary",
    "professional_linguist_credentials",
    "relevant_experience",
    "declaration_independent",
    "declaration_no_conflict_of_interest",
    "declaration_answers_are_own",
    "declaration_did_not_see_other_reviewers",
    "declaration_no_ai_autofill",
    "declaration_no_runtime_predictions_used",
    "declaration_date_utc",
    "signature_or_typed_name",
)

FIXED_PROVENANCE: dict[str, Any] = {
    "review_method": "AI_ASSISTED_HUMAN_VERIFIED",
    "status": "AI_ASSISTED_HUMAN_VERIFIED",
    "user_confirmation": "USER_ACCEPTED_AI_SUGGESTIONS_WITHOUT_CHANGES",
    "independent_human_review": False,
    "row_by_row_independent_review_performed": False,
    "ai_autofill_used": True,
    "user_accepted": True,
    "professional_linguist_adjudication": False,
    "linguist_approved": False,
    "production_approved": False,
    "official_round_a_lock_eligible": False,
    "round_b_authorized": False,
    "frozen_v3_quality_gate_authorized": False,
    "prohibited_for_training": True,
    "eligible_for_frozen_quality_gold": False,
    "declaration_completed": False,
    "declaration_no_ai_autofill": False,
    "professional_linguist_b_is_ai_role_simulation": True,
}

DRAFTS_ROOT = (
    REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3_ai_assisted" / "role_drafts"
)
OPS_PACKAGES = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3"
    / "review_operations"
    / "reviewer_packages"
)
OFFICIAL_INBOX = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3"
    / "review_operations"
    / "round_a_inbox"
)
DEFAULT_EVIDENCE_ROOT = (
    REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3_ai_assisted" / "remaining_roles"
)

_ROUND_B_PLACEHOLDER = re.compile(r"^\(Round B opens", re.IGNORECASE)
_V3R_ID = re.compile(r"^V3R-[0-9a-f]{12}$", re.IGNORECASE)


class Mai07AiAssistedRemainingImportError(ValueError):
    pass


@dataclass(frozen=True)
class CanonicalReviewRecord:
    review_id: str
    input_text: str
    highlighted_span: str
    disposition: str
    confidence: str
    reason_category: str
    natural_context_ok: str
    suspected_ambiguity: str
    reviewer_notes: str
    source_workbook: str
    source_row: int
    role_id: str
    review_method: str
    independent_human_review: bool
    ai_autofill_used: bool
    user_accepted: bool
    professional_linguist_adjudication: bool
    prohibited_for_training: bool
    eligible_for_frozen_quality_gold: bool
    professional_linguist_b_is_ai_role_simulation: bool


@dataclass
class RoleImportResult:
    role_id: str
    ok: bool = False
    records: list[CanonicalReviewRecord] = field(default_factory=list)
    workbook_hashes: dict[str, str] = field(default_factory=dict)
    disposition_distribution: dict[str, int] = field(default_factory=dict)
    confidence_distribution: dict[str, int] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    canonical_jsonl_path: str = ""
    canonical_jsonl_sha256: str = ""


@dataclass
class ImportBundleResult:
    ok: bool
    semantic_hash: str = ""
    total_rows: int = 0
    roles: list[RoleImportResult] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    evidence_root: str = ""
    provenance: dict[str, Any] = field(default_factory=dict)


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return sha256_file(path)


def _write_jsonl(path: Path, records: list[CanonicalReviewRecord]) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        json.dumps(asdict(r), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for r in sorted(records, key=lambda x: (x.role_id, x.review_id))
    ]
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8", newline="\n")
    return sha256_file(path)


def compute_semantic_hash(records: list[CanonicalReviewRecord]) -> str:
    payload = {
        "schema": SCHEMA_ID,
        "phase": PHASE,
        "provenance": {k: FIXED_PROVENANCE[k] for k in sorted(FIXED_PROVENANCE)},
        "records": [
            asdict(r) for r in sorted(records, key=lambda x: (x.role_id, x.review_id))
        ],
    }
    return sha256_bytes(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
            "utf-8"
        )
    )


def assert_official_inbox_untouched() -> None:
    if not OFFICIAL_INBOX.exists():
        return
    hits = list(OFFICIAL_INBOX.rglob("*.xlsx")) + list(OFFICIAL_INBOX.rglob("*.xlsm"))
    if hits:
        raise Mai07AiAssistedRemainingImportError(
            f"official_round_a_inbox_not_empty:{[p.name for p in hits]}"
        )


def assert_not_writing_official_inbox(dest: Path) -> None:
    try:
        dest.resolve().relative_to(OFFICIAL_INBOX.resolve())
    except ValueError:
        return
    raise Mai07AiAssistedRemainingImportError("refusing_write_into_official_round_a_inbox")


def load_official_authority(role_id: str) -> dict[str, tuple[str, str]]:
    src_dir = OPS_PACKAGES / role_id / "round_a"
    if not src_dir.is_dir():
        raise Mai07AiAssistedRemainingImportError(f"missing_official_batches:{role_id}")
    authority: dict[str, tuple[str, str]] = {}
    for path in sorted(src_dir.glob("MokXya_MAI07_V3__*.xlsx")):
        if "__AI_ASSISTED" in path.name:
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for raw in wb["ROUND_A_CONTEXT"].iter_rows(min_row=2, values_only=True):
                if raw is None or not raw[0]:
                    continue
                rid = _cell_str(raw[0])
                authority[rid] = (
                    _cell_str(raw[1] if len(raw) > 1 else ""),
                    _cell_str(raw[2] if len(raw) > 2 else ""),
                )
        finally:
            wb.close()
    if len(authority) != EXPECTED_ROWS_PER_ROLE:
        raise Mai07AiAssistedRemainingImportError(
            f"official_authority_count:{role_id}:{len(authority)}!={EXPECTED_ROWS_PER_ROLE}"
        )
    return authority


def load_source_batch_ids(role_id: str) -> dict[str, list[str]]:
    man_path = OPS_PACKAGES / role_id / "round_a" / "BATCH_MANIFEST.json"
    data = json.loads(man_path.read_text(encoding="utf-8"))
    out: dict[str, list[str]] = {}
    all_ids: list[str] = []
    for batch in data.get("batches") or []:
        fname = str(batch["filename"])
        ids = [str(x) for x in batch["review_ids"]]
        out[fname] = ids
        all_ids.extend(ids)
    if len(all_ids) != EXPECTED_ROWS_PER_ROLE or len(set(all_ids)) != len(all_ids):
        raise Mai07AiAssistedRemainingImportError(f"source_batch_ids_invalid:{role_id}")
    return out


def _reject_macros(path: Path) -> None:
    if path.suffix.lower() == ".xlsm":
        raise Mai07AiAssistedRemainingImportError(f"macro_enabled_extension:{path.name}")
    import zipfile

    with zipfile.ZipFile(path, "r") as zf:
        names = {n.replace("\\", "/").lower() for n in zf.namelist()}
        if any("vbaproject.bin" in n for n in names):
            raise Mai07AiAssistedRemainingImportError(f"macro_vba_project:{path.name}")


def _validate_declaration(wb: Any, workbook_name: str, role_id: str) -> None:
    if "REVIEWER_DECLARATION" not in wb.sheetnames:
        raise Mai07AiAssistedRemainingImportError(f"missing_REVIEWER_DECLARATION:{workbook_name}")
    decl: dict[str, str] = {}
    for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        decl[_cell_str(row[0])] = _cell_str(row[1] if len(row) > 1 else "")
    for field_name in USER_DECLARATION_FIELDS:
        if decl.get(field_name, ""):
            raise Mai07AiAssistedRemainingImportError(
                f"populated_reviewer_declaration:{workbook_name}:{field_name}"
            )
    role_val = decl.get("role_id", "")
    if role_val and role_val != role_id:
        raise Mai07AiAssistedRemainingImportError(
            f"declaration_role_mismatch:{workbook_name}:{role_val}"
        )
    no_ai = decl.get("declaration_no_ai_autofill", "").upper()
    if no_ai in {"TRUE", "YES", "1", "Y"}:
        raise Mai07AiAssistedRemainingImportError(
            f"declaration_no_ai_autofill_claimed_true:{workbook_name}"
        )


def _validate_round_b_uncompleted(wb: Any, workbook_name: str) -> None:
    if "ROUND_B_CANDIDATES" not in wb.sheetnames:
        raise Mai07AiAssistedRemainingImportError(f"missing_ROUND_B_CANDIDATES:{workbook_name}")
    for raw in wb["ROUND_B_CANDIDATES"].iter_rows(min_row=2, values_only=True):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        rid = _cell_str(raw[0] if len(raw) > 0 else "")
        acceptability = _cell_str(raw[5] if len(raw) > 5 else "")
        entered = _cell_str(raw[6] if len(raw) > 6 else "")
        notes = _cell_str(raw[7] if len(raw) > 7 else "")
        if _ROUND_B_PLACEHOLDER.match(rid):
            if acceptability or entered:
                raise Mai07AiAssistedRemainingImportError(
                    f"round_b_placeholder_polluted:{workbook_name}"
                )
            continue
        if acceptability in ROUND_B_ACCEPTABILITY or entered or (
            _V3R_ID.match(rid) and (acceptability or entered or notes)
        ):
            raise Mai07AiAssistedRemainingImportError(
                f"round_b_completed:{workbook_name}:{rid}"
            )


def parse_verified_workbook(
    path: Path,
    *,
    role_id: str,
    expected_ids: list[str],
    authority: Mapping[str, tuple[str, str]],
) -> list[CanonicalReviewRecord]:
    _reject_macros(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise Mai07AiAssistedRemainingImportError(
            f"cannot_open_workbook:{path.name}:{type(exc).__name__}"
        ) from exc

    try:
        missing = set(REQUIRED_SHEETS) - set(wb.sheetnames)
        if missing:
            raise Mai07AiAssistedRemainingImportError(
                f"missing_sheets:{path.name}:{sorted(missing)}"
            )
        _validate_declaration(wb, path.name, role_id)
        _validate_round_b_uncompleted(wb, path.name)

        rows_iter = list(wb["ROUND_A_CONTEXT"].iter_rows(values_only=True))
        header = tuple(_cell_str(c) for c in rows_iter[0])
        if header != ROUND_A_HEADERS:
            raise Mai07AiAssistedRemainingImportError(f"header_mismatch:{path.name}")

        records: list[CanonicalReviewRecord] = []
        seen: set[str] = set()
        for excel_row_idx, raw in enumerate(rows_iter[1:], start=2):
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            cells = {
                ROUND_A_HEADERS[i]: _cell_str(raw[i] if i < len(raw) else "")
                for i in range(len(ROUND_A_HEADERS))
            }
            rid = cells["review_id"]
            if not rid:
                raise Mai07AiAssistedRemainingImportError(
                    f"blank_review_id:{path.name}:row={excel_row_idx}"
                )
            if rid in seen:
                raise Mai07AiAssistedRemainingImportError(
                    f"duplicate_review_ids:{path.name}:{rid}"
                )
            seen.add(rid)

            for field_name in (
                "disposition",
                "confidence",
                "reason_category",
                "natural_context_ok",
                "suspected_ambiguity",
                "reviewer_notes",
                "input_text",
                "highlighted_span",
            ):
                if not cells.get(field_name):
                    raise Mai07AiAssistedRemainingImportError(
                        f"missing_supporting_field:{path.name}:{rid}:{field_name}"
                    )

            disp = cells["disposition"]
            conf = cells["confidence"]
            if disp not in ROUND_A_DISPOSITIONS:
                raise Mai07AiAssistedRemainingImportError(
                    f"unknown_disposition:{path.name}:{rid}:{disp!r}"
                )
            if conf not in CONFIDENCE:
                raise Mai07AiAssistedRemainingImportError(
                    f"unknown_confidence:{path.name}:{rid}:{conf!r}"
                )
            if cells["natural_context_ok"] not in YES_NO:
                raise Mai07AiAssistedRemainingImportError(
                    f"invalid_natural_context_ok:{rid}"
                )
            if cells["suspected_ambiguity"] not in YES_NO:
                raise Mai07AiAssistedRemainingImportError(
                    f"invalid_suspected_ambiguity:{rid}"
                )
            if rid not in authority:
                raise Mai07AiAssistedRemainingImportError(f"unknown_review_id:{role_id}:{rid}")
            auth_text, auth_span = authority[rid]
            if cells["input_text"] != auth_text:
                raise Mai07AiAssistedRemainingImportError(f"changed_input_text:{rid}")
            if cells["highlighted_span"] != auth_span:
                raise Mai07AiAssistedRemainingImportError(f"changed_highlighted_span:{rid}")

            records.append(
                CanonicalReviewRecord(
                    review_id=rid,
                    input_text=cells["input_text"],
                    highlighted_span=cells["highlighted_span"],
                    disposition=disp,
                    confidence=conf,
                    reason_category=cells["reason_category"],
                    natural_context_ok=cells["natural_context_ok"],
                    suspected_ambiguity=cells["suspected_ambiguity"],
                    reviewer_notes=cells["reviewer_notes"],
                    source_workbook=path.name,
                    source_row=excel_row_idx,
                    role_id=role_id,
                    review_method=FIXED_PROVENANCE["review_method"],
                    independent_human_review=False,
                    ai_autofill_used=True,
                    user_accepted=True,
                    professional_linguist_adjudication=False,
                    prohibited_for_training=True,
                    eligible_for_frozen_quality_gold=False,
                    professional_linguist_b_is_ai_role_simulation=(
                        role_id == "PROFESSIONAL_LINGUIST_B"
                    ),
                )
            )
    finally:
        wb.close()

    got_ids = [r.review_id for r in records]
    if set(got_ids) != set(expected_ids) or len(got_ids) != len(expected_ids):
        raise Mai07AiAssistedRemainingImportError(
            f"batch_id_mismatch:{path.name}:got={len(got_ids)}:expected={len(expected_ids)}"
        )
    return records


def _draft_dir_for_role(role_id: str, drafts_root: Path = DRAFTS_ROOT) -> Path:
    return drafts_root / role_id.lower() / "round_a_drafts"


def _source_filename_from_draft(draft_name: str) -> str:
    # ...BATCH_01_of_10__AI_ASSISTED_DRAFT.xlsx → ...BATCH_01_of_10.xlsx
    return draft_name.replace("__AI_ASSISTED_DRAFT", "").replace(
        "__AI_ASSISTED_HUMAN_VERIFIED", ""
    )


def import_role(
    role_id: str,
    *,
    drafts_root: Path = DRAFTS_ROOT,
    evidence_root: Path,
    write_evidence: bool = True,
) -> RoleImportResult:
    result = RoleImportResult(role_id=role_id)
    try:
        if role_id not in ROLES:
            raise Mai07AiAssistedRemainingImportError(f"unexpected_role:{role_id}")
        draft_dir = _draft_dir_for_role(role_id, drafts_root)
        if not draft_dir.is_dir():
            raise Mai07AiAssistedRemainingImportError(f"missing_draft_dir:{draft_dir}")

        authority = load_official_authority(role_id)
        batch_ids = load_source_batch_ids(role_id)
        drafts = sorted(draft_dir.glob("*__AI_ASSISTED_DRAFT.xlsx"))
        if len(drafts) != EXPECTED_WORKBOOKS_PER_ROLE:
            raise Mai07AiAssistedRemainingImportError(
                f"workbook_count:{role_id}:{len(drafts)}!={EXPECTED_WORKBOOKS_PER_ROLE}"
            )

        all_records: list[CanonicalReviewRecord] = []
        workbook_hashes: dict[str, str] = {}
        for draft_path in drafts:
            source_fname = _source_filename_from_draft(draft_path.name)
            if source_fname not in batch_ids:
                raise Mai07AiAssistedRemainingImportError(
                    f"draft_not_in_batch_manifest:{role_id}:{draft_path.name}"
                )
            expected_ids = batch_ids[source_fname]
            records = parse_verified_workbook(
                draft_path,
                role_id=role_id,
                expected_ids=expected_ids,
                authority=authority,
            )
            digest = sha256_file(draft_path)
            verified_name = draft_path.name.replace(
                "__AI_ASSISTED_DRAFT", "__AI_ASSISTED_HUMAN_VERIFIED"
            )
            workbook_hashes[verified_name] = digest
            # rewrite source workbook name on records for evidence naming
            all_records.extend(
                [
                    CanonicalReviewRecord(
                        **{
                            **asdict(r),
                            "source_workbook": verified_name,
                        }
                    )
                    for r in records
                ]
            )

        if len(all_records) != EXPECTED_ROWS_PER_ROLE:
            raise Mai07AiAssistedRemainingImportError(
                f"row_count:{role_id}:{len(all_records)}!={EXPECTED_ROWS_PER_ROLE}"
            )
        if len({r.review_id for r in all_records}) != EXPECTED_ROWS_PER_ROLE:
            raise Mai07AiAssistedRemainingImportError(f"unique_ids:{role_id}")

        for r in all_records:
            if r.independent_human_review or r.professional_linguist_adjudication:
                raise Mai07AiAssistedRemainingImportError("governance_flag_upgrade_attempt")
            if r.eligible_for_frozen_quality_gold:
                raise Mai07AiAssistedRemainingImportError("frozen_gold_true")
            if role_id == "PROFESSIONAL_LINGUIST_B" and not r.professional_linguist_b_is_ai_role_simulation:
                raise Mai07AiAssistedRemainingImportError("linguist_simulation_flag_missing")

        result.records = sorted(all_records, key=lambda x: x.review_id)
        result.workbook_hashes = workbook_hashes
        result.disposition_distribution = dict(
            sorted(Counter(r.disposition for r in all_records).items())
        )
        result.confidence_distribution = dict(
            sorted(Counter(r.confidence for r in all_records).items())
        )

        if write_evidence:
            assert_not_writing_official_inbox(evidence_root)
            role_ev = evidence_root / role_id.lower()
            wb_dir = role_ev / "evidence" / "workbooks"
            wb_dir.mkdir(parents=True, exist_ok=True)
            for draft_path in drafts:
                verified_name = draft_path.name.replace(
                    "__AI_ASSISTED_DRAFT", "__AI_ASSISTED_HUMAN_VERIFIED"
                )
                dest = wb_dir / verified_name
                data = draft_path.read_bytes()
                if sha256_bytes(data) != workbook_hashes[verified_name]:
                    raise Mai07AiAssistedRemainingImportError(f"copy_hash_drift:{verified_name}")
                if dest.exists() and dest.read_bytes() != data:
                    raise Mai07AiAssistedRemainingImportError(
                        f"evidence_immutable_conflict:{verified_name}"
                    )
                if not dest.exists():
                    dest.write_bytes(data)

            # Provenance + manifest
            provenance = {
                **FIXED_PROVENANCE,
                "schema_version": "mai07-ai-assisted-review-provenance-v1",
                "scope": f"MAI-07_V3_{role_id}_ROUND_A",
                "user_confirmation_date": "2026-07-17",
                "engineering_import_ready": True,
                "row_count": EXPECTED_ROWS_PER_ROLE,
                "workbook_count": EXPECTED_WORKBOOKS_PER_ROLE,
                "role_id": role_id,
            }
            if role_id == "PROFESSIONAL_LINGUIST_B":
                provenance["note"] = (
                    "PROFESSIONAL_LINGUIST_B labels are an AI role simulation only; "
                    "not professional-linguist adjudication."
                )
            _write_json(role_ev / "AI_ASSISTED_HUMAN_VERIFICATION_PROVENANCE.json", provenance)

            by_wb = Counter(r.source_workbook for r in all_records)
            manifest = {
                "schema_version": "mai07-ai-assisted-human-verified-manifest-v1",
                "generated_date": "2026-07-17",
                "role_id": role_id,
                "provenance": provenance,
                "files": [
                    {
                        "filename": name,
                        "rows": by_wb[name],
                        "verified_sha256": digest,
                        "size_bytes": (wb_dir / name).stat().st_size,
                        "declaration_blank": True,
                        "round_b_not_completed": True,
                    }
                    for name, digest in sorted(workbook_hashes.items())
                ],
                "all_review_ids_unique": True,
                "all_rows_complete": True,
            }
            _write_json(role_ev / "AI_ASSISTED_HUMAN_VERIFIED_MANIFEST.json", manifest)

            jsonl = role_ev / "canonical" / f"{role_id}_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
            result.canonical_jsonl_sha256 = _write_jsonl(jsonl, result.records)
            result.canonical_jsonl_path = str(jsonl)

        result.ok = True
        return result
    except Mai07AiAssistedRemainingImportError as exc:
        result.errors.append(str(exc))
        result.ok = False
        return result


def import_all_remaining_roles(
    *,
    drafts_root: Path = DRAFTS_ROOT,
    evidence_root: Path | None = None,
    write_evidence: bool = True,
) -> ImportBundleResult:
    bundle = ImportBundleResult(ok=False, provenance=dict(FIXED_PROVENANCE))
    try:
        assert_official_inbox_untouched()
        root = evidence_root or DEFAULT_EVIDENCE_ROOT
        assert_not_writing_official_inbox(root)
        if write_evidence:
            root.mkdir(parents=True, exist_ok=True)

        role_results: list[RoleImportResult] = []
        all_records: list[CanonicalReviewRecord] = []
        for role_id in ROLES:
            rr = import_role(
                role_id,
                drafts_root=drafts_root,
                evidence_root=root,
                write_evidence=write_evidence,
            )
            role_results.append(rr)
            if not rr.ok:
                bundle.errors.extend(rr.errors)
                bundle.roles = role_results
                bundle.ok = False
                return bundle
            all_records.extend(rr.records)

        if len(all_records) != EXPECTED_TOTAL_ROWS:
            raise Mai07AiAssistedRemainingImportError(
                f"total_rows:{len(all_records)}!={EXPECTED_TOTAL_ROWS}"
            )
        # review_ids are role-scoped (blind); uniqueness is per-role already checked
        semantic = compute_semantic_hash(all_records)
        bundle.roles = role_results
        bundle.semantic_hash = semantic
        bundle.total_rows = len(all_records)
        bundle.evidence_root = str(root)
        bundle.provenance = dict(FIXED_PROVENANCE)

        if write_evidence:
            report = {
                "phase": PHASE,
                "schema": SCHEMA_ID,
                "ok": True,
                "imported_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "total_rows": EXPECTED_TOTAL_ROWS,
                "roles": [
                    {
                        "role_id": r.role_id,
                        "row_count": len(r.records),
                        "workbook_count": len(r.workbook_hashes),
                        "unique_review_ids": len({x.review_id for x in r.records}),
                        "disposition_distribution": r.disposition_distribution,
                        "confidence_distribution": r.confidence_distribution,
                        "workbook_hashes": r.workbook_hashes,
                        "canonical_jsonl_path": r.canonical_jsonl_path,
                        "canonical_jsonl_sha256": r.canonical_jsonl_sha256,
                    }
                    for r in role_results
                ],
                "semantic_hash": semantic,
                "provenance": FIXED_PROVENANCE,
                "governance": {
                    "official_round_a_inbox_used": False,
                    "ROUND_A_LOCKED": False,
                    "ROUND_B_READY": False,
                    "QUALITY_GATES_PASSED": False,
                    "LINGUIST_APPROVED": False,
                    "PRODUCTION_APPROVED": False,
                    "MAI-07": "NEEDS_CORRECTIVE_WORK",
                    "MAI-08": "NOT_STARTED",
                    "professional_linguist_b_is_ai_role_simulation": True,
                    "independent_human_review": False,
                    "professional_linguist_adjudication": False,
                    "eligible_for_frozen_quality_gold": False,
                    "prohibited_for_training": True,
                },
            }
            _write_json(root / "reports" / "IMPORT_REPORT.json", report)
            _write_json(
                root / "reports" / "SEMANTIC_HASH.json",
                {
                    "phase": PHASE,
                    "schema": SCHEMA_ID,
                    "semantic_hash": semantic,
                    "total_rows": EXPECTED_TOTAL_ROWS,
                    "algorithm": "sha256(json.dumps({schema,phase,provenance,records_sorted}, sort_keys=True))",
                },
            )
            combined = root / "canonical" / "REMAINING_ROLES_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
            _write_jsonl(combined, all_records)
            (root / "README.md").write_text(
                "\n".join(
                    [
                        "# MAI-07 V3 Remaining Roles — AI-Assisted Human-Verified Engineering Evidence",
                        "",
                        "Status: `AI_ASSISTED_HUMAN_VERIFIED` (engineering import only).",
                        "",
                        "- PRODUCT_POLICY / NEPALI_FLUENT_A / PROFESSIONAL_LINGUIST_B",
                        "- `independent_human_review=false`",
                        "- `PROFESSIONAL_LINGUIST_B` is an **AI role simulation only**",
                        "- `professional_linguist_adjudication=false`",
                        "- `linguist_approved=false` / `production_approved=false`",
                        "- Not official Round A lock / Round B / frozen V3 gold / training",
                        "",
                        f"Semantic hash: `{semantic}`",
                        "",
                    ]
                ),
                encoding="utf-8",
                newline="\n",
            )

        bundle.ok = True
        return bundle
    except Mai07AiAssistedRemainingImportError as exc:
        bundle.errors.append(str(exc))
        bundle.ok = False
        return bundle


def prove_deterministic_reimport(
    *,
    drafts_root: Path = DRAFTS_ROOT,
    work_dir: Path,
) -> dict[str, Any]:
    a = import_all_remaining_roles(
        drafts_root=drafts_root, evidence_root=work_dir / "ev_a", write_evidence=True
    )
    b = import_all_remaining_roles(
        drafts_root=drafts_root, evidence_root=work_dir / "ev_b", write_evidence=True
    )
    if not a.ok or not b.ok:
        raise Mai07AiAssistedRemainingImportError(f"reimport_failed:{a.errors}:{b.errors}")
    if a.semantic_hash != b.semantic_hash:
        raise Mai07AiAssistedRemainingImportError("semantic_hash_not_deterministic")
    j1 = (
        work_dir
        / "ev_a"
        / "canonical"
        / "REMAINING_ROLES_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
    ).read_bytes()
    j2 = (
        work_dir
        / "ev_b"
        / "canonical"
        / "REMAINING_ROLES_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
    ).read_bytes()
    if j1 != j2:
        raise Mai07AiAssistedRemainingImportError("canonical_jsonl_not_deterministic")
    return {
        "ok": True,
        "semantic_hash": a.semantic_hash,
        "canonical_jsonl_sha256": sha256_bytes(j1),
        "total_rows": a.total_rows,
    }


def main(argv: list[str] | None = None) -> int:
    import argparse

    p = argparse.ArgumentParser(description=PHASE)
    p.add_argument("--drafts-root", type=Path, default=DRAFTS_ROOT)
    p.add_argument("--evidence-root", type=Path, default=DEFAULT_EVIDENCE_ROOT)
    p.add_argument("--prove-deterministic", action="store_true")
    args = p.parse_args(argv)
    if args.prove_deterministic:
        proof = prove_deterministic_reimport(
            drafts_root=args.drafts_root,
            work_dir=REPO / "tmp_mai07_remaining_roles_det_proof",
        )
        print(json.dumps(proof, indent=2))
        return 0
    result = import_all_remaining_roles(
        drafts_root=args.drafts_root,
        evidence_root=args.evidence_root,
        write_evidence=True,
    )
    if not result.ok:
        print(json.dumps({"ok": False, "errors": result.errors}, indent=2))
        return 1
    print(
        json.dumps(
            {
                "ok": True,
                "total_rows": result.total_rows,
                "semantic_hash": result.semantic_hash,
                "evidence_root": result.evidence_root,
                "roles": [r.role_id for r in result.roles],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
