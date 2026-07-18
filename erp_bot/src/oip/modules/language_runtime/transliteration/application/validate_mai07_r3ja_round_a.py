"""MAI-07R3J-A — Round A lock-and-validation (governance only).

Validates returned Round A workbooks against the sealed V3 packet.
Does not fill/correct human decisions, release Round B on failure,
modify blind mapping, run model evaluation, or touch runtime.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .mai07_r3ja_v3_agreement import ROUND_A_DISPOSITIONS, cohen_kappa, exact_agreement
from .mai07_r3ja_v3_firewall import REPO

OUT = REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3"
PACKET_MANIFEST = OUT / "V3_PACKET_MANIFEST.json"
BLIND_MAPPING = OUT / "V3_BLIND_MAPPING.json"
SCHEMA = OUT / "V3_REVIEW_SCHEMA.json"
REVIEWERS_DIR = OUT / "reviewers"
INBOX = OUT / "round_a_submissions_inbox"
LOCKED_DIR = OUT / "round_a_locked_submissions"
VALIDATION_DIR = OUT / "round_a_validation"
ROUND_B_RELEASE_DIR = OUT / "round_b_release"

EXPECTED_PACKET_MANIFEST_SHA = "29d16a3ee43d4981515d31a3763aa277a5d98dd5ab84499d58ead7da8723fc6c"
EXPECTED_BLIND_MAPPING_SHA = "d0875db79185b034b080e69f77f1220417cdc24dae5a6fb755a56b472af414f1"

REQUIRED_SHEETS = (
    "START_HERE",
    "REVIEWER_DECLARATION",
    "ROUND_A_CONTEXT",
    "ROUND_B_CANDIDATES",
    "REVIEW_PROGRESS",
    "VALIDATION_ERRORS",
    "SUBMISSION_CHECKLIST",
)

ROUND_A_HEADERS = (
    "review_id",
    "input_text",
    "highlighted_span",
    "disposition",
    "confidence",
    "reason_category",
    "natural_context_ok",
    "suspected_ambiguity",
    "reviewer_notes",
)

CONFIDENCE = frozenset({"HIGH", "MEDIUM", "LOW"})
CSV_INJECT = re.compile(r"^[=+\-@]")

ROLE_FILES = {
    "PRODUCT_POLICY": "MokXya_MAI07_V3__product_policy.xlsx",
    "NEPALI_FLUENT_A": "MokXya_MAI07_V3__nepali_fluent_a.xlsx",
    "PROFESSIONAL_LINGUIST_B": "MokXya_MAI07_V3__professional_linguist_b.xlsx",
    "ACCOUNTING_DOMAIN": "MokXya_MAI07_V3__accounting_domain.xlsx",
}

# Roles required for Round A lock (adjudicator is Round C only).
REQUIRED_ROUND_A_ROLES = tuple(ROLE_FILES.keys())


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _sha_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return _sha(path)


def verify_sealed_authorities() -> dict[str, Any]:
    man_sha = _sha(PACKET_MANIFEST)
    blind_sha = _sha(BLIND_MAPPING)
    errors: list[str] = []
    if man_sha != EXPECTED_PACKET_MANIFEST_SHA:
        errors.append(f"packet_manifest_hash_mismatch:{man_sha}")
    if blind_sha != EXPECTED_BLIND_MAPPING_SHA:
        errors.append(f"blind_mapping_hash_mismatch:{blind_sha}")
    return {
        "ok": not errors,
        "errors": errors,
        "packet_manifest_sha256": man_sha,
        "blind_mapping_sha256": blind_sha,
        "expected_packet_manifest_sha256": EXPECTED_PACKET_MANIFEST_SHA,
        "expected_blind_mapping_sha256": EXPECTED_BLIND_MAPPING_SHA,
    }


def expected_review_ids_for_role(role_id: str) -> set[str]:
    mapping = json.loads(BLIND_MAPPING.read_text(encoding="utf-8"))
    return {r["review_id"] for r in mapping["rows"] if r["role_id"] == role_id}


def discover_returned_workbooks(inbox: Path = INBOX) -> dict[str, Path]:
    """Map role_id → returned workbook path.

    Preference order:
    1. Files in round_a_submissions_inbox/ matching role filenames
    2. Explicit role-named files elsewhere in inbox
    Does NOT treat blank packet templates in reviewers/ as returned submissions.
    """
    found: dict[str, Path] = {}
    inbox.mkdir(parents=True, exist_ok=True)
    for role_id, fname in ROLE_FILES.items():
        p = inbox / fname
        if p.exists():
            found[role_id] = p
            continue
        # allow alternate names: *product_policy*.xlsx etc.
        stub = fname.replace("MokXya_MAI07_V3__", "").replace(".xlsx", "")
        alts = list(inbox.glob(f"*{stub}*.xlsx"))
        if len(alts) == 1:
            found[role_id] = alts[0]
    return found


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def validate_workbook(path: Path, role_id: str) -> dict[str, Any]:
    corrections: list[str] = []
    errors: list[str] = []
    warnings: list[str] = []
    expected_ids = expected_review_ids_for_role(role_id)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {
            "role_id": role_id,
            "path": str(path),
            "ok": False,
            "errors": [f"cannot_open_workbook:{type(exc).__name__}:{exc}"],
            "corrections": ["Re-save as .xlsx without macros and return the intact packet workbook."],
            "declaration_ok": False,
            "rows_expected": len(expected_ids),
            "rows_completed": 0,
            "rows_missing_disposition": len(expected_ids),
            "invalid_enums": [],
            "unknown_ids": [],
            "missing_ids": sorted(expected_ids),
            "duplicate_ids": [],
            "formula_injection_hits": 0,
            "hidden_mapping_exposure": False,
            "product_counted_as_linguist": False,
            "professional_qualification_ok": False,
        }

    # Sheet names
    if tuple(wb.sheetnames) != REQUIRED_SHEETS and set(REQUIRED_SHEETS) - set(wb.sheetnames):
        missing = sorted(set(REQUIRED_SHEETS) - set(wb.sheetnames))
        errors.append(f"missing_sheets:{missing}")
        corrections.append(f"Restore required sheets without renaming: {', '.join(REQUIRED_SHEETS)}")
    if set(wb.sheetnames) - set(REQUIRED_SHEETS):
        extras = sorted(set(wb.sheetnames) - set(REQUIRED_SHEETS))
        # Hidden mapping sheet would be an extra
        if any("mapping" in s.lower() or "blind" in s.lower() for s in extras):
            errors.append(f"hidden_mapping_exposure:{extras}")
            corrections.append("Remove any blind-mapping / import-authority sheets from the returned workbook.")
        else:
            warnings.append(f"extra_sheets:{extras}")

    # Declaration
    decl: dict[str, str] = {}
    if "REVIEWER_DECLARATION" in wb.sheetnames:
        for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            decl[_cell_str(row[0])] = _cell_str(row[1] if len(row) > 1 else "")
    else:
        errors.append("missing_REVIEWER_DECLARATION")
        corrections.append("Complete the REVIEWER_DECLARATION sheet.")

    decl_ok = True
    required_decl = [
        "reviewer_full_name",
        "reviewer_email",
        "declaration_no_ai_autofill",
        "declaration_independent_from_other_reviewers",
        "declaration_no_runtime_predictions_used",
        "declaration_date_utc",
        "signature_or_typed_name",
    ]
    for field in required_decl:
        if not decl.get(field):
            decl_ok = False
            errors.append(f"declaration_missing:{field}")
            corrections.append(f"Fill REVIEWER_DECLARATION field `{field}`.")
    if decl.get("role_id") and decl.get("role_id") != role_id:
        decl_ok = False
        errors.append(f"role_mismatch:declared={decl.get('role_id')}:expected={role_id}")
        corrections.append(f"Keep role_id={role_id}; do not change the assigned role.")
    if not decl.get("role_id"):
        # role_id was prefilled in template; empty means wiped
        decl_ok = False
        errors.append("declaration_missing:role_id")
        corrections.append(f"Restore role_id={role_id} on REVIEWER_DECLARATION.")

    professional_qualification_ok = True
    if role_id == "PROFESSIONAL_LINGUIST_B":
        if not decl.get("qualification_summary") and not decl.get("professional_linguist_credentials"):
            professional_qualification_ok = False
            errors.append("professional_linguist_qualification_missing")
            corrections.append(
                "Professional linguist must fill qualification_summary and/or "
                "professional_linguist_credentials with genuine credentials."
            )
        if not decl.get("professional_linguist_credentials"):
            professional_qualification_ok = False
            errors.append("professional_linguist_credentials_blank")
            corrections.append("Provide professional_linguist_credentials (degree/affiliation/license summary).")

    product_counted_as_linguist = False
    if role_id == "PRODUCT_POLICY":
        # Product workbook must never claim linguist approval fields as authority.
        if decl.get("professional_linguist_credentials") and "linguist" in decl.get(
            "qualification_summary", ""
        ).lower():
            warnings.append("product_policy_has_linguist_text_ignored_for_approval")
        # Explicit gate: product cannot satisfy linguist
        product_counted_as_linguist = False

    # Round A rows
    rows_data: list[dict[str, str]] = []
    header_ok = True
    if "ROUND_A_CONTEXT" not in wb.sheetnames:
        errors.append("missing_ROUND_A_CONTEXT")
        corrections.append("Restore ROUND_A_CONTEXT sheet with original headers.")
        header_ok = False
    else:
        ws = wb["ROUND_A_CONTEXT"]
        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            errors.append("empty_ROUND_A_CONTEXT")
            corrections.append("Do not delete Round A rows.")
            header_ok = False
        else:
            header = tuple(_cell_str(c) for c in rows_iter[0])
            if header != ROUND_A_HEADERS:
                header_ok = False
                errors.append(f"header_mismatch:{header}")
                corrections.append(
                    "Restore exact Round A headers: " + ", ".join(ROUND_A_HEADERS)
                )
            for raw in rows_iter[1:]:
                if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                    continue
                row = {
                    ROUND_A_HEADERS[i]: _cell_str(raw[i] if i < len(raw) else "")
                    for i in range(len(ROUND_A_HEADERS))
                }
                rows_data.append(row)

    wb.close()

    ids = [r["review_id"] for r in rows_data if r["review_id"]]
    id_counts = Counter(ids)
    duplicate_ids = sorted([i for i, n in id_counts.items() if n > 1])
    unknown_ids = sorted(set(ids) - expected_ids)
    missing_ids = sorted(expected_ids - set(ids))
    if duplicate_ids:
        errors.append(f"duplicate_review_ids:{len(duplicate_ids)}")
        corrections.append(f"Remove duplicate review_id rows ({len(duplicate_ids)} ids). Do not invent new IDs.")
    if unknown_ids:
        errors.append(f"unknown_review_ids:{len(unknown_ids)}")
        corrections.append(
            f"Remove {len(unknown_ids)} unknown review_id(s); only use IDs from your assigned workbook."
        )
    if missing_ids:
        errors.append(f"missing_review_ids:{len(missing_ids)}")
        corrections.append(
            f"Restore {len(missing_ids)} missing review_id row(s). Do not delete rows."
        )

    invalid_enums: list[dict[str, str]] = []
    missing_disp = 0
    missing_conf = 0
    formula_hits = 0
    for r in rows_data:
        for field in ("input_text", "highlighted_span", "reviewer_notes", "disposition"):
            val = r.get(field, "")
            if val and CSV_INJECT.match(val.lstrip("'")) and not val.startswith("'"):
                # leading =+@- without escape
                if CSV_INJECT.match(val):
                    formula_hits += 1
        disp = r.get("disposition", "")
        conf = r.get("confidence", "")
        if not disp:
            missing_disp += 1
        elif disp not in ROUND_A_DISPOSITIONS:
            invalid_enums.append({"review_id": r["review_id"], "field": "disposition", "value": disp})
        if disp and not conf:
            missing_conf += 1
        elif conf and conf not in CONFIDENCE:
            invalid_enums.append({"review_id": r["review_id"], "field": "confidence", "value": conf})

    if missing_disp:
        errors.append(f"missing_disposition:{missing_disp}")
        corrections.append(
            f"Complete disposition for all {len(expected_ids)} assigned rows "
            f"({missing_disp} still blank). Allowed enums only."
        )
    if missing_conf:
        errors.append(f"missing_confidence:{missing_conf}")
        corrections.append(f"Set confidence HIGH/MEDIUM/LOW for every labeled row ({missing_conf} missing).")
    if invalid_enums:
        errors.append(f"invalid_enums:{len(invalid_enums)}")
        corrections.append(
            "Replace invalid disposition/confidence values with schema enums only. "
            f"First invalid: {invalid_enums[0]}"
        )
    if formula_hits:
        errors.append(f"formula_injection_hits:{formula_hits}")
        corrections.append("Escape any cell starting with =, +, -, or @ (prefix with ').")

    completed = sum(1 for r in rows_data if r.get("disposition"))
    ok = (
        not errors
        and decl_ok
        and header_ok
        and completed == len(expected_ids)
        and professional_qualification_ok
        and not duplicate_ids
        and not unknown_ids
        and not missing_ids
        and not invalid_enums
        and formula_hits == 0
    )

    # Unique correction list preserving order
    seen: set[str] = set()
    uniq_corr: list[str] = []
    for c in corrections:
        if c not in seen:
            seen.add(c)
            uniq_corr.append(c)

    return {
        "role_id": role_id,
        "path": str(path.relative_to(REPO)).replace("\\", "/") if path.is_relative_to(REPO) else str(path),
        "file_sha256": _sha(path),
        "ok": ok,
        "errors": errors,
        "warnings": warnings,
        "corrections": uniq_corr,
        "declaration_ok": decl_ok,
        "declaration_fields_present": {k: bool(v) for k, v in decl.items()},
        "reviewer_full_name_present": bool(decl.get("reviewer_full_name")),
        "signature_present": bool(decl.get("signature_or_typed_name")),
        "rows_expected": len(expected_ids),
        "rows_present": len(rows_data),
        "rows_completed": completed,
        "rows_missing_disposition": missing_disp,
        "rows_missing_confidence": missing_conf,
        "invalid_enums": invalid_enums[:20],
        "invalid_enum_count": len(invalid_enums),
        "unknown_ids": unknown_ids[:20],
        "unknown_id_count": len(unknown_ids),
        "missing_ids_sample": missing_ids[:20],
        "missing_id_count": len(missing_ids),
        "duplicate_ids": duplicate_ids[:20],
        "duplicate_id_count": len(duplicate_ids),
        "formula_injection_hits": formula_hits,
        "hidden_mapping_exposure": any("hidden_mapping_exposure" in e for e in errors),
        "product_counted_as_linguist": product_counted_as_linguist,
        "professional_qualification_ok": professional_qualification_ok,
        "satisfies_linguist_approval": False,  # never inferred in this phase
        "dispositions": [r["disposition"] for r in rows_data if r.get("disposition")],
        "review_id_to_disposition": {
            r["review_id"]: r["disposition"] for r in rows_data if r.get("disposition")
        },
    }


def detect_identical_response_pattern(reports: list[dict[str, Any]]) -> dict[str, Any]:
    """Flag if two independent reviewers have near-identical full disposition sequences."""
    by_role = {r["role_id"]: r for r in reports if r.get("review_id_to_disposition")}
    pairs = []
    roles = ["NEPALI_FLUENT_A", "PROFESSIONAL_LINGUIST_B", "PRODUCT_POLICY"]
    for i, a in enumerate(roles):
        for b in roles[i + 1 :]:
            if a not in by_role or b not in by_role:
                continue
            da = by_role[a]["review_id_to_disposition"]
            db = by_role[b]["review_id_to_disposition"]
            # Compare on intersection of source items via review_id is role-specific —
            # opaque IDs differ per role, so identical-pattern check uses disposition
            # multiset / sequence length + exact bag equality of disposition lists.
            la = by_role[a].get("dispositions") or []
            lb = by_role[b].get("dispositions") or []
            if not la or not lb or len(la) != len(lb):
                continue
            identical = la == lb and len(la) >= 50
            # Also check if one list is a trivial constant fill
            trivial_a = len(set(la)) == 1 and len(la) >= 50
            trivial_b = len(set(lb)) == 1 and len(lb) >= 50
            pairs.append(
                {
                    "role_a": a,
                    "role_b": b,
                    "identical_sequence": identical,
                    "trivial_constant_a": trivial_a,
                    "trivial_constant_b": trivial_b,
                    "unique_labels_a": len(set(la)),
                    "unique_labels_b": len(set(lb)),
                }
            )
    flagged = [
        p
        for p in pairs
        if p["identical_sequence"] or p["trivial_constant_a"] or p["trivial_constant_b"]
    ]
    return {"pairs": pairs, "flagged": flagged, "ok": not flagged}


def agreement_round_a(reports: list[dict[str, Any]]) -> dict[str, Any]:
    """Round A agreement between Fluent A and Linguist B via source_item join in blind map."""
    mapping = json.loads(BLIND_MAPPING.read_text(encoding="utf-8"))
    by_role_disp = {
        r["role_id"]: r.get("review_id_to_disposition") or {}
        for r in reports
    }
    if "NEPALI_FLUENT_A" not in by_role_disp or "PROFESSIONAL_LINGUIST_B" not in by_role_disp:
        return {"status": "NOT_COMPUTED_MISSING_REVIEWERS", "ok": False}

    # Join on source_item_id
    fluent = {row["source_item_id"]: row["review_id"] for row in mapping["rows"] if row["role_id"] == "NEPALI_FLUENT_A"}
    ling = {row["source_item_id"]: row["review_id"] for row in mapping["rows"] if row["role_id"] == "PROFESSIONAL_LINGUIST_B"}
    labels_a: list[str] = []
    labels_b: list[str] = []
    for sid in sorted(set(fluent) & set(ling)):
        da = by_role_disp["NEPALI_FLUENT_A"].get(fluent[sid])
        db = by_role_disp["PROFESSIONAL_LINGUIST_B"].get(ling[sid])
        if da and db:
            labels_a.append(da)
            labels_b.append(db)
    if not labels_a:
        return {
            "status": "NOT_COMPUTED_NO_OVERLAPPING_LABELS",
            "paired_n": 0,
            "ok": False,
        }
    exact = exact_agreement(labels_a, labels_b)
    kappa = cohen_kappa(labels_a, labels_b)
    return {
        "status": "COMPUTED",
        "paired_n": len(labels_a),
        "exact_disposition_agreement": exact,
        "cohen_kappa": kappa,
        "ok": True,
        "note": "Diagnostic only; adjudication not performed in this phase.",
    }


def run_round_a_validation(*, inbox: Path = INBOX) -> dict[str, Any]:
    VALIDATION_DIR.mkdir(parents=True, exist_ok=True)
    sealed = verify_sealed_authorities()
    if not sealed["ok"]:
        result = {
            "status": "BLOCKED_PACKET_INTEGRITY_FAILED",
            "sealed": sealed,
            "ROUND_A_LOCKED": False,
            "ROUND_B_READY": False,
            "human_answers_altered": False,
            "model_evaluation_performed": False,
        }
        _write_json(VALIDATION_DIR / "MAI_07_V3_ROUND_A_VALIDATION_SUMMARY.json", result)
        return result

    found = discover_returned_workbooks(inbox)
    reports: list[dict[str, Any]] = []
    for role_id in REQUIRED_ROUND_A_ROLES:
        if role_id not in found:
            reports.append(
                {
                    "role_id": role_id,
                    "path": None,
                    "ok": False,
                    "errors": ["NO_RETURNED_WORKBOOK_IN_INBOX"],
                    "corrections": [
                        f"Place completed workbook `{ROLE_FILES[role_id]}` into "
                        f"`docs/mokxya-ai/reviews/mai07_v3/round_a_submissions_inbox/`.",
                        "Do not submit the blank packet template from reviewers/ without completing Round A.",
                        "Complete REVIEWER_DECLARATION (signed) and all Round A disposition+confidence cells.",
                    ],
                    "declaration_ok": False,
                    "rows_expected": len(expected_review_ids_for_role(role_id)),
                    "rows_completed": 0,
                    "rows_missing_disposition": len(expected_review_ids_for_role(role_id)),
                    "invalid_enums": [],
                    "unknown_ids": [],
                    "missing_ids_sample": [],
                    "duplicate_ids": [],
                    "formula_injection_hits": 0,
                    "hidden_mapping_exposure": False,
                    "product_counted_as_linguist": False,
                    "professional_qualification_ok": False,
                    "satisfies_linguist_approval": False,
                }
            )
            continue
        reports.append(validate_workbook(found[role_id], role_id))

    for rep in reports:
        _write_json(
            VALIDATION_DIR / f"MAI_07_V3_ROUND_A_VALIDATION__{rep['role_id']}.json",
            {k: v for k, v in rep.items() if k not in {"dispositions", "review_id_to_disposition"}},
        )

    pattern = detect_identical_response_pattern(reports)
    if not pattern["ok"]:
        for rep in reports:
            if rep["role_id"] in {"NEPALI_FLUENT_A", "PROFESSIONAL_LINGUIST_B", "PRODUCT_POLICY"}:
                rep["ok"] = False
                rep.setdefault("errors", []).append("identical_or_trivial_response_pattern")
                rep.setdefault("corrections", []).append(
                    "Independent reviewers must not copy answers. Re-complete Round A independently."
                )

    all_ok = all(r.get("ok") for r in reports) and sealed["ok"] and pattern["ok"]

    if not all_ok:
        try:
            inbox_rel = str(inbox.relative_to(REPO)).replace("\\", "/")
        except ValueError:
            inbox_rel = str(inbox)
        summary = {
            "status": "BLOCKED_ROUND_A_CORRECTION_REQUIRED",
            "phase": "MAI-07R3J-A-ROUND-A-LOCK-AND-VALIDATION",
            "sealed": sealed,
            "inbox": inbox_rel,
            "files_received": {
                r["role_id"]: r.get("path") for r in reports
            },
            "reviewer_reports": [
                {
                    "role_id": r["role_id"],
                    "ok": r.get("ok"),
                    "declaration_ok": r.get("declaration_ok"),
                    "rows_expected": r.get("rows_expected"),
                    "rows_completed": r.get("rows_completed"),
                    "rows_missing_disposition": r.get("rows_missing_disposition"),
                    "error_count": len(r.get("errors") or []),
                    "errors": r.get("errors"),
                    "corrections": r.get("corrections"),
                    "professional_qualification_ok": r.get("professional_qualification_ok"),
                    "product_counted_as_linguist": False,
                    "satisfies_linguist_approval": False,
                }
                for r in reports
            ],
            "identical_response_pattern": pattern,
            "ROUND_A_LOCKED": False,
            "ROUND_B_READY": False,
            "round_b_released": False,
            "agreement": {"status": "NOT_COMPUTED_ROUND_A_NOT_LOCKED"},
            "QUALITY_GATES_PASSED": False,
            "LINGUIST_APPROVED": False,
            "PRODUCTION_APPROVED": False,
            "MAI_08": "NOT_STARTED",
            "v2_governance": "HISTORICAL_BENCHMARK_EXHAUSTED_FOR_MODEL_SELECTION",
            "human_answers_altered": False,
            "human_answers_generated": False,
            "model_evaluation_performed": False,
            "runtime_modified": False,
            "next_human_action": (
                "Return corrected Round A workbooks to "
                "docs/mokxya-ai/reviews/mai07_v3/round_a_submissions_inbox/ "
                "then re-run MAI-07R3J-A-ROUND-A-LOCK-AND-VALIDATION. "
                "Do not begin Round B."
            ),
            "created_utc": datetime.now(timezone.utc).isoformat(),
        }
        _write_json(VALIDATION_DIR / "MAI_07_V3_ROUND_A_VALIDATION_SUMMARY.json", summary)
        # Inbox README for humans
        (INBOX / "README_PLACE_RETURNED_WORKBOOKS_HERE.md").write_text(
            "# Round A submissions inbox\n\n"
            "Place completed role workbooks here (exact names preferred):\n\n"
            + "\n".join(f"- `{v}`" for v in ROLE_FILES.values())
            + "\n\nBlank templates in `../reviewers/` are **not** accepted as submissions.\n",
            encoding="utf-8",
            newline="\n",
        )
        return summary

    # --- Success path: lock + Round B release ---
    LOCKED_DIR.mkdir(parents=True, exist_ok=True)
    ROUND_B_RELEASE_DIR.mkdir(parents=True, exist_ok=True)
    locked_files: dict[str, Any] = {}
    for role_id, src in found.items():
        dest = LOCKED_DIR / ROLE_FILES[role_id]
        if dest.exists():
            raise FileExistsError(f"already_locked:{dest}")
        shutil.copy2(src, dest)
        raw = _sha(dest)
        # Canonical: role + sorted (review_id, disposition, confidence) from validation extract
        rep = next(r for r in reports if r["role_id"] == role_id)
        # Re-read for canonical export without altering source
        canon_rows = []
        wb = load_workbook(dest, read_only=True, data_only=True)
        for row in wb["ROUND_A_CONTEXT"].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            canon_rows.append(
                {
                    "review_id": _cell_str(row[0]),
                    "disposition": _cell_str(row[3]),
                    "confidence": _cell_str(row[4]),
                    "reason_category": _cell_str(row[5]) if len(row) > 5 else "",
                }
            )
        wb.close()
        canon_rows = sorted(canon_rows, key=lambda r: r["review_id"])
        canon = _sha_bytes(
            json.dumps(canon_rows, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
                "utf-8"
            )
        )
        # Prove original unchanged
        orig_sha = _sha(src)
        locked_files[role_id] = {
            "locked_path": str(dest.relative_to(REPO)).replace("\\", "/"),
            "source_path": str(src.relative_to(REPO)).replace("\\", "/") if src.is_relative_to(REPO) else str(src),
            "raw_sha256": raw,
            "canonical_sha256": canon,
            "source_sha256": orig_sha,
            "source_matches_locked": orig_sha == raw,
        }
        export_path = LOCKED_DIR / f"MAI_07_V3_ROUND_A_RESPONSES__{role_id}.json"
        _write_json(export_path, {"role_id": role_id, "rows": canon_rows})

    agreement = agreement_round_a(reports)

    # Round B packages from sealed templates (already shuffled per role)
    round_b_files: dict[str, str] = {}
    for role_id in ("NEPALI_FLUENT_A", "PROFESSIONAL_LINGUIST_B", "ACCOUNTING_DOMAIN"):
        stub = ROLE_FILES[role_id].replace("MokXya_MAI07_V3__", "").replace(".xlsx", "")
        src_xlsx = REVIEWERS_DIR / ROLE_FILES[role_id]
        # Copy sealed Round B CSV template + workbook note; do not include other reviewers' Round A
        csv_src = REVIEWERS_DIR / f"V3_ROUND_B_TEMPLATE__{stub}.csv"
        csv_dest = ROUND_B_RELEASE_DIR / f"MokXya_MAI07_V3_ROUND_B__{stub}.csv"
        xlsx_dest = ROUND_B_RELEASE_DIR / f"MokXya_MAI07_V3_ROUND_B_PACKAGE__{stub}.xlsx"
        shutil.copy2(csv_src, csv_dest)
        # Build Round-B-focused workbook from sealed reviewer workbook (contains shuffled B sheet)
        shutil.copy2(src_xlsx, xlsx_dest)
        round_b_files[role_id] = {
            "csv": str(csv_dest.relative_to(REPO)).replace("\\", "/"),
            "xlsx_package": str(xlsx_dest.relative_to(REPO)).replace("\\", "/"),
            "csv_sha256": _sha(csv_dest),
            "xlsx_sha256": _sha(xlsx_dest),
            "note": "Round B only; do not overwrite Round A; candidate order is sealed per reviewer.",
        }

    lock_manifest = {
        "schema_version": "1.0.0",
        "phase": "MAI-07R3J-A-ROUND-A-LOCK-AND-VALIDATION",
        "status": "ROUND_B_READY",
        "ROUND_A_LOCKED": True,
        "ROUND_B_READY": True,
        "sealed": sealed,
        "locked_files": locked_files,
        "agreement_round_a": agreement,
        "identical_response_pattern": pattern,
        "QUALITY_GATES_PASSED": False,
        "LINGUIST_APPROVED": False,
        "PRODUCTION_APPROVED": False,
        "MAI_08": "NOT_STARTED",
        "v2_governance": "HISTORICAL_BENCHMARK_EXHAUSTED_FOR_MODEL_SELECTION",
        "adjudication_performed": False,
        "human_answers_altered": False,
        "model_evaluation_performed": False,
        "round_b_files": round_b_files,
        "packet_authority_utc": "2026-07-16T00:00:00+00:00",
    }
    # Strip fluctuating clock from sealed lock body
    lock_path = OUT / "MAI_07_V3_ROUND_A_LOCK_MANIFEST.json"
    lock_sha = _write_json(lock_path, lock_manifest)
    lock_manifest["lock_manifest_sha256"] = lock_sha
    # rewrite with hash included deterministically
    lock_sha = _write_json(lock_path, lock_manifest)

    summary = {
        "status": "ROUND_B_READY",
        "ROUND_A_LOCKED": True,
        "ROUND_B_READY": True,
        "lock_manifest_path": str(lock_path.relative_to(REPO)).replace("\\", "/"),
        "lock_manifest_sha256": lock_sha,
        "locked_files": locked_files,
        "agreement": agreement,
        "round_b_files": round_b_files,
        "sealed": sealed,
        "QUALITY_GATES_PASSED": False,
        "LINGUIST_APPROVED": False,
        "human_answers_altered": False,
        "model_evaluation_performed": False,
        "next_human_action": (
            "Send each reviewer only their Round B package under "
            "docs/mokxya-ai/reviews/mai07_v3/round_b_release/. "
            "Complete Round B independently. Do not share Round A answers."
        ),
    }
    _write_json(VALIDATION_DIR / "MAI_07_V3_ROUND_A_VALIDATION_SUMMARY.json", summary)
    return summary


def main() -> int:
    import argparse

    p = argparse.ArgumentParser(description="MAI-07 V3 Round A lock and validation")
    p.add_argument("--inbox", type=Path, default=INBOX)
    args = p.parse_args()
    result = run_round_a_validation(inbox=args.inbox)
    print(
        json.dumps(
            {
                "status": result.get("status"),
                "ROUND_A_LOCKED": result.get("ROUND_A_LOCKED"),
                "ROUND_B_READY": result.get("ROUND_B_READY"),
                "files_received": result.get("files_received"),
                "lock_manifest_sha256": result.get("lock_manifest_sha256"),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0 if result.get("status") == "ROUND_B_READY" else 2


if __name__ == "__main__":
    raise SystemExit(main())
