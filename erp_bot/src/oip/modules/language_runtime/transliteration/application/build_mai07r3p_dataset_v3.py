"""MAI-07R3P — consume sealed V3 human-review freeze into frozen dataset V3.

Deterministic. Does not run model evaluation, promote runtime, or start MAI-08.
Gold is derived only from freeze-pinned Round A/B artifacts + blind mapping.
"""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[7]
V3_ROOT = REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3"
OPS = V3_ROOT / "review_operations"
FREEZE_MANIFEST = V3_ROOT / "MAI_07_V3_HUMAN_REVIEW_FREEZE_MANIFEST.json"
BLIND_MAPPING = V3_ROOT / "V3_BLIND_MAPPING.json"
ROUND_A_XLSX = {
    "NEPALI_FLUENT_A": V3_ROOT
    / "round_a_locked_submissions"
    / "MokXya_MAI07_V3__nepali_fluent_a.xlsx",
    "PROFESSIONAL_LINGUIST_B": V3_ROOT
    / "round_a_locked_submissions"
    / "MokXya_MAI07_V3__professional_linguist_b.xlsx",
    "ACCOUNTING_DOMAIN": V3_ROOT
    / "round_a_locked_submissions"
    / "MokXya_MAI07_V3__accounting_domain.xlsx",
    "PRODUCT_POLICY": V3_ROOT
    / "round_a_locked_submissions"
    / "MokXya_MAI07_V3__product_policy.xlsx",
}
ROUND_B_RESP = {
    "NEPALI_FLUENT_A": OPS
    / "round_b_locked"
    / "MAI_07_V3_ROUND_B_RESPONSES__NEPALI_FLUENT_A.json",
    "PROFESSIONAL_LINGUIST_B": OPS
    / "round_b_locked"
    / "MAI_07_V3_ROUND_B_RESPONSES__PROFESSIONAL_LINGUIST_B.json",
    "ACCOUNTING_DOMAIN": OPS
    / "round_b_locked"
    / "MAI_07_V3_ROUND_B_RESPONSES__ACCOUNTING_DOMAIN.json",
}

OUT_DIR = REPO / "evals" / "mai07"
FROZEN_V3_DIR = OUT_DIR / "frozen_v3"
MANIFESTS_DIR = OUT_DIR / "manifests"

DATASET_ID = "MAI_07_ROMANIZED_TRANSLITERATION_V3"
SCHEMA_VERSION = "mai07r3p_dataset_v3_1.0.0"
CANONICAL_SCORER_VERSION = "mai-07.r3p.canonical.1.0.0"
AUDIT_SCORER_VERSION = "mai-07.r3p.audit.1.0.0"

EXPECTED_FREEZE_SHA = "adcee2904dc34df8e305d0de9e0a12f61c20b130da86d20ec1f225678587397c"

# Locked before any V3 runtime observation (not fitted to candidate scores).
THRESHOLD_MANIFEST: dict[str, Any] = {
    "schema_version": "1.0.0",
    "threshold_id": "MAI_07_R3P_THRESHOLDS_V3",
    "locked_before_runtime_observation": True,
    "scorer_versions": {
        "canonical": CANONICAL_SCORER_VERSION,
        "audit": AUDIT_SCORER_VERSION,
    },
    "gates": {
        "target_candidate_top1_accuracy": {"op": ">=", "value": 0.88},
        "target_candidate_recall_at_5": {"op": ">=", "value": 0.95},
        "target_candidate_mrr": {"op": ">=", "value": 0.90},
        "core_target_recall_at_5": {"op": ">=", "value": 0.98},
        "unambiguous_target_top1": {"op": ">=", "value": 0.92},
        "english_identity_top1": {"op": ">=", "value": 0.98},
        "false_devanagari_on_english": {"op": "<=", "value": 0.02},
        "protected_mutations": {"op": "==", "value": 0},
        "raw_view_mutations": {"op": "==", "value": 0},
        "deterministic_output_rate": {"op": "==", "value": 1.0},
        "candidate_caps_respected": {"op": "==", "value": 1.0},
        "unresolved_review_accuracy": {"op": ">=", "value": 0.90},
    },
    "note": "Copied from pre-declared V3 metric intent / R3C gate shape; not fitted to candidate scores.",
}

_DEVANAGARI = re.compile(r"[\u0900-\u097F]")


def _sha_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sha_file(path: Path) -> str:
    return _sha_bytes(path.read_bytes())


def _canonical_json(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, indent=2, ensure_ascii=False) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return _sha_bytes(text.encode("utf-8"))


def _contains_devanagari(text: str) -> bool:
    return bool(_DEVANAGARI.search(text or ""))


def _load_round_a_rows(role_id: str) -> dict[str, dict[str, str]]:
    path = ROUND_A_XLSX[role_id]
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dict[str, str]] = {}
    for raw in wb["ROUND_A_CONTEXT"].iter_rows(min_row=2, values_only=True):
        if not raw or raw[0] is None:
            continue
        rid = str(raw[0]).strip()
        out[rid] = {
            "review_id": rid,
            "input_text": "" if raw[1] is None else str(raw[1]),
            "highlighted_span": "" if raw[2] is None else str(raw[2]),
            "disposition": "" if raw[3] is None else str(raw[3]).strip(),
            "confidence": "" if len(raw) < 5 or raw[4] is None else str(raw[4]).strip(),
        }
    wb.close()
    return out


def _load_round_b(role_id: str) -> dict[str, list[dict[str, str]]]:
    path = ROUND_B_RESP[role_id]
    data = json.loads(path.read_text(encoding="utf-8"))
    by_rid: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in data["rows"]:
        by_rid[row["review_id"]].append(
            {
                "candidate_index": str(row["candidate_index"]),
                "candidate_surface": row["candidate_surface"],
                "acceptability": row["acceptability"],
            }
        )
    for rid in by_rid:
        by_rid[rid].sort(
            key=lambda r: int(r["candidate_index"])
            if str(r["candidate_index"]).isdigit()
            else r["candidate_index"]
        )
    return by_rid


def _population_from_disposition(
    disposition: str,
    *,
    preferred_surfaces: list[str],
) -> str:
    if disposition == "DEVANAGARI_TRANSLITERATION_REQUIRED":
        return "TRANSLITERATION_REQUIRED"
    if disposition == "ENGLISH_IDENTITY_REQUIRED":
        return "IDENTITY_REQUIRED"
    if disposition == "TRANSLITERATION_OPTIONAL":
        return "TRANSLITERATION_OPTIONAL"
    if disposition == "NO_TRANSLITERATION_ALLOWED":
        return "NO_TRANSLITERATION_ALLOWED"
    if disposition == "IDENTITY_FIRST_REVIEW_REQUIRED":
        return "IDENTITY_FIRST_REVIEW_REQUIRED"
    if disposition in {"PROTECTED", "NAME_OR_ENTITY", "ACRONYM_OR_IDENTIFIER"}:
        return "PROTECTED_IDENTITY"
    if disposition == "ABSTAIN_CANNOT_DECIDE":
        return "HUMAN_REVIEW_REQUIRED"
    if disposition == "CONTEXT_DEPENDENT":
        if any(_contains_devanagari(s) for s in preferred_surfaces):
            return "CONTEXT_DEPENDENT_NEPALI"
        return "CONTEXT_DEPENDENT_ENGLISH"
    return "HUMAN_REVIEW_REQUIRED"


def _suite_for_population(population: str, pool: str) -> str:
    base = {
        "TRANSLITERATION_REQUIRED": "v3_transliteration_required",
        "IDENTITY_REQUIRED": "v3_identity_required",
        "TRANSLITERATION_OPTIONAL": "v3_transliteration_optional",
        "NO_TRANSLITERATION_ALLOWED": "v3_no_transliteration_allowed",
        "IDENTITY_FIRST_REVIEW_REQUIRED": "v3_identity_first_review",
        "PROTECTED_IDENTITY": "v3_protected_identity",
        "HUMAN_REVIEW_REQUIRED": "v3_human_review_required",
        "CONTEXT_DEPENDENT_ENGLISH": "v3_context_dependent_english",
        "CONTEXT_DEPENDENT_NEPALI": "v3_context_dependent_nepali",
        "INFORMATIONAL_EXCLUDED": "v3_informational_excluded",
    }.get(population, "v3_other")
    return f"{base}__{pool.lower()}"


def build_v3_cases(repo: Path = REPO) -> dict[str, Any]:
    freeze_sha = _sha_file(FREEZE_MANIFEST)
    if freeze_sha != EXPECTED_FREEZE_SHA:
        raise RuntimeError(
            f"freeze_manifest_sha_mismatch:got={freeze_sha}:expected={EXPECTED_FREEZE_SHA}"
        )
    freeze = json.loads(FREEZE_MANIFEST.read_text(encoding="utf-8"))
    if freeze.get("status") != "V3_HUMAN_REVIEW_FREEZE_SEALED":
        raise RuntimeError(f"freeze_not_sealed:{freeze.get('status')}")

    mapping = json.loads(BLIND_MAPPING.read_text(encoding="utf-8"))
    round_a = {role: _load_round_a_rows(role) for role in ROUND_A_XLSX}
    round_b = {role: _load_round_b(role) for role in ROUND_B_RESP}

    # Authority labels: Fluent A primary; Linguist B must match disposition (already kappa=1).
    fluent_map = {
        row["source_item_id"]: row
        for row in mapping["rows"]
        if row["role_id"] == "NEPALI_FLUENT_A"
    }
    ling_map = {
        row["source_item_id"]: row
        for row in mapping["rows"]
        if row["role_id"] == "PROFESSIONAL_LINGUIST_B"
    }
    acct_map = {
        row["source_item_id"]: row
        for row in mapping["rows"]
        if row["role_id"] == "ACCOUNTING_DOMAIN"
    }

    cases: list[dict[str, Any]] = []
    pop_counts: Counter[str] = Counter()
    pool_counts: Counter[str] = Counter()
    rb_counts: Counter[str] = Counter()
    disposition_mismatch = 0

    for sid in sorted(fluent_map):
        frow = fluent_map[sid]
        lrow = ling_map[sid]
        rid_a = frow["review_id"]
        rid_b = lrow["review_id"]
        a = round_a["NEPALI_FLUENT_A"][rid_a]
        b = round_a["PROFESSIONAL_LINGUIST_B"][rid_b]
        if a["disposition"] != b["disposition"]:
            disposition_mismatch += 1
            # Prefer linguist on mismatch (should be zero under sealed lock).
            disposition = b["disposition"]
        else:
            disposition = a["disposition"]

        cands_a = round_b["NEPALI_FLUENT_A"].get(rid_a, [])
        for c in cands_a:
            rb_counts[c["acceptability"]] += 1

        preferred = [
            c["candidate_surface"]
            for c in cands_a
            if c["acceptability"] == "ACCEPTABLE_PREFERRED"
            and c["candidate_surface"] != "NONE_ACCEPTABLE"
        ]
        alternative = [
            c["candidate_surface"]
            for c in cands_a
            if c["acceptability"] == "ACCEPTABLE_ALTERNATIVE"
            and c["candidate_surface"] != "NONE_ACCEPTABLE"
        ]
        # Devanagari targets only for transliteration scoring credit.
        preferred_dev = [s for s in preferred if _contains_devanagari(s)]
        alternative_dev = [s for s in alternative if _contains_devanagari(s)]

        population = _population_from_disposition(
            disposition, preferred_surfaces=preferred
        )
        pool = frow["future_pool_assignment"]
        suite = _suite_for_population(population, pool)
        pop_counts[population] += 1
        pool_counts[pool] += 1

        case = {
            "case_id": sid,
            "source_item_id": sid,
            "family_id": frow["family_id"],
            "suite_id": suite,
            "pool": pool,
            "input_text": a["input_text"],
            "highlighted_span": a["highlighted_span"],
            "primary_population": population,
            "round_a_disposition": disposition,
            "round_a_confidence": a["confidence"],
            "review_ids": {
                "NEPALI_FLUENT_A": rid_a,
                "PROFESSIONAL_LINGUIST_B": rid_b,
                "ACCOUNTING_DOMAIN": acct_map.get(sid, {}).get("review_id"),
                "PRODUCT_POLICY": next(
                    (
                        r["review_id"]
                        for r in mapping["rows"]
                        if r["source_item_id"] == sid and r["role_id"] == "PRODUCT_POLICY"
                    ),
                    None,
                ),
            },
            "candidate_order_fluent_a": list(frow.get("candidate_order") or []),
            "acceptable_preferred": preferred,
            "acceptable_alternative": alternative,
            "acceptable_devanagari_targets": preferred_dev + [
                s for s in alternative_dev if s not in preferred_dev
            ],
            "preferred_devanagari_targets": preferred_dev,
            "round_b_labels_fluent_a": cands_a,
            "provenance_class": frow.get("provenance_class"),
            "prohibited_for_training": True,
            "option_a_mechanical_remap": True,
            "identity_expected": population
            in {"IDENTITY_REQUIRED", "PROTECTED_IDENTITY", "NO_TRANSLITERATION_ALLOWED"},
            "abstention_expected": population == "HUMAN_REVIEW_REQUIRED",
            "requires_domain_review": sid in acct_map,
        }
        cases.append(case)

    if disposition_mismatch:
        # Soft warning recorded in reconciliation; builder still deterministic.
        pass

    reconciliation = {
        "total_cases": len(cases),
        "unique_source_items": len(cases),
        "disposition_mismatch_fluent_vs_linguist": disposition_mismatch,
        "population_counts": dict(sorted(pop_counts.items())),
        "pool_counts": dict(sorted(pool_counts.items())),
        "round_b_label_counts_fluent_a": dict(sorted(rb_counts.items())),
        "freeze_manifest_sha256": freeze_sha,
        "option_a_mechanical_remap": True,
        "parent_freeze_status": freeze.get("status"),
        "blind_mapping_sha256": _sha_file(BLIND_MAPPING),
    }
    return {"cases": cases, "reconciliation": reconciliation, "freeze_sha": freeze_sha}


def write_v3_dataset(repo: Path = REPO) -> dict[str, Any]:
    built = build_v3_cases(repo)
    cases: list[dict[str, Any]] = built["cases"]
    FROZEN_V3_DIR.mkdir(parents=True, exist_ok=True)
    MANIFESTS_DIR.mkdir(parents=True, exist_ok=True)

    # Clear previous suite files in frozen_v3 only
    for old in FROZEN_V3_DIR.glob("*.jsonl"):
        old.unlink()

    by_suite: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for case in cases:
        by_suite[case["suite_id"]].append(case)

    files_meta: list[dict[str, Any]] = []
    for suite_id in sorted(by_suite):
        path = FROZEN_V3_DIR / f"{suite_id}.jsonl"
        lines = [
            _canonical_json(c) for c in sorted(by_suite[suite_id], key=lambda x: x["case_id"])
        ]
        body = ("\n".join(lines) + "\n").encode("utf-8")
        path.write_bytes(body)
        files_meta.append(
            {
                "suite_id": suite_id,
                "path": str(path.relative_to(REPO)).replace("\\", "/"),
                "case_count": len(lines),
                "sha256": _sha_bytes(body),
            }
        )

    # Dataset hash over sorted suite files
    h = hashlib.sha256()
    for f in files_meta:
        h.update(f["suite_id"].encode())
        h.update(b"\0")
        h.update((REPO / f["path"]).read_bytes())
    dataset_hash = h.hexdigest()

    pop_manifest = {
        "schema_version": "1.0.0",
        "dataset_id": DATASET_ID,
        "populations": built["reconciliation"]["population_counts"],
        "pools": built["reconciliation"]["pool_counts"],
        "taxonomy_source": "docs/mokxya-ai/reviews/mai07_v3/V3_POPULATION_TAXONOMY.md",
    }
    pop_sha = _write_json(MANIFESTS_DIR / "MAI_07_R3P_POPULATIONS_V3.manifest.json", pop_manifest)
    thr_sha = _write_json(MANIFESTS_DIR / "MAI_07_R3P_THRESHOLDS_V3.manifest.json", THRESHOLD_MANIFEST)

    manifest = {
        "schema_version": "1.0.0",
        "dataset_id": DATASET_ID,
        "dataset_hash": dataset_hash,
        "builder": "build_mai07r3p_dataset_v3.py",
        "builder_schema_version": SCHEMA_VERSION,
        "canonical_scorer_version": CANONICAL_SCORER_VERSION,
        "audit_scorer_version": AUDIT_SCORER_VERSION,
        "parent_freeze_manifest": "docs/mokxya-ai/reviews/mai07_v3/MAI_07_V3_HUMAN_REVIEW_FREEZE_MANIFEST.json",
        "parent_freeze_sha256": built["freeze_sha"],
        "authority": "ADR_0022",
        "prohibited_for_training": True,
        "option_a_mechanical_remap": True,
        "total_cases": len(cases),
        "files": files_meta,
        "population_manifest_sha256": pop_sha,
        "threshold_manifest_sha256": thr_sha,
        "reconciliation": built["reconciliation"],
        "QUALITY_GATES_PASSED": False,
        "PRODUCTION_APPROVED": False,
        "runtime_evaluation_performed": False,
        "note": (
            "V3 dataset frozen from human-review freeze. "
            "No model predictions observed. Ready for separately authorized one-shot eval."
        ),
    }
    man_sha = _write_json(
        MANIFESTS_DIR / "MAI_07_ROMANIZED_TRANSLITERATION_V3.manifest.json", manifest
    )
    release = {
        "schema_version": "1.0.0",
        "status": "V3_DATASET_FROZEN_AWAITING_AUTHORIZED_EVAL",
        "dataset_id": DATASET_ID,
        "dataset_hash": dataset_hash,
        "dataset_manifest_sha256": man_sha,
        "threshold_manifest_sha256": thr_sha,
        "population_manifest_sha256": pop_sha,
        "QUALITY_GATES_PASSED": False,
        "PRODUCTION_APPROVED": False,
        "LINGUIST_APPROVED": True,
        "candidate_promoted": False,
        "mai_08": "NOT_STARTED",
        "active_runtime_unchanged": "mai-07.1.3-r3f-sealnew",
    }
    # Rebuild of the same frozen hash must not erase a later authorized one-shot outcome.
    rel_path = MANIFESTS_DIR / "MAI_07_R3P_V3_RELEASE_CANDIDATE.manifest.json"
    if rel_path.exists():
        prior = json.loads(rel_path.read_text(encoding="utf-8"))
        if (
            prior.get("dataset_hash") == dataset_hash
            and prior.get("last_attempt_id")
            and prior.get("status")
            not in {None, "V3_DATASET_FROZEN_AWAITING_AUTHORIZED_EVAL"}
        ):
            for key in (
                "status",
                "QUALITY_GATES_PASSED",
                "PRODUCTION_APPROVED",
                "LINGUIST_APPROVED",
                "CUTOVER_AUTHORIZED",
                "candidate_promoted",
                "mai_08",
                "last_attempt_id",
                "last_verdict",
                "last_candidate_runtime",
                "parent_failed_attempt_id",
                "correction_scope",
                "qualified_pack",
                "qualified_pack_content_hash",
                "authority",
                "note",
            ):
                if key in prior:
                    release[key] = prior[key]
            # Rebuild may preserve a completed cutover.
            if prior.get("candidate_promoted") is True:
                release["candidate_promoted"] = True
            elif prior.get("candidate_promoted") is not True:
                release["candidate_promoted"] = False
            if prior.get("active_runtime_version"):
                release["active_runtime_version"] = prior["active_runtime_version"]
                release["active_pack_version"] = prior.get(
                    "active_pack_version", release.get("active_pack_version")
                )
                release.pop("active_runtime_unchanged", None)
            else:
                release["active_runtime_unchanged"] = prior.get(
                    "active_runtime_unchanged", "mai-07.1.13-r3s-active"
                )
    _write_json(rel_path, release)
    return {
        "ok": True,
        "dataset_hash": dataset_hash,
        "dataset_manifest_sha256": man_sha,
        "threshold_manifest_sha256": thr_sha,
        "population_manifest_sha256": pop_sha,
        "total_cases": len(cases),
        "reconciliation": built["reconciliation"],
        "status": release["status"],
    }


def build_twice_and_verify(repo: Path = REPO) -> dict[str, Any]:
    a = write_v3_dataset(repo)
    b = write_v3_dataset(repo)
    if a["dataset_hash"] != b["dataset_hash"]:
        raise RuntimeError("v3_dataset_nondeterministic")
    return a


def main() -> int:
    out = build_twice_and_verify()
    print(json.dumps(out, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
