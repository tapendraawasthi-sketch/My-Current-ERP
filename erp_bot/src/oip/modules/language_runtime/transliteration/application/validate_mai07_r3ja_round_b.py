"""MAI-07R3J-A — Round B lock-and-validation (governance only).

Validates returned Round B workbooks against sealed Round B release templates.
Does not invent/correct human decisions, modify blind mapping, run model
evaluation, or set LINGUIST_APPROVED / QUALITY_GATES_PASSED.
"""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .mai07_r3ja_v3_agreement import (
    ACCEPTANCE_GATES,
    ROUND_B_ACCEPTABILITY,
    cohen_kappa,
    exact_agreement,
)
from .mai07_r3ja_v3_firewall import REPO
from .validate_mai07_r3ja_round_a import (
    EXPECTED_BLIND_MAPPING_SHA,
    EXPECTED_PACKET_MANIFEST_SHA,
    _sha,
    _write_json,
    verify_sealed_authorities,
)

OUT = REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3"
OPS = OUT / "review_operations"
BLIND_MAPPING = OUT / "V3_BLIND_MAPPING.json"
ROUND_B_RELEASE_DIR = OUT / "round_b_release"
ROUND_B_INBOX = OPS / "round_b_inbox"
ROUND_B_LOCKED = OPS / "round_b_locked"
ROUND_B_VALIDATION = OPS / "validation_reports"
ADJUDICATION_DIR = OPS / "adjudication_package"

ROUND_B_ROLES = (
    "NEPALI_FLUENT_A",
    "PROFESSIONAL_LINGUIST_B",
    "ACCOUNTING_DOMAIN",
)

ROUND_B_PACKAGES = {
    "NEPALI_FLUENT_A": "MokXya_MAI07_V3_ROUND_B_PACKAGE__nepali_fluent_a.xlsx",
    "PROFESSIONAL_LINGUIST_B": "MokXya_MAI07_V3_ROUND_B_PACKAGE__professional_linguist_b.xlsx",
    "ACCOUNTING_DOMAIN": "MokXya_MAI07_V3_ROUND_B_PACKAGE__accounting_domain.xlsx",
}

ROUND_B_CSVS = {
    "NEPALI_FLUENT_A": "MokXya_MAI07_V3_ROUND_B__nepali_fluent_a.csv",
    "PROFESSIONAL_LINGUIST_B": "MokXya_MAI07_V3_ROUND_B__professional_linguist_b.csv",
    "ACCOUNTING_DOMAIN": "MokXya_MAI07_V3_ROUND_B__accounting_domain.csv",
}

OFFICIAL_B = frozenset(ROUND_B_ACCEPTABILITY)


def _expected_pairs(role_id: str) -> list[dict[str, str]]:
    path = ROUND_B_RELEASE_DIR / ROUND_B_CSVS[role_id]
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def _load_answers_from_xlsx(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "ROUND_B_CANDIDATES" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"missing_ROUND_B_CANDIDATES:{path.name}")
    out: dict[tuple[str, str], dict[str, str]] = {}
    for raw in wb["ROUND_B_CANDIDATES"].iter_rows(min_row=2, values_only=True):
        if not raw or raw[0] is None:
            continue
        rid = str(raw[0]).strip()
        if rid.startswith("(Round B"):
            continue
        ci = "" if len(raw) < 4 or raw[3] is None else str(raw[3]).strip()
        surface = "" if len(raw) < 5 or raw[4] is None else str(raw[4]).strip()
        acc = "" if len(raw) < 6 or raw[5] is None else str(raw[5]).strip()
        entered = "" if len(raw) < 7 or raw[6] is None else str(raw[6]).strip()
        notes = "" if len(raw) < 8 or raw[7] is None else str(raw[7]).strip()
        out[(rid, ci)] = {
            "review_id": rid,
            "candidate_index": ci,
            "candidate_surface": surface,
            "acceptability": acc,
            "reviewer_entered_candidate": entered,
            "notes": notes,
        }
    wb.close()
    return out


def validate_role_round_b(role_id: str) -> dict[str, Any]:
    inbox = ROUND_B_INBOX / role_id
    pkg_name = ROUND_B_PACKAGES[role_id]
    path = inbox / pkg_name
    errors: list[str] = []
    corrections: list[str] = []
    if not path.exists():
        return {
            "role_id": role_id,
            "ok": False,
            "path": None,
            "errors": ["NO_RETURNED_ROUND_B_PACKAGE"],
            "corrections": [f"Place `{pkg_name}` into round_b_inbox/{role_id}/."],
            "rows_expected": len(_expected_pairs(role_id)),
            "rows_completed": 0,
        }

    expected = _expected_pairs(role_id)
    expected_keys = {(r["review_id"], r["candidate_index"]) for r in expected}
    expected_surface = {
        (r["review_id"], r["candidate_index"]): r["candidate_surface"] for r in expected
    }
    try:
        answers = _load_answers_from_xlsx(path)
    except Exception as exc:  # noqa: BLE001
        return {
            "role_id": role_id,
            "ok": False,
            "path": str(path.as_posix()),
            "errors": [f"read_failed:{exc}"],
            "corrections": ["Restore official Round B package structure."],
            "rows_expected": len(expected),
            "rows_completed": 0,
        }

    got_keys = set(answers)
    missing = sorted(expected_keys - got_keys)
    extra = sorted(got_keys - expected_keys)
    if missing:
        errors.append(f"missing_pairs:{len(missing)}")
        corrections.append(f"Restore {len(missing)} missing Round B candidate rows.")
    if extra:
        errors.append(f"extra_pairs:{len(extra)}")
        corrections.append(f"Remove {len(extra)} unexpected Round B rows.")

    blank = 0
    invalid = 0
    surface_mismatch = 0
    enum_counts: Counter[str] = Counter()
    for key, row in answers.items():
        if key not in expected_keys:
            continue
        if expected_surface[key] != row["candidate_surface"]:
            surface_mismatch += 1
        acc = row["acceptability"]
        if not acc:
            blank += 1
            enum_counts["(blank)"] += 1
        elif acc not in OFFICIAL_B:
            invalid += 1
            enum_counts[acc] += 1
        else:
            enum_counts[acc] += 1

    if blank:
        errors.append(f"missing_acceptability:{blank}")
        corrections.append(f"Fill acceptability for {blank} blank Round B rows.")
    if invalid:
        errors.append(f"invalid_acceptability_enums:{invalid}")
        corrections.append("Use only official Round B acceptability enums.")
    if surface_mismatch:
        errors.append(f"candidate_surface_mismatch:{surface_mismatch}")
        corrections.append("Do not alter sealed candidate_surface values.")

    completed = sum(
        1
        for k, r in answers.items()
        if k in expected_keys and r["acceptability"] in OFFICIAL_B
    )
    ok = not errors and completed == len(expected)
    # Build surface-keyed map for agreement joins
    by_review_surface = {
        (r["review_id"], r["candidate_surface"]): r["acceptability"]
        for r in answers.values()
        if r["acceptability"] in OFFICIAL_B
    }
    by_review_preferred = {}
    for r in answers.values():
        if r["acceptability"] == "ACCEPTABLE_PREFERRED":
            by_review_preferred[r["review_id"]] = r["candidate_surface"]

    return {
        "role_id": role_id,
        "ok": ok,
        "path": str(path.relative_to(REPO)).replace("\\", "/"),
        "file_sha256": _sha(path),
        "errors": errors,
        "corrections": corrections,
        "rows_expected": len(expected),
        "rows_completed": completed,
        "rows_blank": blank,
        "rows_invalid_enum": invalid,
        "acceptability_counts": dict(enum_counts),
        "answers": answers,
        "by_review_surface": by_review_surface,
        "by_review_preferred": by_review_preferred,
        "option_a_remap_notes_present": any(
            "OPTION_A_MECHANICAL_REMAP" in (r.get("notes") or "") for r in answers.values()
        ),
    }


def agreement_round_b(reports: list[dict[str, Any]]) -> dict[str, Any]:
    """Agree Fluent A vs Linguist B on (source_item_id, candidate_surface)."""
    by_role = {r["role_id"]: r for r in reports if r.get("ok")}
    if "NEPALI_FLUENT_A" not in by_role or "PROFESSIONAL_LINGUIST_B" not in by_role:
        return {"status": "NOT_COMPUTED_MISSING_REVIEWERS", "ok": False, "paired_n": 0}

    mapping = json.loads(BLIND_MAPPING.read_text(encoding="utf-8"))
    fluent_ids = {
        row["source_item_id"]: row["review_id"]
        for row in mapping["rows"]
        if row["role_id"] == "NEPALI_FLUENT_A"
    }
    ling_ids = {
        row["source_item_id"]: row["review_id"]
        for row in mapping["rows"]
        if row["role_id"] == "PROFESSIONAL_LINGUIST_B"
    }
    fa = by_role["NEPALI_FLUENT_A"]["by_review_surface"]
    fb = by_role["PROFESSIONAL_LINGUIST_B"]["by_review_surface"]

    labels_a: list[str] = []
    labels_b: list[str] = []
    disagreements: list[dict[str, str]] = []
    for sid in sorted(set(fluent_ids) & set(ling_ids)):
        rid_a = fluent_ids[sid]
        rid_b = ling_ids[sid]
        # All surfaces present for either reviewer on this source item
        surfaces = {
            surf
            for (rid, surf) in set(fa) | set(fb)
            if rid in {rid_a, rid_b}
        }
        for surf in sorted(surfaces):
            la = fa.get((rid_a, surf))
            lb = fb.get((rid_b, surf))
            if not la or not lb:
                continue
            labels_a.append(la)
            labels_b.append(lb)
            if la != lb:
                disagreements.append(
                    {
                        "source_item_id": sid,
                        "candidate_surface": surf,
                        "nepali_fluent_a": la,
                        "professional_linguist_b": lb,
                        "review_id_a": rid_a,
                        "review_id_b": rid_b,
                    }
                )

    if not labels_a:
        return {"status": "NOT_COMPUTED_NO_OVERLAPPING_LABELS", "ok": False, "paired_n": 0}

    exact = exact_agreement(labels_a, labels_b)
    kappa = cohen_kappa(labels_a, labels_b)
    gates = ACCEPTANCE_GATES
    gate_ok = (
        exact >= float(gates["exact_disposition_agreement"])
        and kappa >= float(gates["cohen_kappa_min"])
    )
    return {
        "status": "COMPUTED",
        "paired_n": len(labels_a),
        "exact_acceptability_agreement": exact,
        "cohen_kappa": kappa,
        "disagreement_count": len(disagreements),
        "disagreements_sample": disagreements[:50],
        "gate_exact_min": gates["exact_disposition_agreement"],
        "gate_kappa_min": gates["cohen_kappa_min"],
        "gates_met_diagnostic": gate_ok,
        "ok": True,
        "note": (
            "Diagnostic agreement only. Adjudication + credential verification "
            "still required before LINGUIST_APPROVED."
        ),
    }


def build_adjudication_packet(agreement: dict[str, Any]) -> dict[str, Any]:
    ADJUDICATION_DIR.mkdir(parents=True, exist_ok=True)
    rows = agreement.get("disagreements_sample") or []
    # Full disagreement list may be large; regenerate from sample + note if truncated
    packet = {
        "schema_version": "1.0.0",
        "status": (
            "NO_DISAGREEMENTS"
            if agreement.get("disagreement_count", 0) == 0
            else "AWAITING_INDEPENDENT_ADJUDICATOR"
        ),
        "fields": [
            "source_item_id",
            "candidate_surface",
            "nepali_fluent_a",
            "professional_linguist_b",
            "adjudicated_acceptability",
            "rationale_category",
            "human_review_remains_required",
            "inclusion_decision",
        ],
        "disagreement_count": agreement.get("disagreement_count", 0),
        "rows": [
            {
                **d,
                "adjudicated_acceptability": "",
                "rationale_category": "",
                "human_review_remains_required": True,
                "inclusion_decision": "",
            }
            for d in rows
        ],
        "note": (
            "If disagreement_count > len(rows), full set is in ROUND_B agreement report; "
            "expand packet before adjudication."
        ),
    }
    path = ADJUDICATION_DIR / "MAI_07_V3_ROUND_B_ADJUDICATION_PACKET.json"
    _write_json(path, packet)
    return {
        "path": str(path.relative_to(REPO)).replace("\\", "/"),
        "sha256": _sha(path),
        "status": packet["status"],
        "disagreement_count": packet["disagreement_count"],
    }


def run_round_b_validation(*, inbox: Path = ROUND_B_INBOX) -> dict[str, Any]:
    ROUND_B_VALIDATION.mkdir(parents=True, exist_ok=True)
    ROUND_B_LOCKED.mkdir(parents=True, exist_ok=True)

    sealed = verify_sealed_authorities()
    if not sealed["ok"]:
        result = {
            "status": "BLOCKED_PACKET_INTEGRITY_FAILED",
            "sealed": sealed,
            "ROUND_B_LOCKED": False,
            "LINGUIST_APPROVED": False,
            "QUALITY_GATES_PASSED": False,
        }
        _write_json(ROUND_B_VALIDATION / "ROUND_B_VALIDATION_SUMMARY.json", result)
        return result

    # Require Round A lock still present
    round_a_lock = OUT / "MAI_07_V3_ROUND_A_LOCK_MANIFEST.json"
    if not round_a_lock.exists():
        result = {
            "status": "BLOCKED_ROUND_A_NOT_LOCKED",
            "ROUND_B_LOCKED": False,
            "LINGUIST_APPROVED": False,
            "QUALITY_GATES_PASSED": False,
        }
        _write_json(ROUND_B_VALIDATION / "ROUND_B_VALIDATION_SUMMARY.json", result)
        return result

    reports = [validate_role_round_b(role_id) for role_id in ROUND_B_ROLES]
    for rep in reports:
        slim = {k: v for k, v in rep.items() if k not in {"answers", "by_review_surface", "by_review_preferred"}}
        _write_json(ROUND_B_VALIDATION / f"ROUND_B_VALIDATION__{rep['role_id']}.json", slim)

    all_ok = all(r.get("ok") for r in reports) and sealed["ok"]
    option_a = any(r.get("option_a_remap_notes_present") for r in reports)

    if not all_ok:
        summary = {
            "status": "ROUND_B_CORRECTION_REQUIRED",
            "ROUND_B_LOCKED": False,
            "LINGUIST_APPROVED": False,
            "QUALITY_GATES_PASSED": False,
            "PRODUCTION_APPROVED": False,
            "option_a_mechanical_remap": option_a,
            "roles": [
                {
                    "role_id": r["role_id"],
                    "ok": r.get("ok"),
                    "rows_expected": r.get("rows_expected"),
                    "rows_completed": r.get("rows_completed"),
                    "errors": r.get("errors"),
                    "corrections": r.get("corrections"),
                }
                for r in reports
            ],
            "next_human_action": "Correct Round B packages in round_b_inbox/<ROLE>/ then re-run.",
            "created_utc": datetime.now(timezone.utc).isoformat(),
        }
        _write_json(ROUND_B_VALIDATION / "ROUND_B_VALIDATION_SUMMARY.json", summary)
        return summary

    agreement = agreement_round_b(reports)
    # Expand full disagreement list into adjudication packet when needed
    if agreement.get("ok") and agreement.get("disagreement_count", 0) > 0:
        mapping = json.loads(BLIND_MAPPING.read_text(encoding="utf-8"))
        by_role = {r["role_id"]: r for r in reports}
        fluent_ids = {
            row["source_item_id"]: row["review_id"]
            for row in mapping["rows"]
            if row["role_id"] == "NEPALI_FLUENT_A"
        }
        ling_ids = {
            row["source_item_id"]: row["review_id"]
            for row in mapping["rows"]
            if row["role_id"] == "PROFESSIONAL_LINGUIST_B"
        }
        fa = by_role["NEPALI_FLUENT_A"]["by_review_surface"]
        fb = by_role["PROFESSIONAL_LINGUIST_B"]["by_review_surface"]
        all_d: list[dict[str, str]] = []
        for sid in sorted(set(fluent_ids) & set(ling_ids)):
            rid_a = fluent_ids[sid]
            rid_b = ling_ids[sid]
            surfaces = {surf for (rid, surf) in set(fa) | set(fb) if rid in {rid_a, rid_b}}
            for surf in sorted(surfaces):
                la = fa.get((rid_a, surf))
                lb = fb.get((rid_b, surf))
                if la and lb and la != lb:
                    all_d.append(
                        {
                            "source_item_id": sid,
                            "candidate_surface": surf,
                            "nepali_fluent_a": la,
                            "professional_linguist_b": lb,
                            "review_id_a": rid_a,
                            "review_id_b": rid_b,
                        }
                    )
        agreement["disagreements_sample"] = all_d
        agreement["disagreement_count"] = len(all_d)

    adjudication = build_adjudication_packet(agreement)

    locked_files: dict[str, Any] = {}
    for role_id in ROUND_B_ROLES:
        src = ROUND_B_INBOX / role_id / ROUND_B_PACKAGES[role_id]
        dest = ROUND_B_LOCKED / ROUND_B_PACKAGES[role_id]
        if dest.exists() and _sha(dest) != _sha(src):
            raise FileExistsError(f"already_locked_with_different_content:{dest}")
        if not dest.exists():
            shutil.copy2(src, dest)
        csv_src = ROUND_B_INBOX / role_id / ROUND_B_CSVS[role_id]
        csv_dest = ROUND_B_LOCKED / ROUND_B_CSVS[role_id]
        if csv_src.exists() and not csv_dest.exists():
            shutil.copy2(csv_src, csv_dest)
        # Export canonical JSON responses
        answers = next(r for r in reports if r["role_id"] == role_id)["answers"]
        canon_rows = sorted(
            [
                {
                    "review_id": a["review_id"],
                    "candidate_index": a["candidate_index"],
                    "candidate_surface": a["candidate_surface"],
                    "acceptability": a["acceptability"],
                }
                for a in answers.values()
            ],
            key=lambda r: (r["review_id"], int(r["candidate_index"]) if str(r["candidate_index"]).isdigit() else r["candidate_index"]),
        )
        export = ROUND_B_LOCKED / f"MAI_07_V3_ROUND_B_RESPONSES__{role_id}.json"
        _write_json(export, {"role_id": role_id, "rows": canon_rows})
        locked_files[role_id] = {
            "locked_path": str(dest.relative_to(REPO)).replace("\\", "/"),
            "source_path": str(src.relative_to(REPO)).replace("\\", "/"),
            "raw_sha256": _sha(dest),
            "canonical_sha256": hashlib.sha256(
                json.dumps(canon_rows, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
            ).hexdigest(),
            "source_matches_locked": _sha(src) == _sha(dest),
            "rows": len(canon_rows),
        }

    # Domain completion gate
    acct = next(r for r in reports if r["role_id"] == "ACCOUNTING_DOMAIN")
    domain_completion = (
        acct["rows_completed"] / acct["rows_expected"] if acct["rows_expected"] else 0.0
    )

    lock_manifest = {
        "schema_version": "1.0.0",
        "phase": "MAI-07R3J-A-ROUND-B-LOCK-AND-VALIDATION",
        "status": "ROUND_B_LOCKED",
        "ROUND_A_LOCKED": True,
        "ROUND_B_READY": True,
        "ROUND_B_LOCKED": True,
        "sealed": {
            "blind_mapping_sha256": sealed.get("blind_mapping_sha256"),
            "packet_manifest_sha256": sealed.get("packet_manifest_sha256"),
            "expected_blind_mapping_sha256": EXPECTED_BLIND_MAPPING_SHA,
            "expected_packet_manifest_sha256": EXPECTED_PACKET_MANIFEST_SHA,
            "ok": sealed.get("ok"),
        },
        "locked_files": locked_files,
        "agreement_round_b": {
            k: v for k, v in agreement.items() if k != "disagreements_sample"
        },
        "adjudication_packet": adjudication,
        "option_a_mechanical_remap": option_a,
        "domain_review_completion_technical": domain_completion,
        "QUALITY_GATES_PASSED": False,
        "LINGUIST_APPROVED": False,
        "PRODUCTION_APPROVED": False,
        "MAI_08": "NOT_STARTED",
        "adjudication_performed": False,
        "human_answers_altered": False,
        "model_evaluation_performed": False,
        "authority_note": (
            "Round B locked from inbox submissions"
            + (" (Option A mechanical enum remap)" if option_a else "")
            + ". LINGUIST_APPROVED remains false until adjudication + manual credential verification."
        ),
        "next_human_action": (
            "Send adjudication packet to independent adjudicator if disagreements remain; "
            "then manually verify professional linguist credentials before any approval flags."
            if adjudication.get("disagreement_count", 0) > 0
            else "No Round B disagreements sampled. Still verify linguist credentials manually before LINGUIST_APPROVED."
        ),
        "created_utc": datetime.now(timezone.utc).isoformat(),
    }

    lock_path = OUT / "MAI_07_V3_ROUND_B_LOCK_MANIFEST.json"
    lock_sha = _write_json(lock_path, lock_manifest)
    lock_manifest["lock_manifest_sha256"] = lock_sha
    lock_sha = _write_json(lock_path, lock_manifest)

    result = {
        "status": "ROUND_B_LOCKED",
        "ROUND_A_LOCKED": True,
        "ROUND_B_LOCKED": True,
        "ROUND_B_READY": True,
        "lock_manifest_path": str(lock_path.relative_to(REPO)).replace("\\", "/"),
        "lock_manifest_sha256": lock_sha,
        "locked_files": locked_files,
        "agreement_round_b": lock_manifest["agreement_round_b"],
        "adjudication_packet": adjudication,
        "option_a_mechanical_remap": option_a,
        "QUALITY_GATES_PASSED": False,
        "LINGUIST_APPROVED": False,
        "PRODUCTION_APPROVED": False,
        "human_answers_altered": False,
        "model_evaluation_performed": False,
        "next_human_action": lock_manifest["next_human_action"],
    }
    _write_json(ROUND_B_VALIDATION / "ROUND_B_VALIDATION_SUMMARY.json", result)
    # also copy lock under ops
    shutil.copy2(lock_path, OPS / "hash_manifests" / "MAI_07_V3_ROUND_B_LOCK_MANIFEST.json")
    return result


def main() -> int:
    result = run_round_b_validation()
    print(json.dumps({k: v for k, v in result.items() if k != "locked_files"}, indent=2))
    return 0 if result.get("status") == "ROUND_B_LOCKED" else 2


if __name__ == "__main__":
    raise SystemExit(main())
