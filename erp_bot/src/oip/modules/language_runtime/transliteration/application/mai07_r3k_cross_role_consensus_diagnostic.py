"""MAI-07R3K-AI-ASSISTED-CROSS-ROLE-CONSENSUS-DIAGNOSTIC

Fail-closed consolidation of AI-assisted user-accepted role judgments.
Engineering diagnostics only — not gold, not human IRR, not official Round A.
"""

from __future__ import annotations

import hashlib
import json
import shutil
from collections import Counter, defaultdict
from dataclasses import asdict
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from .import_mai07_r3j_ai_assisted_accounting import (
    CanonicalReviewRecord as AcctRecord,
    compute_semantic_hash as compute_accounting_semantic_hash,
)
from .import_mai07_r3j_ai_assisted_remaining_roles import (
    CanonicalReviewRecord as RemRecord,
    compute_semantic_hash as compute_remaining_semantic_hash,
)
from .mai07_r3ja_v3_agreement import ROUND_A_DISPOSITIONS
from .mai07_r3ja_v3_firewall import REPO
from .mai07_r3k_cross_role_contracts import (
    AgreementDiagnosticV1,
    ConsensusDiagnosticReportV1,
    CrossRoleDecisionV1,
    DisagreementReasonV1,
    RoleJudgmentV1,
    RiskQueueItemV1,
)
from .mai07_r3k_hash_contract import (
    display_abbreviation,
    format_display_citation,
    require_full_sha256,
    typed_hash_fields,
    validate_machine_hash_map,
)
from .validate_mai07_r3ja_round_a import CONFIDENCE

PHASE = "MAI-07R3K-AI-ASSISTED-CROSS-ROLE-CONSENSUS-DIAGNOSTIC"
SCHEMA_ID = "mai07_v3_ai_assisted_cross_role_consensus_diagnostic_v1"

EXPECTED_ACCT_SEMANTIC = "b96bec29e30ddcdc6dce1a5ef09a2003ee9de003a336cd98b43341c6e55e363b"
EXPECTED_REM_SEMANTIC = "1cc783d79cc3cc5f3f2daa288ae8b4721238fed584dbfb540597c8f883a8f4a1"
EXPECTED_UNIQUE_CASES = 1111
EXPECTED_TOTAL_JUDGMENTS = 3944
EXPECTED_FOUR_ROLE = 611
EXPECTED_THREE_ROLE = 500

ACCT_JSONL = (
    REPO
    / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/accounting_domain/canonical"
    / "ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
)
REM_JSONL = (
    REPO
    / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/remaining_roles/canonical"
    / "REMAINING_ROLES_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
)
BLIND_MAPPING = REPO / "docs/mokxya-ai/reviews/mai07_v3/V3_BLIND_MAPPING.json"
EXPECTED_BLIND_MAPPING_SHA = "d0875db79185b034b080e69f77f1220417cdc24dae5a6fb755a56b472af414f1"
DRAFT_AUDITS = {
    "PRODUCT_POLICY": REPO
    / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/role_drafts/product_policy/round_a_drafts/AI_ASSISTED_DRAFT_AUDIT.json",
    "NEPALI_FLUENT_A": REPO
    / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/role_drafts/nepali_fluent_a/round_a_drafts/AI_ASSISTED_DRAFT_AUDIT.json",
    "PROFESSIONAL_LINGUIST_B": REPO
    / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/role_drafts/professional_linguist_b/round_a_drafts/AI_ASSISTED_DRAFT_AUDIT.json",
}
OFFICIAL_INBOX = (
    REPO / "docs/mokxya-ai/reviews/mai07_v3/review_operations/round_a_inbox"
)
DEFAULT_OUT = (
    REPO / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/cross_role_diagnostic"
)

FIXED_PROVENANCE: dict[str, Any] = {
    "status": "AI_ASSISTED_CROSS_ROLE_CONSENSUS_DIAGNOSTIC",
    "independent_human_review": False,
    "professional_linguist_adjudication": False,
    "linguist_approved": False,
    "production_approved": False,
    "official_round_a_lock_eligible": False,
    "round_b_authorized": False,
    "frozen_v3_quality_gate_authorized": False,
    "prohibited_for_training": True,
    "eligible_for_frozen_quality_gold": False,
    "majority_voting_is_gold": False,
    "agreement_is_independent_human_irr": False,
    "ai_autofill_used": True,
    "user_accepted": True,
}

REQUIRED_ROLES_ALWAYS = ("PRODUCT_POLICY", "NEPALI_FLUENT_A", "PROFESSIONAL_LINGUIST_B")
ACCOUNTING_ROLE = "ACCOUNTING_DOMAIN"
YES_NO = frozenset({"YES", "NO"})
REVIEW_DISPOSITIONS = frozenset(
    {"IDENTITY_FIRST_REVIEW_REQUIRED", "CONTEXT_DEPENDENT", "ABSTAIN_CANNOT_DECIDE"}
)
REQUIRED_DECISION_DISPOSITIONS = frozenset(
    {
        "ENGLISH_IDENTITY_REQUIRED",
        "DEVANAGARI_TRANSLITERATION_REQUIRED",
        "NO_TRANSLITERATION_ALLOWED",
        "ACRONYM_OR_IDENTIFIER",
        "PROTECTED",
        "NAME_OR_ENTITY",
        "TRANSLITERATION_OPTIONAL",
    }
)
SAFETY_SENSITIVE = frozenset(
    {
        "PROTECTED",
        "ACRONYM_OR_IDENTIFIER",
        "ENGLISH_IDENTITY_REQUIRED",
        "DEVANAGARI_TRANSLITERATION_REQUIRED",
        "ABSTAIN_CANNOT_DECIDE",
    }
)

_PACKET_SEED = "mai07-r3k-targeted-packet-20260717"


class Mai07R3KDiagnosticError(ValueError):
    pass


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return sha256_file(path)


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        json.dumps(r, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for r in rows
    ]
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8", newline="\n")
    return sha256_file(path)


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise Mai07R3KDiagnosticError(f"missing_jsonl:{path}")
    out: list[dict[str, Any]] = []
    for ln in path.read_text(encoding="utf-8").splitlines():
        if ln.strip():
            out.append(json.loads(ln))
    return out


def assert_official_inbox_empty() -> None:
    if not OFFICIAL_INBOX.exists():
        return
    hits = list(OFFICIAL_INBOX.rglob("*.xlsx")) + list(OFFICIAL_INBOX.rglob("*.xlsm"))
    if hits:
        raise Mai07R3KDiagnosticError(f"official_inbox_not_empty:{[p.name for p in hits]}")


def verify_input_semantic_hashes() -> dict[str, str]:
    acct_rows = load_jsonl(ACCT_JSONL)
    rem_rows = load_jsonl(REM_JSONL)
    if len(acct_rows) != 611:
        raise Mai07R3KDiagnosticError(f"acct_row_count:{len(acct_rows)}")
    if len(rem_rows) != 3333:
        raise Mai07R3KDiagnosticError(f"rem_row_count:{len(rem_rows)}")

    acct_recs = []
    for r in acct_rows:
        acct_recs.append(
            AcctRecord(
                review_id=r["review_id"],
                input_text=r["input_text"],
                highlighted_span=r["highlighted_span"],
                disposition=r["disposition"],
                confidence=r["confidence"],
                reason_category=r["reason_category"],
                natural_context_ok=r["natural_context_ok"],
                suspected_ambiguity=r["suspected_ambiguity"],
                reviewer_notes=r["reviewer_notes"],
                source_workbook=r["source_workbook"],
                source_row=int(r["source_row"]),
                role_id=r["role_id"],
                review_method=r["review_method"],
                independent_human_review=bool(r["independent_human_review"]),
                ai_autofill_used=bool(r["ai_autofill_used"]),
                user_accepted=bool(r["user_accepted"]),
                professional_linguist_adjudication=bool(r["professional_linguist_adjudication"]),
                prohibited_for_training=bool(r["prohibited_for_training"]),
                eligible_for_frozen_quality_gold=bool(r["eligible_for_frozen_quality_gold"]),
            )
        )
    rem_recs = []
    for r in rem_rows:
        rem_recs.append(
            RemRecord(
                review_id=r["review_id"],
                input_text=r["input_text"],
                highlighted_span=r["highlighted_span"],
                disposition=r["disposition"],
                confidence=r["confidence"],
                reason_category=r["reason_category"],
                natural_context_ok=r["natural_context_ok"],
                suspected_ambiguity=r["suspected_ambiguity"],
                reviewer_notes=r["reviewer_notes"],
                source_workbook=r["source_workbook"],
                source_row=int(r["source_row"]),
                role_id=r["role_id"],
                review_method=r["review_method"],
                independent_human_review=bool(r["independent_human_review"]),
                ai_autofill_used=bool(r["ai_autofill_used"]),
                user_accepted=bool(r["user_accepted"]),
                professional_linguist_adjudication=bool(r["professional_linguist_adjudication"]),
                prohibited_for_training=bool(r["prohibited_for_training"]),
                eligible_for_frozen_quality_gold=bool(r["eligible_for_frozen_quality_gold"]),
                professional_linguist_b_is_ai_role_simulation=bool(
                    r.get("professional_linguist_b_is_ai_role_simulation", False)
                ),
            )
        )
    acct_h = compute_accounting_semantic_hash(acct_recs)
    rem_h = compute_remaining_semantic_hash(rem_recs)
    if acct_h != EXPECTED_ACCT_SEMANTIC:
        raise Mai07R3KDiagnosticError(
            f"acct_semantic_mismatch:expected={EXPECTED_ACCT_SEMANTIC}:actual={acct_h}"
        )
    if rem_h != EXPECTED_REM_SEMANTIC:
        raise Mai07R3KDiagnosticError(
            f"rem_semantic_mismatch:expected={EXPECTED_REM_SEMANTIC}:actual={rem_h}"
        )
    for r in acct_recs + rem_recs:  # type: ignore[operator]
        if r.independent_human_review or r.professional_linguist_adjudication:
            raise Mai07R3KDiagnosticError("false_provenance_independent_or_linguist")
        if r.eligible_for_frozen_quality_gold:
            raise Mai07R3KDiagnosticError("false_provenance_frozen_gold")
        if r.disposition not in ROUND_A_DISPOSITIONS:
            raise Mai07R3KDiagnosticError(f"unknown_disposition:{r.disposition}")
        if r.confidence not in CONFIDENCE:
            raise Mai07R3KDiagnosticError(f"unknown_confidence:{r.confidence}")
        if r.natural_context_ok not in YES_NO or r.suspected_ambiguity not in YES_NO:
            raise Mai07R3KDiagnosticError(f"invalid_yes_no:{r.review_id}")
    return {"accounting": acct_h, "remaining": rem_h}


def load_blind_mapping() -> dict[str, dict[str, str]]:
    """review_id -> {role_id, source_item_id}"""
    if sha256_file(BLIND_MAPPING) != EXPECTED_BLIND_MAPPING_SHA:
        raise Mai07R3KDiagnosticError("blind_mapping_hash_mismatch")
    data = json.loads(BLIND_MAPPING.read_text(encoding="utf-8"))
    out: dict[str, dict[str, str]] = {}
    for row in data["rows"]:
        rid = str(row["review_id"])
        if rid in out:
            raise Mai07R3KDiagnosticError(f"duplicate_mapping_review_id:{rid}")
        out[rid] = {"role_id": str(row["role_id"]), "source_item_id": str(row["source_item_id"])}
    return out


def load_decision_sources() -> dict[str, str]:
    """remaining-role review_id -> ACCOUNTING_VERIFIED_CONTENT_MAP | HEURISTIC_V1"""
    out: dict[str, str] = {}
    for role_id, path in DRAFT_AUDITS.items():
        if not path.is_file():
            raise Mai07R3KDiagnosticError(f"missing_draft_audit:{role_id}")
        audit = json.loads(path.read_text(encoding="utf-8"))
        for row in audit["rows"]:
            rid = str(row["review_id"])
            src = str(row["source"])
            if src not in {"ACCOUNTING_VERIFIED_CONTENT_MAP", "HEURISTIC_V1"}:
                raise Mai07R3KDiagnosticError(f"unknown_decision_source:{src}")
            if rid in out and out[rid] != src:
                raise Mai07R3KDiagnosticError(f"decision_source_conflict:{rid}")
            out[rid] = src
    return out


def _diagnostic_case_id(source_item_id: str) -> str:
    digest = hashlib.sha256(f"{_PACKET_SEED}:{source_item_id}".encode("utf-8")).hexdigest()[:12]
    return f"V3DX-{digest}"


def build_judgments(
    acct_rows: list[dict[str, Any]],
    rem_rows: list[dict[str, Any]],
    mapping: Mapping[str, Mapping[str, str]],
    decision_sources: Mapping[str, str],
) -> dict[str, list[RoleJudgmentV1]]:
    by_case: dict[str, list[RoleJudgmentV1]] = defaultdict(list)
    seen_role_case: set[tuple[str, str]] = set()

    def add(row: dict[str, Any], decision_source: str) -> None:
        rid = str(row["review_id"])
        if rid not in mapping:
            raise Mai07R3KDiagnosticError(f"review_id_not_in_mapping:{rid}")
        meta = mapping[rid]
        if meta["role_id"] != row["role_id"]:
            raise Mai07R3KDiagnosticError(f"role_mismatch_mapping:{rid}")
        key = (meta["source_item_id"], row["role_id"])
        if key in seen_role_case:
            raise Mai07R3KDiagnosticError(f"duplicate_role_decision:{key}")
        seen_role_case.add(key)
        # authority text must match across roles for same case — checked later
        j = RoleJudgmentV1(
            role_id=str(row["role_id"]),
            review_id=rid,
            disposition=str(row["disposition"]),
            confidence=str(row["confidence"]),
            reason_category=str(row["reason_category"]),
            natural_context_ok=str(row["natural_context_ok"]),
            suspected_ambiguity=str(row["suspected_ambiguity"]),
            reviewer_notes=str(row["reviewer_notes"]),
            source_workbook=str(row["source_workbook"]),
            source_row=int(row["source_row"]),
            decision_source=decision_source,  # type: ignore[arg-type]
            review_method=str(row["review_method"]),
            independent_human_review=False,
            ai_autofill_used=True,
            user_accepted=True,
            professional_linguist_adjudication=False,
            prohibited_for_training=True,
            eligible_for_frozen_quality_gold=False,
            professional_linguist_b_is_ai_role_simulation=bool(
                row.get("professional_linguist_b_is_ai_role_simulation", False)
            )
            or row["role_id"] == "PROFESSIONAL_LINGUIST_B",
        )
        by_case[meta["source_item_id"]].append(j)

    for row in acct_rows:
        add(row, "ACCOUNTING_DOMAIN_VERIFIED_IMPORT")
    for row in rem_rows:
        src = decision_sources.get(str(row["review_id"]))
        if not src:
            raise Mai07R3KDiagnosticError(f"missing_decision_source:{row['review_id']}")
        add(row, src)
    return by_case


def _exact_agreement_status(disps: list[str]) -> str:
    if len(disps) <= 1:
        return "SINGLE_ROLE"
    counts = Counter(disps)
    if len(counts) == 1:
        return "UNANIMOUS"
    if len(counts) == len(disps):
        return "ALL_DIFFERENT"
    top = counts.most_common(1)[0][1]
    if top > len(disps) / 2:
        return "MAJORITY"
    return "SPLIT_NO_MAJORITY"


def _disagreement_and_risks(judgments: list[RoleJudgmentV1]) -> tuple[list[DisagreementReasonV1], list[str]]:
    reasons: list[DisagreementReasonV1] = []
    flags: list[str] = []
    disps = [j.disposition for j in judgments]
    confs = [j.confidence for j in judgments]
    nats = [j.natural_context_ok for j in judgments]
    ambs = [j.suspected_ambiguity for j in judgments]
    disp_set = set(disps)

    if len(disp_set) > 1:
        reasons.append(DisagreementReasonV1("DISPOSITION_DISAGREE", ",".join(sorted(disp_set))))
        flags.append("DISPOSITION_DISAGREE")
    if any(c in {"LOW", "MEDIUM"} for c in confs):
        flags.append("LOW_OR_MEDIUM_CONFIDENCE")
    if any(a == "YES" for a in ambs):
        flags.append("SUSPECTED_AMBIGUITY_YES")
    if "ABSTAIN_CANNOT_DECIDE" in disp_set:
        flags.append("ABSTAIN_PRESENT")
    if disp_set & {"CONTEXT_DEPENDENT", "IDENTITY_FIRST_REVIEW_REQUIRED", "TRANSLITERATION_OPTIONAL"}:
        flags.append("SOFT_OR_OPTIONAL_DISPOSITION")
    if len(nats) > 1 and len(set(nats)) > 1:
        reasons.append(DisagreementReasonV1("NATURAL_CONTEXT_OK_CONFLICT", ",".join(nats)))
        flags.append("NATURAL_CONTEXT_OK_CONFLICT")
    if {"ACRONYM_OR_IDENTIFIER", "PROTECTED"} <= disp_set or (
        "ACRONYM_OR_IDENTIFIER" in disp_set and "NAME_OR_ENTITY" in disp_set
    ):
        flags.append("IDENTIFIER_CLASSIFICATION_CONFLICT")
        reasons.append(DisagreementReasonV1("IDENTIFIER_CLASSIFICATION_CONFLICT", ",".join(sorted(disp_set))))
    if "ENGLISH_IDENTITY_REQUIRED" in disp_set and "DEVANAGARI_TRANSLITERATION_REQUIRED" in disp_set:
        flags.append("ENGLISH_VS_DEVANAGARI_CONFLICT")
        reasons.append(DisagreementReasonV1("ENGLISH_VS_DEVANAGARI_CONFLICT", "ENGLISH+DEVANAGARI"))
    if "NAME_OR_ENTITY" in disp_set and (
        "DEVANAGARI_TRANSLITERATION_REQUIRED" in disp_set or "ENGLISH_IDENTITY_REQUIRED" in disp_set
    ):
        flags.append("NAME_ENTITY_POLICY_CONFLICT")
    if "PROTECTED" in disp_set and len(disp_set) > 1:
        flags.append("PROTECTED_VS_OTHER_CONFLICT")
    has_review = bool(disp_set & REVIEW_DISPOSITIONS)
    has_required = bool(disp_set & REQUIRED_DECISION_DISPOSITIONS)
    if has_review and has_required:
        flags.append("REVIEW_VS_REQUIRED_CONFLICT")
        reasons.append(DisagreementReasonV1("REVIEW_VS_REQUIRED_CONFLICT", "mixed review/required"))
    if any(j.decision_source == "HEURISTIC_V1" for j in judgments) and (
        disp_set & SAFETY_SENSITIVE or "LOW_OR_MEDIUM_CONFIDENCE" in flags
    ):
        flags.append("HEURISTIC_V1_SAFETY_SENSITIVE")
    return reasons, sorted(set(flags))


def _risk_tier(flags: list[str]) -> str:
    critical = {
        "ENGLISH_VS_DEVANAGARI_CONFLICT",
        "PROTECTED_VS_OTHER_CONFLICT",
        "IDENTIFIER_CLASSIFICATION_CONFLICT",
    }
    high = {
        "DISPOSITION_DISAGREE",
        "REVIEW_VS_REQUIRED_CONFLICT",
        "HEURISTIC_V1_SAFETY_SENSITIVE",
        "ABSTAIN_PRESENT",
    }
    medium = {
        "LOW_OR_MEDIUM_CONFIDENCE",
        "SUSPECTED_AMBIGUITY_YES",
        "SOFT_OR_OPTIONAL_DISPOSITION",
        "NATURAL_CONTEXT_OK_CONFLICT",
        "NAME_ENTITY_POLICY_CONFLICT",
    }
    if set(flags) & critical:
        return "TIER_1_CRITICAL"
    if set(flags) & high:
        return "TIER_2_HIGH"
    if set(flags) & medium:
        return "TIER_3_MEDIUM"
    if flags:
        return "TIER_4_LOW"
    return "TIER_4_LOW"


def consolidate_with_text(
    by_case: Mapping[str, list[RoleJudgmentV1]],
    case_text: Mapping[str, tuple[str, str]],
) -> list[CrossRoleDecisionV1]:
    decisions: list[CrossRoleDecisionV1] = []
    for source_item_id in sorted(by_case.keys()):
        judgments = sorted(by_case[source_item_id], key=lambda j: j.role_id)
        roles = {j.role_id for j in judgments}
        for req in REQUIRED_ROLES_ALWAYS:
            if req not in roles:
                raise Mai07R3KDiagnosticError(
                    f"missing_required_role:{source_item_id}:{req}"
                )
        has_acct = ACCOUNTING_ROLE in roles
        expected_n = 4 if has_acct else 3
        if len(judgments) != expected_n:
            raise Mai07R3KDiagnosticError(
                f"role_count:{source_item_id}:{len(judgments)}!={expected_n}"
            )
        text, span = case_text[source_item_id]
        # authority consistency
        for j in judgments:
            # text recovered separately — verify all roles share same text via case_text only
            pass
        # Verify text identity was established when building case_text
        disps = [j.disposition for j in judgments]
        status = _exact_agreement_status(disps)
        reasons, flags = _disagreement_and_risks(judgments)
        if not has_acct and len(judgments) != 3:
            flags.append("UNEXPECTED_ROLE_EVIDENCE")
        if has_acct and len(judgments) != 4:
            flags.append("UNEXPECTED_ROLE_EVIDENCE")
            reasons.append(DisagreementReasonV1("UNEXPECTED_ROLE_EVIDENCE", str(len(judgments))))

        decisions.append(
            CrossRoleDecisionV1(
                source_item_id=source_item_id,
                diagnostic_case_id=_diagnostic_case_id(source_item_id),
                input_text=text,
                highlighted_span=span,
                role_count=len(judgments),
                has_accounting_domain=has_acct,
                judgments=tuple(judgments),
                disposition_set=tuple(sorted(set(disps))),
                exact_agreement_status=status,  # type: ignore[arg-type]
                dispositions_unanimous=(status == "UNANIMOUS"),
                disagreement_reasons=tuple(reasons),
                risk_flags=tuple(flags),
                decision_sources=tuple(sorted({j.decision_source for j in judgments})),
                heuristic_v1_present=any(j.decision_source == "HEURISTIC_V1" for j in judgments),
                accounting_map_present=any(
                    j.decision_source == "ACCOUNTING_VERIFIED_CONTENT_MAP" for j in judgments
                ),
                majority_as_gold=False,
            )
        )
    if len(decisions) != EXPECTED_UNIQUE_CASES:
        raise Mai07R3KDiagnosticError(f"unique_cases:{len(decisions)}!={EXPECTED_UNIQUE_CASES}")
    four = sum(1 for d in decisions if d.role_count == 4)
    three = sum(1 for d in decisions if d.role_count == 3)
    if four != EXPECTED_FOUR_ROLE or three != EXPECTED_THREE_ROLE:
        raise Mai07R3KDiagnosticError(f"four_three:{four}/{three}")
    return decisions


def build_case_text_map(
    acct_rows: list[dict[str, Any]],
    rem_rows: list[dict[str, Any]],
    mapping: Mapping[str, Mapping[str, str]],
) -> dict[str, tuple[str, str]]:
    case_text: dict[str, tuple[str, str]] = {}
    for row in acct_rows + rem_rows:
        sid = mapping[row["review_id"]]["source_item_id"]
        pair = (str(row["input_text"]), str(row["highlighted_span"]))
        if sid in case_text and case_text[sid] != pair:
            raise Mai07R3KDiagnosticError(f"altered_authority_text_span:{sid}")
        case_text[sid] = pair
    return case_text


def compute_agreement(decisions: list[CrossRoleDecisionV1]) -> AgreementDiagnosticV1:
    three = [d for d in decisions]
    four = [d for d in decisions if d.role_count == 4]

    def exact_disp(d: CrossRoleDecisionV1, roles: tuple[str, ...] | None = None) -> bool:
        js = d.judgments
        if roles:
            js = tuple(j for j in js if j.role_id in roles)
        return len({j.disposition for j in js}) == 1

    three_roles = REQUIRED_ROLES_ALWAYS
    three_agree = sum(1 for d in three if exact_disp(d, three_roles))
    four_agree = sum(1 for d in four if exact_disp(d))

    unanimous = sum(1 for d in decisions if d.exact_agreement_status == "UNANIMOUS")
    majority = sum(1 for d in decisions if d.exact_agreement_status == "MAJORITY")
    all_diff = sum(1 for d in decisions if d.exact_agreement_status == "ALL_DIFFERENT")
    split = sum(1 for d in decisions if d.exact_agreement_status == "SPLIT_NO_MAJORITY")
    abstain = sum(1 for d in decisions if "ABSTAIN_PRESENT" in d.risk_flags)
    review_req = sum(1 for d in decisions if "REVIEW_VS_REQUIRED_CONFLICT" in d.risk_flags)

    # By disposition: for each disposition, among cases where ALL roles chose it / where any chose it
    by_disp: dict[str, Any] = {}
    for disp in ROUND_A_DISPOSITIONS:
        any_n = sum(1 for d in decisions if disp in d.disposition_set)
        all_n = sum(1 for d in decisions if d.disposition_set == (disp,))
        by_disp[disp] = {
            "cases_any_role": any_n,
            "cases_unanimous_this_disposition": all_n,
        }

    conf_agree = sum(
        1 for d in decisions if len({j.confidence for j in d.judgments}) == 1
    )
    nat_agree = sum(
        1 for d in decisions if len({j.natural_context_ok for j in d.judgments}) == 1
    )
    amb_agree = sum(
        1 for d in decisions if len({j.suspected_ambiguity for j in d.judgments}) == 1
    )

    # Inherited map vs heuristic: bucket by whether HEURISTIC_V1 present on remaining roles
    map_only = [d for d in decisions if d.accounting_map_present and not d.heuristic_v1_present]
    heur = [d for d in decisions if d.heuristic_v1_present]
    by_src = {
        "ACCOUNTING_MAP_INHERITED_NO_HEURISTIC": {
            "n": len(map_only),
            "three_role_exact_agree": sum(1 for d in map_only if exact_disp(d, three_roles)),
        },
        "HEURISTIC_V1_PRESENT": {
            "n": len(heur),
            "three_role_exact_agree": sum(1 for d in heur if exact_disp(d, three_roles)),
        },
    }

    # Pairwise
    pairs = [
        ("PRODUCT_POLICY", "NEPALI_FLUENT_A"),
        ("PRODUCT_POLICY", "PROFESSIONAL_LINGUIST_B"),
        ("NEPALI_FLUENT_A", "PROFESSIONAL_LINGUIST_B"),
        ("PRODUCT_POLICY", "ACCOUNTING_DOMAIN"),
        ("NEPALI_FLUENT_A", "ACCOUNTING_DOMAIN"),
        ("PROFESSIONAL_LINGUIST_B", "ACCOUNTING_DOMAIN"),
    ]
    pairwise: dict[str, Any] = {}
    for a, b in pairs:
        eligible = [
            d
            for d in decisions
            if {a, b}.issubset({j.role_id for j in d.judgments})
        ]
        agree = 0
        for d in eligible:
            da = next(j.disposition for j in d.judgments if j.role_id == a)
            db = next(j.disposition for j in d.judgments if j.role_id == b)
            if da == db:
                agree += 1
        pairwise[f"{a}__{b}"] = {
            "n": len(eligible),
            "exact_disposition_agree": agree,
            "rate": (agree / len(eligible)) if eligible else None,
        }

    taxonomy = Counter()
    for d in decisions:
        for f in d.risk_flags:
            if f.endswith("_CONFLICT") or f in {
                "ABSTAIN_PRESENT",
                "DISPOSITION_DISAGREE",
                "HEURISTIC_V1_SAFETY_SENSITIVE",
            }:
                taxonomy[f] += 1

    # Non-independent AI-output similarity: simple percent agreement only (explicitly labeled)
    # Do NOT call this Cohen's kappa / human IRR.
    ai_sim = {
        "label": "NON_INDEPENDENT_AI_OUTPUT_SIMILARITY_ONLY",
        "not_human_inter_rater_reliability": True,
        "three_role_percent_agreement": three_agree / len(three) if three else None,
        "four_role_percent_agreement": four_agree / len(four) if four else None,
        "note": "Percent agreement among AI-assisted role simulations; not Cohen kappa / Krippendorff alpha as human IRR.",
    }

    return AgreementDiagnosticV1(
        three_role_exact_disposition_agreement_rate=three_agree / len(three),
        three_role_exact_disposition_agree_count=three_agree,
        three_role_n=len(three),
        four_role_exact_disposition_agreement_rate=(four_agree / len(four)) if four else 0.0,
        four_role_exact_disposition_agree_count=four_agree,
        four_role_n=len(four),
        unanimous_count=unanimous,
        majority_count=majority,
        all_different_count=all_diff,
        split_no_majority_count=split,
        abstention_containing_count=abstain,
        review_vs_required_conflict_count=review_req,
        agreement_by_disposition=by_disp,
        agreement_by_confidence={
            "cases_all_roles_same_confidence": conf_agree,
            "n": len(decisions),
            "rate": conf_agree / len(decisions),
        },
        natural_context_ok_agreement={
            "cases_all_roles_same": nat_agree,
            "n": len(decisions),
            "rate": nat_agree / len(decisions),
        },
        suspected_ambiguity_agreement={
            "cases_all_roles_same": amb_agree,
            "n": len(decisions),
            "rate": amb_agree / len(decisions),
        },
        agreement_by_decision_source_bucket=by_src,
        pairwise_role_disposition_agreement=pairwise,
        conflict_taxonomy_counts=dict(sorted(taxonomy.items())),
        ai_output_similarity_metrics_non_independent=ai_sim,
    )


def build_risk_queue(decisions: list[CrossRoleDecisionV1]) -> list[RiskQueueItemV1]:
    queue: list[RiskQueueItemV1] = []
    for d in decisions:
        reasons, flags = list(d.disagreement_reasons), list(d.risk_flags)
        # Inclusion rules
        include = bool(flags) or len(d.disposition_set) > 1
        if not include:
            continue
        # Always include if any inclusion criterion from spec
        codes = list(flags)
        if not codes and len(d.disposition_set) > 1:
            codes = ["DISPOSITION_DISAGREE"]
        tier = _risk_tier(codes)
        queue.append(
            RiskQueueItemV1(
                diagnostic_case_id=d.diagnostic_case_id,
                source_item_id=d.source_item_id,
                risk_tier=tier,  # type: ignore[arg-type]
                reason_codes=tuple(sorted(set(codes))),
                input_text=d.input_text,
                highlighted_span=d.highlighted_span,
                role_count=d.role_count,
                disposition_set=d.disposition_set,
            )
        )
    # Deterministic order: tier, case id
    tier_order = {"TIER_1_CRITICAL": 0, "TIER_2_HIGH": 1, "TIER_3_MEDIUM": 2, "TIER_4_LOW": 3}
    queue.sort(key=lambda x: (tier_order[x.risk_tier], x.diagnostic_case_id))
    return queue


def compute_diagnostic_semantic_hash(
    decisions: list[CrossRoleDecisionV1],
    risk_queue: list[RiskQueueItemV1],
    agreement: AgreementDiagnosticV1,
) -> str:
    payload = {
        "schema": SCHEMA_ID,
        "phase": PHASE,
        "provenance": FIXED_PROVENANCE,
        "decisions": [asdict(d) for d in decisions],
        "risk_queue": [asdict(r) for r in risk_queue],
        "agreement": asdict(agreement),
        "majority_as_gold": False,
    }
    return sha256_bytes(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
            "utf-8"
        )
    )


def write_blinded_targeted_packet(
    risk_queue: list[RiskQueueItemV1],
    decisions: list[CrossRoleDecisionV1],
    out_dir: Path,
) -> dict[str, Any]:
    """Reviewer-facing packet: opaque id + text + span only. No AI labels."""
    packet_dir = out_dir / "targeted_review_packet"
    private_dir = out_dir / "private_adjudication_import_only"
    packet_dir.mkdir(parents=True, exist_ok=True)
    private_dir.mkdir(parents=True, exist_ok=True)

    # Cap packet to risk queue (all risk items) — full diagnostic set for human targeting
    items = risk_queue
    by_case = {d.diagnostic_case_id: d for d in decisions}

    # Blinded workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TARGETED_REVIEW"
    headers = (
        "diagnostic_review_id",
        "input_text",
        "highlighted_span",
        "context_note",
        "disposition",
        "confidence",
        "reviewer_notes",
    )
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    private_rows = []
    for item in items:
        # Opaque ID distinct from diagnostic_case_id for reviewer surface
        opaque = "TGT-" + hashlib.sha256(
            f"{_PACKET_SEED}:tgt:{item.diagnostic_case_id}".encode("utf-8")
        ).hexdigest()[:12]
        ctx = (
            "Review the highlighted span in context. Use allowed disposition enums only. "
            "Do not consult AI prior labels or other reviewers."
        )
        ws.append([opaque, item.input_text, item.highlighted_span, ctx, "", "", ""])
        d = by_case[item.diagnostic_case_id]
        private_rows.append(
            {
                "opaque_diagnostic_review_id": opaque,
                "diagnostic_case_id": item.diagnostic_case_id,
                "source_item_id": item.source_item_id,
                "risk_tier": item.risk_tier,
                "reason_codes": list(item.reason_codes),
                "ai_disposition_set": list(item.disposition_set),
                "role_judgments": [asdict(j) for j in d.judgments],
                "use": "adjudication_import_only",
            }
        )

    # Dropdown for disposition / confidence on blank columns
    if items:
        dv_d = DataValidation(
            type="list",
            formula1='"' + ",".join(ROUND_A_DISPOSITIONS) + '"',
            allow_blank=True,
        )
        dv_c = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=True)
        ws.add_data_validation(dv_d)
        ws.add_data_validation(dv_c)
        dv_d.add(f"E2:E{len(items)+1}")
        dv_c.add(f"F2:F{len(items)+1}")

    xlsx_path = packet_dir / "MokXya_MAI07_V3_R3K_TARGETED_BLIND_REVIEW.xlsx"
    wb.save(xlsx_path)

    # Leakage scan of reviewer-facing sheet values
    forbidden_tokens = (
        "PRODUCT_POLICY",
        "NEPALI_FLUENT_A",
        "PROFESSIONAL_LINGUIST_B",
        "ACCOUNTING_DOMAIN",
        "HEURISTIC_V1",
        "ACCOUNTING_VERIFIED",
        "majority_as_gold",
        "UNANIMOUS",
        "TIER_1",
    )
    for row in ws.iter_rows(min_row=2, values_only=True):
        blob = " ".join("" if c is None else str(c) for c in row)
        for tok in forbidden_tokens:
            if tok in blob:
                raise Mai07R3KDiagnosticError(f"packet_leakage:{tok}")
        # AI dispositions must not appear in filled columns A-D (E-G blank)
        for disp in ROUND_A_DISPOSITIONS:
            if row[4] or row[5] or row[6]:
                raise Mai07R3KDiagnosticError("packet_prefilled_answers")

    instructions = packet_dir / "INSTRUCTIONS.md"
    instructions.write_text(
        "\n".join(
            [
                "# MAI-07R3K Targeted Blind Review (AI-assisted diagnostic follow-up)",
                "",
                "This packet is **not** official independent Round A.",
                "It is a targeted engineering follow-up derived from AI-assisted cross-role disagreements.",
                "",
                "## Do",
                "- Fill disposition and confidence for each row using allowed enums.",
                "- Judge from input text and highlighted span only.",
                "",
                "## Do not",
                "- Do not treat this as professional-linguist or independent Round A evidence.",
                "- Do not place completed files in the official V3 `round_a_inbox`.",
                "- Do not request or use AI prior labels, role names, or agreement scores.",
                "",
                f"Rows: {len(items)}",
                "",
            ]
        ),
        encoding="utf-8",
        newline="\n",
    )

    _write_jsonl(private_dir / "PRIVATE_BLIND_MAPPING_ADJUDICATION_IMPORT_ONLY.jsonl", private_rows)
    _write_json(
        private_dir / "PRIVATE_MANIFEST.json",
        {
            "use": "adjudication_import_only",
            "prohibited_for_reviewer_distribution": True,
            "item_count": len(private_rows),
            "not_official_round_a": True,
        },
    )

    # Ensure workbook bytes don't embed private mapping sheet
    from openpyxl import load_workbook

    check = load_workbook(xlsx_path, read_only=True)
    try:
        if set(check.sheetnames) != {"TARGETED_REVIEW"}:
            raise Mai07R3KDiagnosticError(f"unexpected_sheets:{check.sheetnames}")
    finally:
        check.close()

    return {
        "item_count": len(items),
        "xlsx_path": str(xlsx_path),
        "xlsx_sha256": sha256_file(xlsx_path),
        "leakage_scan_passed": True,
    }


def run_diagnostic(*, out_root: Path | None = None, write: bool = True) -> dict[str, Any]:
    assert_official_inbox_empty()
    hashes = verify_input_semantic_hashes()
    mapping = load_blind_mapping()
    decision_sources = load_decision_sources()
    acct_rows = load_jsonl(ACCT_JSONL)
    rem_rows = load_jsonl(REM_JSONL)

    if len(acct_rows) + len(rem_rows) != EXPECTED_TOTAL_JUDGMENTS:
        raise Mai07R3KDiagnosticError("judgment_count")

    by_case = build_judgments(acct_rows, rem_rows, mapping, decision_sources)
    case_text = build_case_text_map(acct_rows, rem_rows, mapping)
    decisions = consolidate_with_text(by_case, case_text)
    agreement = compute_agreement(decisions)
    risk_queue = build_risk_queue(decisions)
    semantic = compute_diagnostic_semantic_hash(decisions, risk_queue, agreement)

    role_counts = Counter()
    for d in decisions:
        for j in d.judgments:
            role_counts[j.role_id] += 1

    # Hash contract: full 64-char typed fields only; never mix raw ZIP with semantic.
    acct_sem = require_full_sha256(hashes["accounting"], label="input_accounting_semantic_hash")
    rem_sem = require_full_sha256(hashes["remaining"], label="input_remaining_semantic_hash")
    sem = require_full_sha256(semantic, label="r3k_semantic_sha256")
    zip_path = REPO / "MokXya_MAI07_V3_ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED_READY.zip"
    zip_raw = require_full_sha256(sha256_file(zip_path), label="accounting_package_zip_raw_sha256")
    hash_contract = typed_hash_fields(
        accounting_package_zip_raw_sha256=zip_raw,
        accounting_import_semantic_sha256=acct_sem,
        remaining_roles_import_semantic_sha256=rem_sem,
        r3k_semantic_sha256=sem,
    )
    validate_machine_hash_map(hash_contract)
    # Display abbreviations derived from ONE full hash each (never prefix∥suffix across fields).
    display_citations = {
        "accounting_import_semantic_display": display_abbreviation(acct_sem),
        "accounting_package_zip_raw_display": display_abbreviation(zip_raw),
        "remaining_roles_import_semantic_display": display_abbreviation(rem_sem),
        "r3k_semantic_display": display_abbreviation(sem),
    }

    report = ConsensusDiagnosticReportV1(
        phase=PHASE,
        schema=SCHEMA_ID,
        ok=True,
        semantic_hash=sem,
        input_accounting_semantic_hash=acct_sem,
        input_remaining_semantic_hash=rem_sem,
        unique_source_item_ids=len(decisions),
        total_role_judgments=EXPECTED_TOTAL_JUDGMENTS,
        four_role_cases=EXPECTED_FOUR_ROLE,
        three_role_cases=EXPECTED_THREE_ROLE,
        role_counts=dict(sorted(role_counts.items())),
        agreement=agreement,
        risk_queue_count=len(risk_queue),
        risk_tier_counts=dict(sorted(Counter(r.risk_tier for r in risk_queue).items())),
        targeted_packet_item_count=len(risk_queue),
        governance={
            "QUALITY_GATES_PASSED": False,
            "LINGUIST_APPROVED": False,
            "PRODUCTION_APPROVED": False,
            "ROUND_A_LOCKED": False,
            "ROUND_B_READY": False,
            "majority_voting_is_gold": False,
            "agreement_is_independent_human_irr": False,
            "MAI-07": "NEEDS_CORRECTIVE_WORK",
            "MAI-08": "NOT_STARTED",
        },
        provenance=FIXED_PROVENANCE,
    )

    result: dict[str, Any] = {
        "ok": True,
        "semantic_hash": sem,
        "report": asdict(report),
        "decisions": decisions,
        "risk_queue": risk_queue,
        "agreement": agreement,
    }

    if write:
        root = out_root or DEFAULT_OUT
        root.mkdir(parents=True, exist_ok=True)
        historical_root = root.resolve() == DEFAULT_OUT.resolve()
        decisions_path = root / "canonical" / "CROSS_ROLE_DECISIONS.jsonl"
        if historical_root and decisions_path.is_file():
            # Never silently overwrite sealed R3K historical evidence (CASE A / REPORT_ONLY).
            existing_sem = require_full_sha256(
                json.loads((root / "reports" / "SEMANTIC_HASH.json").read_text(encoding="utf-8"))[
                    "semantic_hash"
                ],
                label="existing_r3k_semantic",
            )
            if existing_sem != sem:
                raise Mai07R3KDiagnosticError(
                    f"historical_r3k_semantic_drift:{existing_sem}!={sem}"
                )
            canon = [asdict(d) for d in decisions]
            risk_rows = [asdict(r) for r in risk_queue]
            rebuilt_dec = (
                "\n".join(
                    json.dumps(r, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                    for r in canon
                )
                + ("\n" if canon else "")
            ).encode("utf-8")
            rebuilt_risk = (
                "\n".join(
                    json.dumps(r, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                    for r in risk_rows
                )
                + ("\n" if risk_rows else "")
            ).encode("utf-8")
            if rebuilt_dec != decisions_path.read_bytes():
                raise Mai07R3KDiagnosticError("historical_rebuild_decisions_byte_mismatch")
            if rebuilt_risk != (root / "canonical" / "RISK_QUEUE.jsonl").read_bytes():
                raise Mai07R3KDiagnosticError("historical_rebuild_risk_queue_byte_mismatch")
            # New sidecar only — does not alter historical report/canonical files.
            _write_json(root / "reports" / "HASH_CONTRACT.json", hash_contract)
            result["hash_contract"] = hash_contract
            result["evidence_root"] = str(root)
            result["historical_outputs_preserved"] = True
        else:
            canon = [asdict(d) for d in decisions]
            _write_jsonl(decisions_path, canon)
            _write_jsonl(
                root / "canonical" / "RISK_QUEUE.jsonl",
                [asdict(r) for r in risk_queue],
            )
            report_payload = asdict(report)
            report_payload["hash_contract"] = {
                **hash_contract,
                "display_citations": display_citations,
                "machine_hashes_must_be_64_lowercase_hex": True,
                "raw_and_semantic_are_distinct_fields": True,
                "display_abbreviation_must_derive_from_single_full_hash": True,
            }
            _write_json(root / "reports" / "CONSENSUS_DIAGNOSTIC_REPORT.json", report_payload)
            _write_json(root / "reports" / "AGREEMENT_DIAGNOSTIC.json", asdict(agreement))
            _write_json(
                root / "reports" / "SEMANTIC_HASH.json",
                {
                    "phase": PHASE,
                    "schema": SCHEMA_ID,
                    "semantic_hash": sem,
                    "unique_cases": EXPECTED_UNIQUE_CASES,
                    "total_judgments": EXPECTED_TOTAL_JUDGMENTS,
                    "hash_contract": hash_contract,
                },
            )
            _write_json(root / "reports" / "HASH_CONTRACT.json", hash_contract)
            packet_meta = write_blinded_targeted_packet(risk_queue, decisions, root)
            result["packet"] = packet_meta
            _write_json(root / "reports" / "TARGETED_PACKET_META.json", packet_meta)
            (root / "README.md").write_text(
                "\n".join(
                    [
                        "# MAI-07R3K AI-Assisted Cross-Role Consensus Diagnostic",
                        "",
                        "Engineering diagnostics only. **Not** independent human IRR, **not** gold,",
                        "**not** official Round A, **not** majority-as-authority.",
                        "",
                        f"Semantic hash (full): `{sem}`",
                        format_display_citation("R3K semantic (display)", sem),
                        format_display_citation("Accounting import semantic (display)", acct_sem),
                        format_display_citation("Accounting ZIP raw (display)", zip_raw),
                        format_display_citation("Remaining-roles import semantic (display)", rem_sem),
                        f"Cases: {EXPECTED_UNIQUE_CASES} · Judgments: {EXPECTED_TOTAL_JUDGMENTS}",
                        f"Risk queue / targeted packet rows: {len(risk_queue)}",
                        "",
                    ]
                ),
                encoding="utf-8",
                newline="\n",
            )
            result["hash_contract"] = hash_contract
            result["evidence_root"] = str(root)
            result["historical_outputs_preserved"] = False

    assert_official_inbox_empty()
    return result


def prove_deterministic_rerun(work_dir: Path) -> dict[str, Any]:
    a = run_diagnostic(out_root=work_dir / "a", write=True)
    b = run_diagnostic(out_root=work_dir / "b", write=True)
    if a["semantic_hash"] != b["semantic_hash"]:
        raise Mai07R3KDiagnosticError("semantic_hash_not_deterministic")
    p1 = (work_dir / "a/canonical/CROSS_ROLE_DECISIONS.jsonl").read_bytes()
    p2 = (work_dir / "b/canonical/CROSS_ROLE_DECISIONS.jsonl").read_bytes()
    if p1 != p2:
        raise Mai07R3KDiagnosticError("canonical_not_deterministic")
    q1 = (work_dir / "a/canonical/RISK_QUEUE.jsonl").read_bytes()
    q2 = (work_dir / "b/canonical/RISK_QUEUE.jsonl").read_bytes()
    if q1 != q2:
        raise Mai07R3KDiagnosticError("risk_queue_not_deterministic")
    return {
        "ok": True,
        "semantic_hash": a["semantic_hash"],
        "canonical_sha256": sha256_bytes(p1),
        "risk_queue_sha256": sha256_bytes(q1),
    }


def main(argv: list[str] | None = None) -> int:
    import argparse

    p = argparse.ArgumentParser(description=PHASE)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--prove-deterministic", action="store_true")
    args = p.parse_args(argv)
    try:
        if args.prove_deterministic:
            proof = prove_deterministic_rerun(REPO / "tmp_mai07_r3k_det_proof")
            print(json.dumps(proof, indent=2))
            return 0
        result = run_diagnostic(out_root=args.out, write=True)
        print(
            json.dumps(
                {
                    "ok": True,
                    "semantic_hash": result["semantic_hash"],
                    "unique_cases": EXPECTED_UNIQUE_CASES,
                    "total_judgments": EXPECTED_TOTAL_JUDGMENTS,
                    "risk_queue_count": len(result["risk_queue"]),
                    "three_role_agree_rate": result["agreement"].three_role_exact_disposition_agreement_rate,
                    "four_role_agree_rate": result["agreement"].four_role_exact_disposition_agreement_rate,
                },
                indent=2,
            )
        )
        return 0
    except Mai07R3KDiagnosticError as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, indent=2))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
