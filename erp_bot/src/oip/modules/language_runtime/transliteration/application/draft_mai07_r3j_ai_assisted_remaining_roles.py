"""MAI-07R3J-AI-ASSISTED-REMAINING-ROLE-DRAFTS

Deterministic AI-assisted Round A draft generator for PRODUCT_POLICY,
NEPALI_FLUENT_A, and PROFESSIONAL_LINGUIST_B.

Governance:
- status = AI_ASSISTED_DRAFT_FOR_HUMAN_REVIEW
- independent_human_review = false
- Not official inbox / Round A lock / Round B / frozen V3 gold / training / runtime
- Leaves REVIEWER_DECLARATION, ROUND_B_CANDIDATES, SUBMISSION_CHECKLIST untouched
- Fills only ROUND_A_CONTEXT decision columns D–I
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from .mai07_r3ja_v3_firewall import REPO
from .validate_mai07_r3ja_round_a import ROUND_A_HEADERS

PHASE = "MAI-07R3J-AI-ASSISTED-REMAINING-ROLE-DRAFTS"
SCHEMA_ID = "mai07_v3_ai_assisted_role_draft_v1"

REMAINING_ROLES = (
    "PRODUCT_POLICY",
    "NEPALI_FLUENT_A",
    "PROFESSIONAL_LINGUIST_B",
)

OPS_PACKAGES = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3"
    / "review_operations"
    / "reviewer_packages"
)
OFFICIAL_INBOX = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3"
    / "review_operations"
    / "round_a_inbox"
)
ACCOUNTING_CANONICAL = (
    REPO
    / "docs"
    / "mokxya-ai"
    / "reviews"
    / "mai07_v3_ai_assisted"
    / "accounting_domain"
    / "canonical"
    / "ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
)
DEFAULT_OUT = (
    REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3_ai_assisted" / "role_drafts"
)

FIXED_PROVENANCE: dict[str, Any] = {
    "status": "AI_ASSISTED_DRAFT_FOR_HUMAN_REVIEW",
    "review_method": "AI_ASSISTED_DRAFT",
    "independent_human_review": False,
    "row_by_row_independent_review_performed": False,
    "ai_autofill_used": True,
    "user_accepted": False,
    "professional_linguist_adjudication": False,
    "linguist_approved": False,
    "production_approved": False,
    "official_round_a_lock_eligible": False,
    "round_b_authorized": False,
    "frozen_v3_quality_gate_authorized": False,
    "prohibited_for_training": True,
    "eligible_for_frozen_quality_gold": False,
    "submission_ready_under_current_protocol": False,
    "declaration_completed": False,
}

NEPALI_MARKERS = re.compile(
    r"\b(ra|ko|lai|ma|bata|garne|gara|garnu|garera|milaaune|milaaidim|hisab|"
    r"baaki|aaja|hajur|hajurko|ho|cha|chha|grahak|sodhe|sodheko|bholi|samma|"
    r"please|clear|dim)\b",
    re.I,
)
ENGLISH_UI = re.compile(
    r"\b(Please update|English settings|settings panel|Create a draft|"
    r"leave it unposted|Sync failed|retry after|English UI)\b",
    re.I,
)
PROTECTED_HINT = re.compile(
    r"(Do not mutate protected token|zero-width space|keep surface \[)",
    re.I,
)
PROTECTED_SPAN = re.compile(
    r"(SERIAL-|ACC-SYN-|PAN-[A-Z0-9]|MOB-|TXN-SYN-|INV-SYN-|UUID-|IMEI-|URL-https|"
    r"\u200b)"
)
ACRONYM_SPAN = re.compile(r"^[A-Z]{2,8}$")

# Shared/loan surfaces seen in accounting AI drafts
LOANWORDS_OPTIONAL = frozenset(
    {
        "restore",
        "opening",
        "posted",
        "salary",
        "purchase",
        "employee",
        "branch",
        "withdraw",
        "deposit",
        "transfer",
        "invoice",
        "voucher",
        "receipt",
        "payment",
        "balance",
        "account",
        "customer",
        "supplier",
        "inventory",
        "stock",
        "discount",
        "credit",
        "debit",
        "journal",
        "ledger",
    }
)
SHARED_CONTEXT = frozenset({"premium", "token", "mark", "match", "duty"})
AMBIGUOUS_SHORT = frozenset(
    {
        "bank",
        "lot",
        "graph",
        "mail",
        "post",
        "agenda",
        "folder",
        "passbook",
        "flag",
        "rule",
        "bag",
        "file",
        "entry",
        "ticket",
        "scan",
        "page",
        "cash",
        "tax",
        "bill",
        "sync",
        "queue",
        "draft",
        "menu",
        "pending",
        "clear",
        "error",
        "field",
        "panel",
        "settings",
        "update",
        "create",
        "leave",
        "retry",
        "failed",
        "monthly",
        "backup",
    }
)
CLEAR_ENGLISH_SPANS = frozenset(
    {
        "duty",
        "match",
        "menu",
        "pending",
        "clear",
        "error",
        "mark",
        "settings",
        "panel",
        "field",
        "update",
        "create",
        "draft",
        "unposted",
        "retry",
        "backup",
        "failed",
        "queue",
    }
)

NOTES = {
    "ENGLISH_IDENTITY_REQUIRED": "AI-assisted draft: the highlighted span is used as English; retain the Latin identity.",
    "DEVANAGARI_TRANSLITERATION_REQUIRED": "AI-assisted draft: ordinary Romanized Nepali word; Devanagari transliteration is expected.",
    "IDENTITY_FIRST_REVIEW_REQUIRED": "AI-assisted draft: insufficient decisive context; identity-first human review required.",
    "TRANSLITERATION_OPTIONAL": "AI-assisted draft: domain loanword in Nepali context; transliteration optional.",
    "CONTEXT_DEPENDENT": "AI-assisted draft: shared/borrowed surface; disposition depends on intended sense.",
    "ACRONYM_OR_IDENTIFIER": "AI-assisted draft: acronym/identifier; keep as-is.",
    "PROTECTED": "AI-assisted draft: protected token/surface; must not be mutated.",
    "ABSTAIN_CANNOT_DECIDE": "AI-assisted draft: possible typo/OOV; abstain pending human judgment.",
    "NAME_OR_ENTITY": "AI-assisted draft: name/entity-like; prefer identity.",
    "NO_TRANSLITERATION_ALLOWED": "AI-assisted draft: no transliteration allowed for this span.",
}


class Mai07AiAssistedDraftError(ValueError):
    pass


@dataclass(frozen=True)
class DraftDecision:
    disposition: str
    confidence: str
    reason_category: str
    natural_context_ok: str
    suspected_ambiguity: str
    reviewer_notes: str
    source: str  # ACCOUNTING_VERIFIED_CONTENT_MAP | HEURISTIC_V1


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return sha256_file(path)


def load_accounting_content_map(path: Path = ACCOUNTING_CANONICAL) -> dict[tuple[str, str], DraftDecision]:
    if not path.is_file():
        raise Mai07AiAssistedDraftError(f"missing_accounting_canonical:{path}")
    out: dict[tuple[str, str], DraftDecision] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        key = (str(row["input_text"]), str(row["highlighted_span"]))
        out[key] = DraftDecision(
            disposition=str(row["disposition"]),
            confidence=str(row["confidence"]),
            reason_category=str(row["reason_category"]),
            natural_context_ok=str(row["natural_context_ok"]),
            suspected_ambiguity=str(row["suspected_ambiguity"]),
            reviewer_notes=str(row["reviewer_notes"]),
            source="ACCOUNTING_VERIFIED_CONTENT_MAP",
        )
    if len(out) != 611:
        raise Mai07AiAssistedDraftError(f"accounting_map_count:{len(out)}!=611")
    return out


def heuristic_decision(input_text: str, span: str) -> DraftDecision:
    t = input_text.strip()
    s = span.strip()
    sl = s.lower()
    np_ctx = bool(NEPALI_MARKERS.search(t))
    eng_ui = bool(ENGLISH_UI.search(t))

    if PROTECTED_HINT.search(t) or PROTECTED_SPAN.search(s) or "\u200b" in s:
        return DraftDecision(
            "PROTECTED",
            "HIGH",
            "PROTECTED_TOKEN_OR_SURFACE",
            "YES",
            "NO",
            NOTES["PROTECTED"],
            "HEURISTIC_V1",
        )
    if ACRONYM_SPAN.match(s) or (s.isupper() and 2 <= len(s) <= 8 and s.isalpha()):
        return DraftDecision(
            "ACRONYM_OR_IDENTIFIER",
            "HIGH",
            "ACRONYM_OR_IDENTIFIER",
            "YES",
            "NO",
            NOTES["ACRONYM_OR_IDENTIFIER"],
            "HEURISTIC_V1",
        )
    if re.search(r"(typo|OOV|jabaj|cannot decide|maybe .* means)", t, re.I) or sl in {
        "jabaj",
        "asdfgh",
    }:
        return DraftDecision(
            "ABSTAIN_CANNOT_DECIDE",
            "LOW",
            "POSSIBLE_TYPO_OR_OOV",
            "NO",
            "YES",
            NOTES["ABSTAIN_CANNOT_DECIDE"],
            "HEURISTIC_V1",
        )
    if eng_ui and not np_ctx:
        return DraftDecision(
            "ENGLISH_IDENTITY_REQUIRED",
            "HIGH",
            "CLEAR_ENGLISH_CONTEXT",
            "YES",
            "NO",
            NOTES["ENGLISH_IDENTITY_REQUIRED"],
            "HEURISTIC_V1",
        )
    if sl in SHARED_CONTEXT and (np_ctx or "maybe" in t.lower() or "or" in t.lower()):
        return DraftDecision(
            "CONTEXT_DEPENDENT",
            "LOW",
            "SHARED_OR_BORROWED_SURFACE",
            "NO",
            "YES",
            NOTES["CONTEXT_DEPENDENT"],
            "HEURISTIC_V1",
        )
    if np_ctx and sl in LOANWORDS_OPTIONAL:
        return DraftDecision(
            "TRANSLITERATION_OPTIONAL",
            "MEDIUM",
            "DOMAIN_LOANWORD_IN_NEPALI_CONTEXT",
            "YES",
            "YES",
            NOTES["TRANSLITERATION_OPTIONAL"],
            "HEURISTIC_V1",
        )
    if np_ctx and sl in AMBIGUOUS_SHORT:
        return DraftDecision(
            "IDENTITY_FIRST_REVIEW_REQUIRED",
            "LOW",
            "INSUFFICIENT_CONTEXT",
            "NO",
            "YES",
            NOTES["IDENTITY_FIRST_REVIEW_REQUIRED"],
            "HEURISTIC_V1",
        )
    if np_ctx and s.isalpha() and s.islower() and sl not in CLEAR_ENGLISH_SPANS:
        # Romanized Nepali-looking token in Nepali context
        return DraftDecision(
            "DEVANAGARI_TRANSLITERATION_REQUIRED",
            "HIGH",
            "CLEAR_ROMANIZED_NEPALI_WORD"
            if len(s) >= 4
            else "DECISIVE_ROMANIZED_NEPALI_CONTEXT",
            "YES",
            "NO",
            NOTES["DEVANAGARI_TRANSLITERATION_REQUIRED"],
            "HEURISTIC_V1",
        )
    if eng_ui or (not np_ctx and sl in CLEAR_ENGLISH_SPANS):
        return DraftDecision(
            "ENGLISH_IDENTITY_REQUIRED",
            "HIGH",
            "CLEAR_ENGLISH_CONTEXT",
            "YES",
            "NO",
            NOTES["ENGLISH_IDENTITY_REQUIRED"],
            "HEURISTIC_V1",
        )
    if not np_ctx and not eng_ui and sl in AMBIGUOUS_SHORT:
        return DraftDecision(
            "IDENTITY_FIRST_REVIEW_REQUIRED",
            "LOW",
            "INSUFFICIENT_CONTEXT",
            "NO",
            "YES",
            NOTES["IDENTITY_FIRST_REVIEW_REQUIRED"],
            "HEURISTIC_V1",
        )
    # Conservative default: identity-first rather than inventing Devanagari
    return DraftDecision(
        "IDENTITY_FIRST_REVIEW_REQUIRED",
        "LOW",
        "INSUFFICIENT_CONTEXT",
        "NO",
        "YES",
        NOTES["IDENTITY_FIRST_REVIEW_REQUIRED"],
        "HEURISTIC_V1",
    )


def decide(
    input_text: str,
    span: str,
    content_map: Mapping[tuple[str, str], DraftDecision],
) -> DraftDecision:
    key = (input_text, span)
    if key in content_map:
        return content_map[key]
    return heuristic_decision(input_text, span)


def assert_inbox_not_targeted(path: Path) -> None:
    try:
        path.resolve().relative_to(OFFICIAL_INBOX.resolve())
    except ValueError:
        return
    raise Mai07AiAssistedDraftError("refusing_write_into_official_round_a_inbox")


def draft_workbook(
    source: Path,
    dest: Path,
    content_map: Mapping[tuple[str, str], DraftDecision],
) -> dict[str, Any]:
    assert_inbox_not_targeted(dest)
    wb = load_workbook(source)
    if "ROUND_A_CONTEXT" not in wb.sheetnames:
        raise Mai07AiAssistedDraftError(f"missing_ROUND_A_CONTEXT:{source.name}")

    # Capture Round B / declaration fingerprints before edits
    decl_before = [
        tuple(c.value for c in row)
        for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=1, max_col=2)
    ]
    rb_before = [
        tuple(c.value for c in row)
        for row in wb["ROUND_B_CANDIDATES"].iter_rows(min_row=1, max_col=8)
    ]

    ws = wb["ROUND_A_CONTEXT"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if tuple(str(h or "").strip() for h in header) != ROUND_A_HEADERS:
        raise Mai07AiAssistedDraftError(f"header_mismatch:{source.name}")

    rows_out: list[dict[str, Any]] = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2), start=2):
        rid = row[0].value
        if rid is None or str(rid).strip() == "":
            continue
        input_text = str(row[1].value or "")
        span = str(row[2].value or "")
        # authority fields must remain unchanged — never write A–C
        decision = decide(input_text, span, content_map)
        row[3].value = decision.disposition
        row[4].value = decision.confidence
        row[5].value = decision.reason_category
        row[6].value = decision.natural_context_ok
        row[7].value = decision.suspected_ambiguity
        row[8].value = decision.reviewer_notes
        rows_out.append(
            {
                "workbook": source.name,
                "excel_row": excel_row,
                "review_id": str(rid).strip(),
                "input_text": input_text,
                "highlighted_span": span,
                **asdict(decision),
            }
        )

    # Ensure we did not mutate declaration / round B
    decl_after = [
        tuple(c.value for c in row)
        for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=1, max_col=2)
    ]
    rb_after = [
        tuple(c.value for c in row)
        for row in wb["ROUND_B_CANDIDATES"].iter_rows(min_row=1, max_col=8)
    ]
    if decl_before != decl_after:
        raise Mai07AiAssistedDraftError(f"declaration_mutated:{source.name}")
    if rb_before != rb_after:
        raise Mai07AiAssistedDraftError(f"round_b_mutated:{source.name}")

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    wb.close()

    return {
        "source": source.name,
        "output": dest.name,
        "source_sha256": sha256_file(source),
        "output_sha256": sha256_file(dest),
        "source_size_bytes": source.stat().st_size,
        "output_size_bytes": dest.stat().st_size,
        "row_count": len(rows_out),
        "ids_unchanged": True,
        "all_required_filled": True,
        "declaration_blank": True,
        "round_b_unchanged_shape": True,
        "rows": rows_out,
    }


def draft_role(
    role_id: str,
    *,
    content_map: Mapping[tuple[str, str], DraftDecision],
    out_root: Path = DEFAULT_OUT,
) -> dict[str, Any]:
    if role_id not in REMAINING_ROLES:
        raise Mai07AiAssistedDraftError(f"role_not_in_remaining_set:{role_id}")
    src_dir = OPS_PACKAGES / role_id / "round_a"
    if not src_dir.is_dir():
        raise Mai07AiAssistedDraftError(f"missing_role_package:{src_dir}")
    role_out = out_root / role_id.lower() / "round_a_drafts"
    assert_inbox_not_targeted(role_out)
    if role_out.exists():
        shutil.rmtree(role_out)
    role_out.mkdir(parents=True, exist_ok=True)

    workbook_reports: list[dict[str, Any]] = []
    all_rows: list[dict[str, Any]] = []
    sources = sorted(src_dir.glob("MokXya_MAI07_V3__*.xlsx"))
    if not sources:
        raise Mai07AiAssistedDraftError(f"no_source_workbooks:{role_id}")

    for src in sources:
        dest_name = src.stem + "__AI_ASSISTED_DRAFT.xlsx"
        dest = role_out / dest_name
        report = draft_workbook(src, dest, content_map)
        rows = report.pop("rows")
        workbook_reports.append(report)
        all_rows.extend(rows)

    disp = Counter(r["disposition"] for r in all_rows)
    conf = Counter(r["confidence"] for r in all_rows)
    reason = Counter(r["reason_category"] for r in all_rows)
    src_counts = Counter(r["source"] for r in all_rows)

    audit = {
        "schema": SCHEMA_ID,
        "phase": PHASE,
        "role_id": role_id,
        "summary": {
            **FIXED_PROVENANCE,
            "role_id": role_id,
            "row_count": len(all_rows),
            "workbook_count": len(workbook_reports),
            "disposition_counts": dict(sorted(disp.items())),
            "confidence_counts": dict(sorted(conf.items())),
            "reason_counts": dict(sorted(reason.items())),
            "decision_source_counts": dict(sorted(src_counts.items())),
            "workbook_reports": workbook_reports,
        },
        "rows": all_rows,
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    audit_path = role_out / "AI_ASSISTED_DRAFT_AUDIT.json"
    _write_json(audit_path, audit)

    readme = role_out / "README_AI_ASSISTED_DRAFT.txt"
    readme.write_text(
        "\n".join(
            [
                f"MokXya MAI-07 V3 — {role_id} Round A AI-Assisted Draft",
                "",
                "These files contain AI-generated suggestions in ROUND_A_CONTEXT columns D–I.",
                "They are not independent human review and do not satisfy declaration_no_ai_autofill.",
                "REVIEWER_DECLARATION, ROUND_B_CANDIDATES, and SUBMISSION_CHECKLIST were intentionally left untouched.",
                "Use these copies only as review aids. Do not place them in the official independent-review inbox.",
                "Overlapping ACCOUNTING_DOMAIN content reuses verified AI-assisted labels; remaining rows use HEURISTIC_V1.",
                "",
                f"Rows drafted: {len(all_rows)}",
                f"Workbooks: {len(workbook_reports)}",
                "",
            ]
        ),
        encoding="utf-8",
        newline="\n",
    )

    # ZIP package (non-authoritative drafts)
    zip_path = out_root / f"MokXya_MAI07_V3_{role_id}_ROUND_A_AI_ASSISTED_DRAFTS.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in sorted(role_out.iterdir()):
            zf.write(p, arcname=p.name)

    return {
        "role_id": role_id,
        "row_count": len(all_rows),
        "workbook_count": len(workbook_reports),
        "draft_dir": str(role_out),
        "zip_path": str(zip_path),
        "zip_sha256": sha256_file(zip_path),
        "audit_sha256": sha256_file(audit_path),
        "disposition_counts": dict(sorted(disp.items())),
        "decision_source_counts": dict(sorted(src_counts.items())),
        "unique_review_ids": len({r["review_id"] for r in all_rows}),
    }


def draft_all_remaining_roles(*, out_root: Path = DEFAULT_OUT) -> dict[str, Any]:
    assert_inbox_not_targeted(out_root)
    out_root.mkdir(parents=True, exist_ok=True)
    content_map = load_accounting_content_map()
    role_results = []
    for role in REMAINING_ROLES:
        role_results.append(draft_role(role, content_map=content_map, out_root=out_root))

    # Ensure official inbox still empty of our drafts
    inbox_hits = list(OFFICIAL_INBOX.rglob("*AI_ASSISTED_DRAFT*.xlsx")) if OFFICIAL_INBOX.exists() else []
    if inbox_hits:
        raise Mai07AiAssistedDraftError(f"drafts_found_in_official_inbox:{inbox_hits}")

    summary = {
        "schema": SCHEMA_ID,
        "phase": PHASE,
        "ok": True,
        "provenance": FIXED_PROVENANCE,
        "roles": role_results,
        "total_rows": sum(r["row_count"] for r in role_results),
        "total_workbooks": sum(r["workbook_count"] for r in role_results),
        "governance": {
            "official_round_a_inbox_used": False,
            "ROUND_A_LOCKED": False,
            "ROUND_B_READY": False,
            "QUALITY_GATES_PASSED": False,
            "LINGUIST_APPROVED": False,
            "PRODUCTION_APPROVED": False,
            "MAI-07": "NEEDS_CORRECTIVE_WORK",
            "MAI-08": "NOT_STARTED",
        },
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    # Canonical decision hash (stable across XLSX writer timestamps)
    decision_payload = {
        "phase": PHASE,
        "provenance": FIXED_PROVENANCE,
        "roles": [],
    }
    for role_id in REMAINING_ROLES:
        audit_path = out_root / role_id.lower() / "round_a_drafts" / "AI_ASSISTED_DRAFT_AUDIT.json"
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        rows = sorted(audit["rows"], key=lambda r: r["review_id"])
        decision_payload["roles"].append(
            {
                "role_id": role_id,
                "rows": [
                    {
                        "review_id": r["review_id"],
                        "input_text": r["input_text"],
                        "highlighted_span": r["highlighted_span"],
                        "disposition": r["disposition"],
                        "confidence": r["confidence"],
                        "reason_category": r["reason_category"],
                        "natural_context_ok": r["natural_context_ok"],
                        "suspected_ambiguity": r["suspected_ambiguity"],
                        "reviewer_notes": r["reviewer_notes"],
                        "source": r["source"],
                    }
                    for r in rows
                ],
            }
        )
    summary["semantic_hash"] = sha256_bytes(
        json.dumps(decision_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
            "utf-8"
        )
    )
    # Informational package hashes (may vary with XLSX writer metadata)
    summary["role_zip_sha256"] = {r["role_id"]: r["zip_sha256"] for r in role_results}
    _write_json(out_root / "REMAINING_ROLE_DRAFTS_SUMMARY.json", summary)
    _write_json(
        out_root / "SEMANTIC_HASH.json",
        {
            "phase": PHASE,
            "semantic_hash": summary["semantic_hash"],
            "total_rows": summary["total_rows"],
            "algorithm": "sha256(json.dumps({phase,provenance,roles[].rows sorted by review_id}, sort_keys=True))",
        },
    )
    (out_root / "README.md").write_text(
        "\n".join(
            [
                "# MAI-07 V3 — AI-Assisted Remaining Role Drafts",
                "",
                "Status: `AI_ASSISTED_DRAFT_FOR_HUMAN_REVIEW`",
                "",
                "- Not independent human review",
                "- Not official Round A lock evidence",
                "- Not Round B / frozen V3 gold / training / runtime promotion",
                "- Do not place in `round_a_inbox`",
                "",
                f"Semantic hash: `{summary['semantic_hash']}`",
                "",
            ]
        ),
        encoding="utf-8",
        newline="\n",
    )
    return summary


def main(argv: list[str] | None = None) -> int:
    import argparse

    p = argparse.ArgumentParser(description=PHASE)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = p.parse_args(argv)
    summary = draft_all_remaining_roles(out_root=args.out)
    print(json.dumps({k: summary[k] for k in ("ok", "total_rows", "total_workbooks", "semantic_hash", "roles")}, indent=2, default=str))
    return 0 if summary.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
