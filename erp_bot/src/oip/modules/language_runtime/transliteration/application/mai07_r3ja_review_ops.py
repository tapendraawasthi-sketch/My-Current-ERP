"""MAI-07R3J-A automated review operations (mechanical workflow only).

Does not generate/infer/copy/alter human answers.
Does not treat AI as human reviewers.
Does not modify runtime/resources or run model evaluation.
Does not implement MAI-08.
"""

from __future__ import annotations

import hashlib
import html
import json
import re
import shutil
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .mai07_r3ja_v3_agreement import ROUND_A_DISPOSITIONS, ROUND_B_ACCEPTABILITY
from .mai07_r3ja_v3_firewall import REPO, assert_source_code_firewall
from .validate_mai07_r3ja_round_a import (
    EXPECTED_BLIND_MAPPING_SHA,
    EXPECTED_PACKET_MANIFEST_SHA,
    ROLE_FILES,
    ROUND_A_HEADERS,
    REQUIRED_SHEETS,
    validate_workbook,
    verify_sealed_authorities,
)

PHASE = "MAI-07R3J-A-AUTOMATED-REVIEW-OPERATIONS"
PACKET_ROOT = REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3"
OPS = PACKET_ROOT / "review_operations"
REVIEWERS_SRC = PACKET_ROOT / "reviewers"
BLIND_MAPPING = PACKET_ROOT / "V3_BLIND_MAPPING.json"
PACKET_MANIFEST = PACKET_ROOT / "V3_PACKET_MANIFEST.json"

BATCH_SIZE = 120
BATCH_SEED = "mai07-r3ja-review-ops-batch-20260716"

STATES = (
    "PACKET_VERIFIED",
    "ROUND_A_PACKAGES_READY",
    "WAITING_FOR_ROUND_A_SUBMISSIONS",
    "ROUND_A_CORRECTION_REQUIRED",
    "ROUND_A_LOCKED",
    "ROUND_B_PACKAGES_READY",
    "WAITING_FOR_ROUND_B_SUBMISSIONS",
    "ROUND_B_CORRECTION_REQUIRED",
    "ROUND_B_LOCKED",
    "AGREEMENT_ANALYZED",
    "ADJUDICATION_REQUIRED",
    "WAITING_FOR_ADJUDICATION",
    "ADJUDICATION_CORRECTION_REQUIRED",
    "REVIEW_COMPLETE_READY_FOR_R3J_B",
)

ROUND_A_ROLES = (
    "PRODUCT_POLICY",
    "NEPALI_FLUENT_A",
    "PROFESSIONAL_LINGUIST_B",
    "ACCOUNTING_DOMAIN",
)

ROLE_DISPLAY = {
    "PRODUCT_POLICY": "Product Policy Reviewer",
    "NEPALI_FLUENT_A": "Nepali-Fluent Reviewer A",
    "PROFESSIONAL_LINGUIST_B": "Professional Nepali Linguist Reviewer B",
    "ACCOUNTING_DOMAIN": "Nepal Accounting/Business Domain Reviewer",
    "INDEPENDENT_ADJUDICATOR": "Independent Adjudicator",
}

FORBIDDEN_NAME = re.compile(r"[\x00-\x1f]")
CSV_INJECT = re.compile(r"^[=+\-@]")
CONFIDENCE = ("HIGH", "MEDIUM", "LOW")


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _sha_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _write_json(path: Path, obj: Any) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    path.write_text(text, encoding="utf-8", newline="\n")
    return _sha(path)


def _write_text(path: Path, text: str) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8", newline="\n")
    return _sha(path)


def _safe_filename(name: str) -> str:
    base_only = Path(name).name
    if base_only.lower().endswith((".xlsm", ".xltm")) or name.lower().endswith((".xlsm", ".xltm")):
        raise ValueError(f"MACRO_ENABLED_FILE_REJECTED:{name}")
    if (
        "/" in name
        or "\\" in name
        or ".." in name
        or FORBIDDEN_NAME.search(name)
        or base_only != name
    ):
        raise ValueError(f"PATH_TRAVERSAL_OR_FORBIDDEN_FILENAME:{name}")
    return base_only


def ensure_ops_dirs() -> None:
    for rel in (
        "",
        "reviewer_packages",
        "round_a_inbox",
        "round_a_locked",
        "round_b_packages",
        "round_b_inbox",
        "round_b_locked",
        "adjudication_package",
        "adjudication_inbox",
        "final_locked_reviews",
        "validation_reports",
        "agreement_reports",
        "hash_manifests",
        "scripts",
    ):
        (OPS / rel).mkdir(parents=True, exist_ok=True)


def load_status() -> dict[str, Any]:
    path = OPS / "REVIEW_STATUS.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {
        "schema_version": "1.0.0",
        "phase": PHASE,
        "state": "PACKET_VERIFIED",
        "QUALITY_GATES_PASSED": False,
        "LINGUIST_APPROVED": False,
        "PRODUCTION_APPROVED": False,
        "MAI_08": "NOT_STARTED",
        "human_blocker": "BLOCKED_PENDING_INDEPENDENT_HUMAN_REVIEW",
        "credential_status": "CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION",
        "human_answers_generated": False,
        "model_evaluation_performed": False,
        "roles": {},
    }


def save_status(status: dict[str, Any]) -> str:
    status = dict(status)
    status["updated_utc"] = datetime.now(timezone.utc).isoformat()
    status["packet_manifest_sha256"] = EXPECTED_PACKET_MANIFEST_SHA
    status["blind_mapping_sha256"] = EXPECTED_BLIND_MAPPING_SHA
    return _write_json(OPS / "REVIEW_STATUS.json", status)


def stabilize_xlsx(path: Path) -> None:
    import io

    fixed_core = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        b'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        b'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        b"<dc:creator>MAI-07R3J-A-REVIEW-OPS</dc:creator>"
        b'<dcterms:created xsi:type="dcterms:W3CDTF">2026-07-16T00:00:00Z</dcterms:created>'
        b'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-07-16T00:00:00Z</dcterms:modified>'
        b"<cp:lastModifiedBy>MAI-07R3J-A-REVIEW-OPS</cp:lastModifiedBy>"
        b"</cp:coreProperties>"
    )
    raw = path.read_bytes()
    zin = zipfile.ZipFile(io.BytesIO(raw), "r")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in sorted(zin.infolist(), key=lambda i: i.filename):
            payload = fixed_core if info.filename == "docProps/core.xml" else zin.read(info.filename)
            new = zipfile.ZipInfo(filename=info.filename, date_time=(2026, 7, 16, 0, 0, 0))
            new.compress_type = zipfile.ZIP_DEFLATED
            new.create_system = 0
            zout.writestr(new, payload)
    zin.close()
    path.write_bytes(buf.getvalue())


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1E2433")
    font = Font(color="FFFFFF", bold=True, size=10)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True)


def _csv_cell(value: Any) -> str:
    s = "" if value is None else str(value)
    if CSV_INJECT.match(s):
        return "'" + s
    return s


def read_round_a_rows(src_xlsx: Path) -> list[list[Any]]:
    wb = load_workbook(src_xlsx, read_only=True, data_only=True)
    ws = wb["ROUND_A_CONTEXT"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = tuple(str(c or "") for c in rows[0])
    if header != ROUND_A_HEADERS:
        raise ValueError(f"header_mismatch:{src_xlsx.name}")
    out: list[list[Any]] = []
    for raw in rows[1:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        # Force blank answer fields — never carry filled answers into packages
        row = [
            raw[0],
            _csv_cell(raw[1]),
            _csv_cell(raw[2]),
            "",  # disposition blank
            "",  # confidence blank
            "",
            "",
            "",
            "",
        ]
        out.append(row)
    return out


def deterministic_batches(rows: list[list[Any]], role_id: str) -> list[list[list[Any]]]:
    """Split rows into batches of BATCH_SIZE; last batch may be smaller (>=1)."""
    # Stable order already by opaque review_id from source packet
    batches: list[list[list[Any]]] = []
    for i in range(0, len(rows), BATCH_SIZE):
        batches.append(rows[i : i + BATCH_SIZE])
    # Reconcile coverage
    flat = [r[0] for b in batches for r in b]
    if len(flat) != len(rows) or len(set(flat)) != len(flat):
        raise RuntimeError(f"batch_coverage_failed:{role_id}")
    return batches


def build_batch_workbook(
    *,
    role_id: str,
    batch_index: int,
    batch_total: int,
    rows: list[list[Any]],
    dest: Path,
) -> str:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "START_HERE"
    for i, line in enumerate(
        [
            f"MAI-07 V3 Round A — {ROLE_DISPLAY[role_id]}",
            f"Batch {batch_index} of {batch_total}",
            f"Assigned rows in this batch: {len(rows)}",
            "",
            "Edit only: REVIEWER_DECLARATION (once per role), ROUND_A_CONTEXT.",
            "Do not change review_id values or sheet names.",
            "Leave answer cells blank until you decide; do not use AI autofill.",
            "Save as .xlsx (no macros). Return this batch file to the coordinator inbox.",
            "Do not begin Round B until Round A is locked.",
        ],
        start=1,
    ):
        ws0.cell(i, 1, line)
    ws0.column_dimensions["A"].width = 100

    ws1 = wb.create_sheet("REVIEWER_DECLARATION")
    for row in [
        ["field", "value"],
        ["reviewer_full_name", ""],
        ["reviewer_email", ""],
        ["role_id", role_id],
        ["contact_reference_id", ""],
        ["qualification_summary", ""],
        ["professional_linguist_credentials", ""],
        ["relevant_experience", ""],
        ["declaration_independent", ""],
        ["declaration_no_conflict_of_interest", ""],
        ["declaration_answers_are_own", ""],
        ["declaration_did_not_see_other_reviewers", ""],
        ["declaration_no_ai_autofill", ""],
        ["declaration_no_runtime_predictions_used", ""],
        ["declaration_date_utc", ""],
        ["signature_or_typed_name", ""],
        ["credential_status", "CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION"],
    ]:
        ws1.append(row)
    _style_header(ws1, 2)

    wsA = wb.create_sheet("ROUND_A_CONTEXT")
    wsA.append(list(ROUND_A_HEADERS))
    _style_header(wsA, len(ROUND_A_HEADERS))
    for r in rows:
        wsA.append(list(r))
    dv_disp = DataValidation(
        type="list",
        formula1='"' + ",".join(ROUND_A_DISPOSITIONS) + '"',
        allow_blank=True,
    )
    dv_conf = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=True)
    wsA.add_data_validation(dv_disp)
    wsA.add_data_validation(dv_conf)
    if rows:
        dv_disp.add(f"D2:D{len(rows)+1}")
        dv_conf.add(f"E2:E{len(rows)+1}")
    for i in range(1, 10):
        wsA.column_dimensions[get_column_letter(i)].width = 18 if i != 2 else 55

    # Placeholder Round B sheet — locked closed until Round A lock (blank instruction only)
    wsB = wb.create_sheet("ROUND_B_CANDIDATES")
    wsB.append(
        [
            "review_id",
            "input_text",
            "highlighted_span",
            "candidate_index",
            "candidate_surface",
            "acceptability",
            "reviewer_entered_candidate",
            "notes",
        ]
    )
    _style_header(wsB, 8)
    wsB.append(["(Round B opens only after Round A lock — do not fill this sheet yet)", "", "", "", "", "", "", ""])

    wsP = wb.create_sheet("REVIEW_PROGRESS")
    wsP.append(["metric", "value"])
    wsP.append(["batch_index", batch_index])
    wsP.append(["batch_total", batch_total])
    wsP.append(["rows_in_batch", len(rows)])
    wsP.append(["dispositions_completed", f"=COUNTA(ROUND_A_CONTEXT!D2:D{len(rows)+1})"])
    _style_header(wsP, 2)

    wsE = wb.create_sheet("VALIDATION_ERRORS")
    wsE.append(["error_code", "detail"])
    wsE.append(["(empty until steward validation)", ""])
    _style_header(wsE, 2)

    wsC = wb.create_sheet("SUBMISSION_CHECKLIST")
    for line in [
        ["checklist_item", "done_Y_N"],
        ["declaration_completed_for_role", ""],
        ["all_batch_rows_labeled_or_abstained", ""],
        ["no_ai_autofill", ""],
        ["did_not_see_other_reviewers", ""],
        ["saved_as_xlsx_no_macros", ""],
        ["returned_to_inbox", ""],
    ]:
        wsC.append(line)
    _style_header(wsC, 2)

    # Ensure sheet order matches required names (subset + Round B placeholder)
    dest.parent.mkdir(parents=True, exist_ok=True)
    from datetime import datetime as dt
    from datetime import timezone as tz

    fixed = dt(2026, 7, 16, 0, 0, 0, tzinfo=tz.utc)
    wb.properties.creator = "MAI-07R3J-A-REVIEW-OPS"
    wb.properties.lastModifiedBy = "MAI-07R3J-A-REVIEW-OPS"
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.save(dest)
    stabilize_xlsx(dest)
    return _sha(dest)


def write_role_instructions(role_id: str, dest: Path, batch_total: int, row_total: int) -> str:
    text = f"""# Instructions — {ROLE_DISPLAY[role_id]}

Role ID: `{role_id}`
Round: A (context / disposition)
Batches: {batch_total}
Assigned rows: {row_total}

## What you receive
- This instructions file
- Declaration fields inside each batch workbook (complete once)
- Batch workbooks with your assigned opaque review IDs only

## What you must NOT receive / use
- Blind mapping
- Other reviewers' packages
- Runtime predictions or expected answers
- Population / pool labels
- Failure history

## How to complete
1. Open batch 01 first; complete REVIEWER_DECLARATION.
2. For every row: set disposition (allowed enum) and confidence (HIGH/MEDIUM/LOW).
3. ABSTAIN_CANNOT_DECIDE is allowed; it is not a pass.
4. Save as `.xlsx` without macros.
5. Return each completed batch to the coordinator Round A inbox.

## Allowed Round A dispositions
{chr(10).join('- ' + d for d in ROUND_A_DISPOSITIONS)}

## Credential note
Automation records:
`CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION`
Professional linguist approval cannot be inferred until a human coordinator verifies credentials.
"""
    return _write_text(dest, text)


def package_role_round_a(role_id: str) -> dict[str, Any]:
    src = REVIEWERS_SRC / ROLE_FILES[role_id]
    if not src.exists():
        raise FileNotFoundError(src)
    rows = read_round_a_rows(src)
    batches = deterministic_batches(rows, role_id)
    role_dir = OPS / "reviewer_packages" / role_id / "round_a"
    if role_dir.exists():
        shutil.rmtree(role_dir)
    role_dir.mkdir(parents=True)
    batch_meta = []
    all_ids: list[str] = []
    for i, batch_rows in enumerate(batches, start=1):
        fname = f"MokXya_MAI07_V3__{role_id}__ROUND_A__BATCH_{i:02d}_of_{len(batches):02d}.xlsx"
        path = role_dir / fname
        sha = build_batch_workbook(
            role_id=role_id,
            batch_index=i,
            batch_total=len(batches),
            rows=batch_rows,
            dest=path,
        )
        ids = [str(r[0]) for r in batch_rows]
        all_ids.extend(ids)
        batch_meta.append(
            {
                "batch_index": i,
                "batch_total": len(batches),
                "filename": fname,
                "row_count": len(batch_rows),
                "review_ids": ids,
                "sha256": sha,
            }
        )
    if len(all_ids) != len(set(all_ids)):
        raise RuntimeError(f"duplicate_ids_in_batches:{role_id}")
    write_role_instructions(
        role_id,
        role_dir / "INSTRUCTIONS.md",
        len(batches),
        len(rows),
    )
    _write_text(
        role_dir / "SAVE_AND_RETURN.md",
        f"""# Save and return — {role_id}

1. Complete all batches ({len(batches)}).
2. Save each as `.xlsx` (no `.xlsm`).
3. Place completed batch files into:
   `docs/mokxya-ai/reviews/mai07_v3/review_operations/round_a_inbox/{role_id}/`
4. Do not rename batch files.
5. Do not include blind mapping or other roles' files.
""",
    )
    _write_text(
        role_dir / "PROGRESS_CHECKLIST.md",
        f"""# Progress checklist — {role_id}

- [ ] Declaration completed
- [ ] Batch files completed: 0 / {len(batches)}
- [ ] No AI autofill
- [ ] Did not see other reviewers' answers
- [ ] Returned to round_a_inbox/{role_id}/
""",
    )
    manifest = {
        "role_id": role_id,
        "round": "A",
        "source_workbook": ROLE_FILES[role_id],
        "source_sha256": _sha(src),
        "batch_size_target": BATCH_SIZE,
        "batch_count": len(batches),
        "row_total": len(rows),
        "batches": batch_meta,
        "all_review_ids_sha256": _sha_bytes(
            json.dumps(sorted(all_ids), separators=(",", ":")).encode("utf-8")
        ),
        "blank_answers": True,
        "prohibited_for_training": True,
    }
    man_sha = _write_json(role_dir / "BATCH_MANIFEST.json", manifest)

    # ZIP package
    zip_path = OPS / "reviewer_packages" / f"MokXya_MAI07_V3_ROUND_A_PACKAGE__{role_id}.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in sorted(role_dir.rglob("*")):
            if p.is_file():
                # reject accidental mapping
                if "blind" in p.name.lower() or "mapping" in p.name.lower():
                    raise RuntimeError(f"forbidden_file_in_package:{p}")
                arc = f"{role_id}/round_a/{p.relative_to(role_dir).as_posix()}"
                info = zipfile.ZipInfo(arc, date_time=(2026, 7, 16, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                zf.writestr(info, p.read_bytes())
    # Ensure no blind mapping content in zip
    with zipfile.ZipFile(zip_path, "r") as zf:
        names = zf.namelist()
        if any("blind" in n.lower() or "mapping" in n.lower() for n in names):
            raise RuntimeError("blind_mapping_leaked_into_zip")
    return {
        "role_id": role_id,
        "batch_count": len(batches),
        "row_total": len(rows),
        "batch_manifest_sha256": man_sha,
        "zip_path": str(zip_path.relative_to(REPO)).replace("\\", "/"),
        "zip_sha256": _sha(zip_path),
        "inbox_path": str((OPS / "round_a_inbox" / role_id).relative_to(REPO)).replace("\\", "/"),
    }


def write_start_here() -> None:
    _write_text(
        OPS / "START_HERE.md",
        """# MAI-07 V3 Review Operations — START HERE

## What this automates
Mechanical packaging, validation, locking, agreement reports, and adjudication packaging.

## What only humans do
1. Choose real reviewers
2. Send role-specific ZIPs
3. Receive completed packages
4. Place returns in the role inbox folders
5. Verify reviewer identity/credentials manually
6. Send disagreement packet to the adjudicator (after Round B lock + agreement)

## One-click
- `RUN_REVIEW_WORKFLOW.bat` — advance workflow safely
- `CHECK_REVIEW_STATUS.bat` — print / open status

## Critical rules
- AI does **not** fill answers
- Round B never releases before Round A lock
- `LINGUIST_APPROVED` stays false until full professional workflow + manual credential verification
- Do not edit blind mapping
""",
    )
    _write_text(
        OPS / "START_HERE_NP.md",
        """# MAI-07 V3 समीक्षा सञ्चालन — यहाँबाट सुरु गर्नुहोस्

## मानवले मात्र गर्ने
1. वास्तविक reviewer छान्नुहोस्
2. भूमिकाअनुसार ZIP पठाउनुहोस्
3. पूरा प्याकेज फर्काउनुहोस्
4. inbox फोल्डरमा राख्नुहोस्
5. परिचय/योग्यता म्यानुअली प्रमाणित गर्नुहोस्
6. असहमति प्याकेट adjudicator लाई पठाउनुहोस् (Round B lock पछि)

AI ले उत्तर भर्दैन। Round A lock अघि Round B खुल्दैन।
""",
    )


def write_bat_scripts() -> None:
    run_bat = r"""@echo off
setlocal EnableExtensions
REM MAI-07 V3 Review Operations — one-click workflow (Windows)
cd /d "%~dp0"
REM Ascend to repository root (review_operations -> mai07_v3 -> reviews -> mokxya-ai -> docs -> repo)
cd ..\..\..\..\..
if not exist "erp_bot\src\oip\modules\language_runtime\transliteration\application\mai07_r3ja_review_ops.py" (
  echo ERROR: repository root not detected from script location.
  exit /b 2
)
set PYTHONPATH=erp_bot\src
where python >nul 2>&1
if errorlevel 1 (
  echo ERROR: python not found on PATH.
  exit /b 2
)
python -m src.oip.modules.language_runtime.transliteration.application.mai07_r3ja_review_ops --run
set EXITCODE=%ERRORLEVEL%
echo.
echo Status file: docs\mokxya-ai\reviews\mai07_v3\review_operations\REVIEW_STATUS.json
echo Dashboard:   docs\mokxya-ai\reviews\mai07_v3\review_operations\REVIEW_STATUS.html
exit /b %EXITCODE%
"""
    check_bat = r"""@echo off
setlocal EnableExtensions
cd /d "%~dp0"
cd ..\..\..\..\..
if not exist "erp_bot\src\oip\modules\language_runtime\transliteration\application\mai07_r3ja_review_ops.py" (
  echo ERROR: repository root not detected.
  exit /b 2
)
set PYTHONPATH=erp_bot\src
python -m src.oip.modules.language_runtime.transliteration.application.mai07_r3ja_review_ops --status
if exist "docs\mokxya-ai\reviews\mai07_v3\review_operations\REVIEW_STATUS.html" (
  start "" "docs\mokxya-ai\reviews\mai07_v3\review_operations\REVIEW_STATUS.html"
)
exit /b 0
"""
    _write_text(OPS / "RUN_REVIEW_WORKFLOW.bat", run_bat.replace("\n", "\r\n"))
    _write_text(OPS / "CHECK_REVIEW_STATUS.bat", check_bat.replace("\n", "\r\n"))
    # Also copy into scripts/
    shutil.copy2(OPS / "RUN_REVIEW_WORKFLOW.bat", OPS / "scripts" / "RUN_REVIEW_WORKFLOW.bat")
    shutil.copy2(OPS / "CHECK_REVIEW_STATUS.bat", OPS / "scripts" / "CHECK_REVIEW_STATUS.bat")


def render_dashboard(status: dict[str, Any]) -> str:
    roles = status.get("roles") or {}
    rows = []
    for role_id, info in sorted(roles.items()):
        rows.append(
            "<tr>"
            f"<td>{html.escape(role_id)}</td>"
            f"<td>{html.escape(str(info.get('batch_count', '')))}</td>"
            f"<td>{html.escape(str(info.get('batches_returned', 0)))}</td>"
            f"<td>{html.escape(str(info.get('completion_pct', '0%')))}</td>"
            f"<td>{html.escape(str(info.get('declaration_status', 'pending')))}</td>"
            f"<td>{html.escape(str(info.get('credential_status', 'CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION')))}</td>"
            "</tr>"
        )
    body = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"/>
<title>MAI-07 V3 Review Status</title>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:24px;background:#f5f6fa;color:#1e2433}}
h1{{font-size:18px}} h2{{font-size:14px;margin-top:24px}}
.card{{background:#fff;border:1px solid #d1d5db;padding:16px;margin:12px 0}}
table{{border-collapse:collapse;width:100%;font-size:12px}}
th,td{{border:1px solid #e5e7eb;padding:8px;text-align:left}}
th{{background:#1e2433;color:#fff}}
.badge{{display:inline-block;padding:4px 8px;background:#eef2ff;border:1px solid #c7d2fe;font-size:12px}}
.warn{{background:#fff7ed;border-color:#fdba74}}
</style></head><body>
<h1>MAI-07 V3 Review Operations</h1>
<div class="card"><span class="badge">State: {html.escape(str(status.get('state')))}</span>
<span class="badge warn">Human blocker: {html.escape(str(status.get('human_blocker')))}</span>
<p>QUALITY_GATES_PASSED={status.get('QUALITY_GATES_PASSED')} ·
LINGUIST_APPROVED={status.get('LINGUIST_APPROVED')} ·
MAI_08={html.escape(str(status.get('MAI_08')))}</p>
<p><b>Next action:</b> {html.escape(str(status.get('next_human_action','')))}</p>
</div>
<div class="card"><h2>Round A lock / Round B</h2>
<p>ROUND_A_LOCKED={status.get('ROUND_A_LOCKED', False)} ·
ROUND_B_READY={status.get('ROUND_B_READY', False)} ·
ROUND_B_LOCKED={status.get('ROUND_B_LOCKED', False)}</p>
<p>Credential status: {html.escape(str(status.get('credential_status')))}</p>
</div>
<div class="card"><h2>Reviewer progress</h2>
<table><thead><tr>
<th>Role</th><th>Batches</th><th>Returned</th><th>Completion</th><th>Declaration</th><th>Credentials</th>
</tr></thead><tbody>
{''.join(rows) if rows else '<tr><td colspan="6">No roles packaged yet</td></tr>'}
</tbody></table></div>
<div class="card"><h2>Security</h2>
<p>Blind mapping is not shown. Macro-enabled files are rejected. No answers are auto-filled.</p>
</div>
</body></html>
"""
    return _write_text(OPS / "REVIEW_STATUS.html", body)


def detect_returned_batches(role_id: str) -> list[Path]:
    inbox = OPS / "round_a_inbox" / role_id
    inbox.mkdir(parents=True, exist_ok=True)
    files = []
    for p in sorted(inbox.glob("*.xlsx")):
        name = _safe_filename(p.name)
        if name.lower().endswith((".xlsm", ".xltm")):
            raise ValueError(f"MACRO_ENABLED_FILE_REJECTED:{name}")
        upper = name.upper()
        if "AI_ASSISTED" in upper or "__AI_ASSISTED_DRAFT" in upper:
            raise ValueError(
                "AI_ASSISTED_ARTIFACT_FORBIDDEN_IN_OFFICIAL_INBOX:"
                f"{role_id}:{name}"
            )
        files.append(inbox / name)
    return files


def write_round_a_missing_reports(status: dict[str, Any]) -> dict[str, Any]:
    """Write exact per-role missing-batch reports without inspecting/creating answers."""
    summary: dict[str, Any] = {
        "status": "WAITING_FOR_ROUND_A_SUBMISSIONS",
        "human_answers_generated": False,
        "human_answers_altered": False,
        "roles": {},
    }
    for role_id in ROUND_A_ROLES:
        manifest_path = (
            OPS / "reviewer_packages" / role_id / "round_a" / "BATCH_MANIFEST.json"
        )
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        returned = {p.name for p in detect_returned_batches(role_id)}
        expected = [b["filename"] for b in manifest["batches"]]
        missing = [name for name in expected if name not in returned]
        missing_rows = sum(
            b["row_count"] for b in manifest["batches"] if b["filename"] in missing
        )
        role_report = {
            "role_id": role_id,
            "state": "WAITING_FOR_ROUND_A_SUBMISSIONS" if not returned else "ROUND_A_CORRECTION_REQUIRED",
            "expected_batch_count": len(expected),
            "returned_batch_count": len(returned),
            "missing_batch_count": len(missing),
            "expected_row_count": manifest["row_total"],
            "missing_row_count": missing_rows,
            "missing_files": missing,
            "unexpected_files": sorted(returned - set(expected)),
            "declaration_status": status["roles"].get(role_id, {}).get(
                "declaration_status", "pending_submission"
            ),
            "credential_status": "CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION",
            "correction_requirements": [
                "Return every listed batch using its exact filename.",
                "Complete the signed reviewer declaration.",
                "Complete every assigned disposition and confidence cell using allowed enums.",
                "Do not rename sheets, headers, or review IDs.",
                "Save as .xlsx without formulas or macros.",
            ],
            "inbox_path": status["roles"].get(role_id, {}).get("inbox_path"),
            "human_answers_generated": False,
            "human_answers_altered": False,
        }
        if role_id == "PROFESSIONAL_LINGUIST_B":
            role_report["correction_requirements"].append(
                "Declare professional-linguist qualifications; manual credential verification remains required."
            )
        _write_json(
            OPS / "validation_reports" / f"ROUND_A_MISSING__{role_id}.json",
            role_report,
        )
        summary["roles"][role_id] = role_report
    _write_json(
        OPS / "validation_reports" / "ROUND_A_MISSING_SUBMISSIONS_SUMMARY.json",
        summary,
    )
    return summary


def initialize_packages() -> dict[str, Any]:
    """Verify seals, build Round A packages/ZIPs, set WAITING_FOR_ROUND_A_SUBMISSIONS."""
    ensure_ops_dirs()
    # Firewall on this module
    viol = assert_source_code_firewall(Path(__file__))
    if viol:
        raise RuntimeError(f"firewall:{viol}")
    sealed = verify_sealed_authorities()
    if not sealed["ok"]:
        raise RuntimeError(f"sealed_authority_failed:{sealed['errors']}")

    write_start_here()
    write_bat_scripts()

    role_results = {}
    for role_id in ROUND_A_ROLES:
        (OPS / "round_a_inbox" / role_id).mkdir(parents=True, exist_ok=True)
        role_results[role_id] = package_role_round_a(role_id)

    hashes = {
        "packet_manifest_sha256": sealed["packet_manifest_sha256"],
        "blind_mapping_sha256": sealed["blind_mapping_sha256"],
        "role_packages": {
            k: {"zip_sha256": v["zip_sha256"], "batch_manifest_sha256": v["batch_manifest_sha256"]}
            for k, v in role_results.items()
        },
    }
    hash_sha = _write_json(OPS / "hash_manifests" / "ROUND_A_PACKAGE_HASH_MANIFEST.json", hashes)

    status = load_status()
    status.update(
        {
            "state": "WAITING_FOR_ROUND_A_SUBMISSIONS",
            "prior_states": ["PACKET_VERIFIED", "ROUND_A_PACKAGES_READY", "WAITING_FOR_ROUND_A_SUBMISSIONS"],
            "MAI_07R3J_A_REVIEW_OPS": "PASSED_AUTOMATION",
            "human_blocker": "BLOCKED_PENDING_INDEPENDENT_HUMAN_REVIEW",
            "QUALITY_GATES_PASSED": False,
            "LINGUIST_APPROVED": False,
            "PRODUCTION_APPROVED": False,
            "MAI_08": "NOT_STARTED",
            "ROUND_A_LOCKED": False,
            "ROUND_B_READY": False,
            "ROUND_B_LOCKED": False,
            "credential_status": "CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION",
            "human_answers_generated": False,
            "model_evaluation_performed": False,
            "package_hash_manifest_sha256": hash_sha,
            "roles": {
                rid: {
                    "batch_count": info["batch_count"],
                    "row_total": info["row_total"],
                    "batches_returned": 0,
                    "completion_pct": "0%",
                    "declaration_status": "pending_submission",
                    "credential_status": "CREDENTIALS_DECLARED_PENDING_MANUAL_VERIFICATION",
                    "zip_path": info["zip_path"],
                    "inbox_path": info["inbox_path"],
                }
                for rid, info in role_results.items()
            },
            "next_human_action": (
                "Send each Round A ZIP to the chosen real reviewer. "
                "Place completed batch .xlsx files into review_operations/round_a_inbox/<ROLE>/ "
                "then run RUN_REVIEW_WORKFLOW.bat."
            ),
            "adjudicator_package_ready": False,
        }
    )
    save_status(status)
    render_dashboard(status)
    return {
        "status": "PASSED_AUTOMATION",
        "state": status["state"],
        "sealed": sealed,
        "roles": role_results,
        "dashboard": str((OPS / "REVIEW_STATUS.html").relative_to(REPO)).replace("\\", "/"),
    }


def advance_workflow() -> dict[str, Any]:
    """Detect state and advance only when gates pass. Never invent answers."""
    ensure_ops_dirs()
    sealed = verify_sealed_authorities()
    if not sealed["ok"]:
        return {"status": "BLOCKED_PACKET_INTEGRITY_FAILED", "sealed": sealed}

    status = load_status()
    state = status.get("state", "PACKET_VERIFIED")

    # If packages not built yet, initialize
    zips_exist = all(
        (OPS / "reviewer_packages" / f"MokXya_MAI07_V3_ROUND_A_PACKAGE__{r}.zip").exists()
        for r in ROUND_A_ROLES
    )
    if not zips_exist or state in {"PACKET_VERIFIED", "ROUND_A_PACKAGES_READY"}:
        return initialize_packages()

    # Waiting for Round A — validate any returns; do not skip to Round B
    if state in {"WAITING_FOR_ROUND_A_SUBMISSIONS", "ROUND_A_CORRECTION_REQUIRED"}:
        any_files = False
        for role_id in ROUND_A_ROLES:
            files = detect_returned_batches(role_id)
            status["roles"].setdefault(role_id, {})["batches_returned"] = len(files)
            if files:
                any_files = True
                # Validate each batch file mechanically (identity/path only at batch level;
                # full role merge validation happens when all batches present)
                for f in files:
                    _safe_filename(f.name)
            expected = status["roles"].get(role_id, {}).get("batch_count") or 0
            returned = status["roles"].get(role_id, {}).get("batches_returned") or 0
            pct = int(100 * returned / expected) if expected else 0
            status["roles"][role_id]["completion_pct"] = f"{pct}%"

        if not any_files:
            status["state"] = "WAITING_FOR_ROUND_A_SUBMISSIONS"
            missing_report = write_round_a_missing_reports(status)
            status["next_human_action"] = (
                "No Round A submissions in inbox yet. Send ZIPs to reviewers; "
                "place completed batches under round_a_inbox/<ROLE>/."
            )
            save_status(status)
            render_dashboard(status)
            return {
                "status": "WAITING_FOR_ROUND_A_SUBMISSIONS",
                "state": status["state"],
                "roles": status["roles"],
                "missing_report": missing_report,
            }

        # Partial returns → correction/waiting, never Round B
        incomplete = [
            rid
            for rid in ROUND_A_ROLES
            if (status["roles"].get(rid, {}).get("batches_returned") or 0)
            < (status["roles"].get(rid, {}).get("batch_count") or 0)
        ]
        if incomplete:
            status["state"] = "ROUND_A_CORRECTION_REQUIRED"
            status["next_human_action"] = (
                "Incomplete Round A returns for: " + ", ".join(incomplete) + ". "
                "Return remaining batches; do not release Round B."
            )
            # Write correction report
            report = {
                "status": "ROUND_A_CORRECTION_REQUIRED",
                "incomplete_roles": incomplete,
                "note": "No human answers were altered.",
            }
            _write_json(OPS / "validation_reports" / "ROUND_A_CORRECTION_REQUIRED.json", report)
            save_status(status)
            render_dashboard(status)
            return report

        # All batches present for all roles — attempt merge + full validation via existing validator path
        # For now require merged files in legacy inbox OR validate batch completeness only and
        # instruct merge step. Auto-merge batches into canonical workbook for validator.
        merge_errors = []
        for role_id in ROUND_A_ROLES:
            try:
                merge_round_a_batches(role_id)
            except Exception as exc:  # noqa: BLE001
                merge_errors.append(f"{role_id}:{exc}")
        if merge_errors:
            status["state"] = "ROUND_A_CORRECTION_REQUIRED"
            status["next_human_action"] = "Fix Round A merge/validation errors: " + "; ".join(merge_errors)
            save_status(status)
            render_dashboard(status)
            return {"status": "ROUND_A_CORRECTION_REQUIRED", "errors": merge_errors}

        # Delegate to existing Round A validator using merged files in ops round_a_inbox root copies
        from .validate_mai07_r3ja_round_a import INBOX as LEGACY_INBOX
        from .validate_mai07_r3ja_round_a import run_round_a_validation

        # Copy merged canonical files into legacy inbox expected by validator
        LEGACY_INBOX.mkdir(parents=True, exist_ok=True)
        for role_id, fname in ROLE_FILES.items():
            src = OPS / "round_a_inbox" / role_id / fname
            if src.exists():
                shutil.copy2(src, LEGACY_INBOX / fname)
        result = run_round_a_validation(inbox=LEGACY_INBOX)
        if result.get("status") != "ROUND_B_READY":
            status["state"] = "ROUND_A_CORRECTION_REQUIRED"
            status["next_human_action"] = result.get("next_human_action") or "Correct Round A submissions."
            _write_json(OPS / "validation_reports" / "ROUND_A_VALIDATION_SUMMARY.json", result)
            save_status(status)
            render_dashboard(status)
            return result

        status["state"] = "ROUND_B_PACKAGES_READY"
        status["ROUND_A_LOCKED"] = True
        status["ROUND_B_READY"] = True
        # Copy Round B packages into ops folder
        prepare_round_b_packages()
        status["state"] = "WAITING_FOR_ROUND_B_SUBMISSIONS"
        status["next_human_action"] = (
            "Round A locked. Send Round B packages from round_b_packages/ to eligible reviewers."
        )
        save_status(status)
        render_dashboard(status)
        return {"status": "WAITING_FOR_ROUND_B_SUBMISSIONS", "round_a_lock": result}

    # Later states: keep waiting / report only until submissions exist
    save_status(status)
    render_dashboard(status)
    return {"status": status.get("state"), "state": status.get("state"), "note": "No further auto-advance without submissions."}


def merge_round_a_batches(role_id: str) -> Path:
    """Merge returned batch xlsx into canonical ROLE_FILES name. Does not invent answers."""
    inbox = OPS / "round_a_inbox" / role_id
    man = json.loads(
        (OPS / "reviewer_packages" / role_id / "round_a" / "BATCH_MANIFEST.json").read_text(
            encoding="utf-8"
        )
    )
    expected_batches = man["batch_count"]
    files = detect_returned_batches(role_id)
    if len(files) != expected_batches:
        raise RuntimeError(f"batch_count_mismatch:{len(files)}!={expected_batches}")

    # Read rows from each batch, keyed by review_id
    by_id: dict[str, list[Any]] = {}
    for f in files:
        wb = load_workbook(f, read_only=True, data_only=True)
        if "ROUND_A_CONTEXT" not in wb.sheetnames:
            wb.close()
            raise RuntimeError(f"missing_ROUND_A_CONTEXT:{f.name}")
        rows = list(wb["ROUND_A_CONTEXT"].iter_rows(values_only=True))
        wb.close()
        for raw in rows[1:]:
            if not raw or raw[0] is None:
                continue
            rid = str(raw[0]).strip()
            # Preserve human answers exactly as submitted (no alteration)
            by_id[rid] = [raw[i] if i < len(raw) else "" for i in range(9)]

    expected_ids = []
    for b in man["batches"]:
        expected_ids.extend(b["review_ids"])
    if set(by_id) != set(expected_ids):
        missing = sorted(set(expected_ids) - set(by_id))
        extra = sorted(set(by_id) - set(expected_ids))
        raise RuntimeError(f"id_set_mismatch:missing={len(missing)}:extra={len(extra)}")

    # Build merged workbook with human values preserved
    ordered = [by_id[i] for i in expected_ids]
    # Convert to list rows with escaped text cells but preserve disposition/confidence as-is
    rows_out = []
    for raw in ordered:
        rows_out.append(
            [
                raw[0],
                _csv_cell(raw[1]),
                _csv_cell(raw[2]),
                "" if raw[3] is None else str(raw[3]).strip(),
                "" if raw[4] is None else str(raw[4]).strip(),
                "" if raw[5] is None else str(raw[5]).strip(),
                "" if raw[6] is None else str(raw[6]).strip(),
                "" if raw[7] is None else str(raw[7]).strip(),
                _csv_cell(raw[8]) if raw[8] is not None else "",
            ]
        )
    dest = inbox / ROLE_FILES[role_id]
    build_batch_workbook(
        role_id=role_id,
        batch_index=1,
        batch_total=1,
        rows=[[r[0], r[1], r[2], "", "", "", "", "", ""] for r in rows_out],
        dest=dest,
    )
    # Re-open and write actual human dispositions without changing them
    wb = load_workbook(dest)
    ws = wb["ROUND_A_CONTEXT"]
    for i, r in enumerate(rows_out, start=2):
        ws.cell(i, 4, r[3])
        ws.cell(i, 5, r[4])
        ws.cell(i, 6, r[5])
        ws.cell(i, 7, r[6])
        ws.cell(i, 8, r[7])
        ws.cell(i, 9, r[8])
    wb.save(dest)
    stabilize_xlsx(dest)
    # Preserve originals already in inbox; copy merge to locked staging later
    return dest


def prepare_round_b_packages() -> dict[str, str]:
    """Copy sealed Round B templates into round_b_packages after Round A lock only."""
    out: dict[str, str] = {}
    for role_id in ("NEPALI_FLUENT_A", "PROFESSIONAL_LINGUIST_B", "ACCOUNTING_DOMAIN"):
        stub = ROLE_FILES[role_id].replace("MokXya_MAI07_V3__", "").replace(".xlsx", "")
        csv_src = REVIEWERS_SRC / f"V3_ROUND_B_TEMPLATE__{stub}.csv"
        xlsx_src = REVIEWERS_SRC / ROLE_FILES[role_id]
        dest_dir = OPS / "round_b_packages" / role_id
        dest_dir.mkdir(parents=True, exist_ok=True)
        csv_dest = dest_dir / f"MokXya_MAI07_V3_ROUND_B__{stub}.csv"
        xlsx_dest = dest_dir / f"MokXya_MAI07_V3_ROUND_B_PACKAGE__{stub}.xlsx"
        shutil.copy2(csv_src, csv_dest)
        shutil.copy2(xlsx_src, xlsx_dest)
        zip_path = OPS / "round_b_packages" / f"MokXya_MAI07_V3_ROUND_B_PACKAGE__{role_id}.zip"
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for p in (csv_dest, xlsx_dest):
                info = zipfile.ZipInfo(f"{role_id}/{p.name}", date_time=(2026, 7, 16, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                zf.writestr(info, p.read_bytes())
        out[role_id] = str(zip_path.relative_to(REPO)).replace("\\", "/")
        (OPS / "round_b_inbox" / role_id).mkdir(parents=True, exist_ok=True)
    return out


def print_status() -> dict[str, Any]:
    status = load_status()
    render_dashboard(status)
    print(json.dumps({
        "state": status.get("state"),
        "ROUND_A_LOCKED": status.get("ROUND_A_LOCKED"),
        "ROUND_B_READY": status.get("ROUND_B_READY"),
        "roles": {
            rid: {
                "batches_returned": info.get("batches_returned"),
                "batch_count": info.get("batch_count"),
                "completion_pct": info.get("completion_pct"),
                "zip_path": info.get("zip_path"),
                "inbox_path": info.get("inbox_path"),
            }
            for rid, info in (status.get("roles") or {}).items()
        },
        "next_human_action": status.get("next_human_action"),
        "LINGUIST_APPROVED": status.get("LINGUIST_APPROVED"),
        "QUALITY_GATES_PASSED": status.get("QUALITY_GATES_PASSED"),
        "MAI_08": status.get("MAI_08"),
    }, indent=2, sort_keys=True))
    return status


def main() -> int:
    import argparse

    p = argparse.ArgumentParser(description="MAI-07 V3 automated review operations")
    p.add_argument("--run", action="store_true")
    p.add_argument("--status", action="store_true")
    p.add_argument("--init", action="store_true")
    args = p.parse_args()
    if args.status:
        print_status()
        return 0
    if args.init:
        out = initialize_packages()
        print(json.dumps(out, indent=2, sort_keys=True))
        return 0
    # default --run
    out = advance_workflow()
    print(json.dumps(out, indent=2, sort_keys=True, default=str))
    state = out.get("state") or out.get("status")
    return 0 if state in {"WAITING_FOR_ROUND_A_SUBMISSIONS", "WAITING_FOR_ROUND_B_SUBMISSIONS", "REVIEW_COMPLETE_READY_FOR_R3J_B", "PASSED_AUTOMATION"} or out.get("status") == "PASSED_AUTOMATION" else 1


if __name__ == "__main__":
    raise SystemExit(main())
