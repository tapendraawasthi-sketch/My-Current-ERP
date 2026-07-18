"""MAI-07R3J-A — build independent V3 governance/review packet.

Governance and review-packet phase only.
Does not modify transliteration runtime/resources.
Does not open V1/V2 case bodies or frozen predictions.
Does not evaluate candidates or implement MAI-08.
"""

from __future__ import annotations

import csv
import hashlib
import json
import random
import re
from datetime import timezone  # noqa: F401 — reserved for steward tooling
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .mai07_r3ja_v3_agreement import (
    ACCEPTANCE_GATES,
    ROUND_A_DISPOSITIONS,
    ROUND_B_ACCEPTABILITY,
)
from .mai07_r3ja_v3_firewall import (
    REPO,
    assert_path_allowed,
    assert_source_code_firewall,
    snapshot_historical_hashes,
)
from .mai07_r3ja_v3_independent_corpus import (
    PHASE_ID,
    PROHIBITED_FOR_TRAINING,
    SPLIT_SEED,
    V3SourceItem,
    build_independent_corpus,
    coverage_report,
    family_pool_assignment,
)

OUT = REPO / "docs" / "mokxya-ai" / "reviews" / "mai07_v3"
PACKET_SEED = "mai07-r3ja-packet-20260716"
PACKET_AUTHORITY_UTC = "2026-07-16T00:00:00+00:00"
CSV_INJECT = re.compile(r"^[=+\-@]")

REVIEWER_ROLES = (
    {
        "role_id": "PRODUCT_POLICY",
        "display": "Product Policy Reviewer",
        "file_stub": "product_policy",
        "rounds": ("A",),
        "satisfies_linguist_approval": False,
    },
    {
        "role_id": "NEPALI_FLUENT_A",
        "display": "Nepali-Fluent Reviewer A",
        "file_stub": "nepali_fluent_a",
        "rounds": ("A", "B"),
        "satisfies_linguist_approval": False,
    },
    {
        "role_id": "PROFESSIONAL_LINGUIST_B",
        "display": "Professional Nepali Linguist Reviewer B",
        "file_stub": "professional_linguist_b",
        "rounds": ("A", "B"),
        "satisfies_linguist_approval": True,  # only if completed + declared + adjudicated later
    },
    {
        "role_id": "ACCOUNTING_DOMAIN",
        "display": "Nepal Accounting/Business Domain Reviewer",
        "file_stub": "accounting_domain",
        "rounds": ("A", "B"),
        "satisfies_linguist_approval": False,
    },
    {
        "role_id": "INDEPENDENT_ADJUDICATOR",
        "display": "Independent Adjudicator",
        "file_stub": "adjudicator",
        "rounds": ("C",),
        "satisfies_linguist_approval": False,
    },
)


def _sha_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sha_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_text(path: Path, text: str) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8", newline="\n")
    return _sha_file(path)


def _write_json(path: Path, obj: Any) -> str:
    return _write_text(path, json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True) + "\n")


def _csv_cell(value: Any) -> str:
    s = "" if value is None else str(value)
    if CSV_INJECT.match(s):
        return "'" + s
    return s


def opaque_review_id(source_item_id: str, role_id: str) -> str:
    return "V3R-" + _sha_bytes(f"{PACKET_SEED}:{role_id}:{source_item_id}".encode("utf-8"))[:12]


def shuffle_candidates(cands: list[str], review_id: str, role_id: str) -> list[str]:
    rng = random.Random(f"{PACKET_SEED}:B:{role_id}:{review_id}")
    out = list(cands)
    # Always include identity-like surface and none option markers for Round B UI
    if "NONE_ACCEPTABLE" not in out:
        out.append("NONE_ACCEPTABLE")
    rng.shuffle(out)
    return out


def build_blind_mapping(items: list[V3SourceItem], role_assignments: dict[str, list[str]]) -> dict[str, Any]:
    rows = []
    for it in items:
        pool = family_pool_assignment(it.family_id)
        for role_id, source_ids in role_assignments.items():
            if it.source_item_id not in source_ids:
                continue
            rid = opaque_review_id(it.source_item_id, role_id)
            cands = shuffle_candidates(list(it.proposed_round_b_candidates), rid, role_id)
            rows.append(
                {
                    "review_id": rid,
                    "role_id": role_id,
                    "source_item_id": it.source_item_id,
                    "family_id": it.family_id,
                    "future_pool_assignment": pool,
                    "provenance_class": it.provenance_class,
                    "candidate_order": cands,
                    "counterfactual_group_id": it.counterfactual_group_id,
                    "counterfactual_role": it.counterfactual_role,
                }
            )
    return {
        "schema_version": "mai07_v3_blind_mapping.1.0.0",
        "use": "adjudication_import_only",
        "prohibited": [
            "reviewer_facing",
            "runtime_import",
            "training",
            "model_gold",
        ],
        "split_seed": SPLIT_SEED,
        "packet_seed": PACKET_SEED,
        "prohibited_for_training": True,
        "rows": rows,
    }


def assign_items_to_roles(items: list[V3SourceItem]) -> dict[str, list[str]]:
    """Each scored item gets ≥2 independent reviewers (Fluent A + Linguist B).

    Product policy: all Round A. Domain: accounting-tagged. Adjudicator: none yet.
    """
    all_ids = [it.source_item_id for it in items]
    accounting_ids = [
        it.source_item_id
        for it in items
        if "accounting_business" in it.design_tags or "shared_context" in it.design_tags
    ]
    return {
        "PRODUCT_POLICY": list(all_ids),
        "NEPALI_FLUENT_A": list(all_ids),
        "PROFESSIONAL_LINGUIST_B": list(all_ids),
        "ACCOUNTING_DOMAIN": accounting_ids,
        "INDEPENDENT_ADJUDICATOR": [],
    }


def write_round_a_csv(path: Path, items: list[V3SourceItem], role_id: str) -> str:
    headers = [
        "review_id",
        "input_text",
        "highlighted_span",
        "disposition",
        "confidence",
        "reason_category",
        "natural_context_ok",
        "suspected_ambiguity",
        "reviewer_notes",
    ]
    lines = [",".join(headers)]
    for it in items:
        rid = opaque_review_id(it.source_item_id, role_id)
        row = [
            rid,
            _csv_cell(it.input_text),
            _csv_cell(it.highlighted_span),
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        lines.append(",".join(row))
    # Use csv module for proper quoting
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        for it in items:
            rid = opaque_review_id(it.source_item_id, role_id)
            w.writerow(
                [
                    rid,
                    _csv_cell(it.input_text),
                    _csv_cell(it.highlighted_span),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
    return _sha_file(path)


def write_round_b_csv(
    path: Path,
    items: list[V3SourceItem],
    role_id: str,
    mapping_by_review: dict[str, list[str]],
) -> str:
    headers = [
        "review_id",
        "input_text",
        "highlighted_span",
        "candidate_index",
        "candidate_surface",
        "acceptability",
        "reviewer_entered_candidate",
        "notes",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        for it in items:
            rid = opaque_review_id(it.source_item_id, role_id)
            cands = mapping_by_review.get(rid, shuffle_candidates(list(it.proposed_round_b_candidates), rid, role_id))
            for idx, surf in enumerate(cands):
                w.writerow(
                    [
                        rid,
                        _csv_cell(it.input_text),
                        _csv_cell(it.highlighted_span),
                        idx,
                        _csv_cell(surf),
                        "",
                        "",
                        "",
                    ]
                )
    return _sha_file(path)


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1E2433")
    font = Font(color="FFFFFF", bold=True, size=10)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True)


def _stabilize_xlsx_zip(path: Path) -> None:
    """Rewrite XLSX zip entries with fixed timestamps/order/core props for deterministic SHA-256."""
    import io
    import zipfile

    fixed_core = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        b'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        b'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        b"<dc:creator>MAI-07R3J-A</dc:creator>"
        b'<dcterms:created xsi:type="dcterms:W3CDTF">2026-07-16T00:00:00Z</dcterms:created>'
        b'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-07-16T00:00:00Z</dcterms:modified>'
        b"<cp:lastModifiedBy>MAI-07R3J-A</cp:lastModifiedBy>"
        b"</cp:coreProperties>"
    )
    raw = path.read_bytes()
    zin = zipfile.ZipFile(io.BytesIO(raw), "r")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in sorted(zin.infolist(), key=lambda i: i.filename):
            payload = fixed_core if info.filename == "docProps/core.xml" else zin.read(info.filename)
            new = zipfile.ZipInfo(filename=info.filename, date_time=(2026, 7, 16, 0, 0, 0))
            new.compress_type = zipfile.ZIP_DEFLATED
            new.create_system = 0
            zout.writestr(new, payload)
    zin.close()
    path.write_bytes(buf.getvalue())


def build_workbook(
    path: Path,
    *,
    role: dict[str, Any],
    round_a_rows: list[list[Any]] | None,
    round_b_rows: list[list[Any]] | None,
    is_adjudicator: bool = False,
) -> str:
    wb = Workbook()
    # START_HERE
    ws0 = wb.active
    ws0.title = "START_HERE"
    instructions = [
        "MAI-07 V3 Independent Review Packet — START HERE",
        f"Role: {role['display']} ({role['role_id']})",
        "",
        "Which sheets you edit:",
        "- REVIEWER_DECLARATION (required before any labels)",
        "- ROUND_A_CONTEXT (Round A only — no candidates/ranks/scores shown)",
        "- ROUND_B_CANDIDATES (only after Round A is lock-hashed; Round B cannot overwrite Round A)",
        "- REVIEW_PROGRESS (read counters; do not invent completion)",
        "",
        "Required columns: disposition, confidence (Round A); acceptability (Round B).",
        "How to save: File → Save As .xlsx keeping the same filename pattern.",
        "What NOT to change: hidden mapping sheets (none present), import schemas, review_id values, candidate_index.",
        "When Round B may begin: ONLY after your Round A file is returned and hash-locked by the packet steward.",
        "Where to return: return this completed workbook to the MokXya review steward (not to runtime engineers for gold mining).",
        "",
        "Do not claim LINGUIST_APPROVED. Product-owner review alone cannot satisfy professional linguist approval.",
        "Do not use AI to autofill human-review fields.",
        "CANNOT_DECIDE / ABSTAIN is allowed; it is not a pass.",
        "prohibited_for_training=true for all items.",
    ]
    for i, line in enumerate(instructions, start=1):
        ws0.cell(i, 1, line)
    ws0.column_dimensions["A"].width = 110

    # REVIEWER_DECLARATION
    ws1 = wb.create_sheet("REVIEWER_DECLARATION")
    decls = [
        ["field", "value"],
        ["reviewer_full_name", ""],
        ["reviewer_email", ""],
        ["role_id", role["role_id"]],
        ["qualification_summary", ""],
        ["professional_linguist_credentials", ""],
        ["declaration_no_ai_autofill", ""],
        ["declaration_independent_from_other_reviewers", ""],
        ["declaration_no_runtime_predictions_used", ""],
        ["declaration_date_utc", ""],
        ["signature_or_typed_name", ""],
    ]
    for r, row in enumerate(decls, start=1):
        for c, v in enumerate(row, start=1):
            ws1.cell(r, c, v)
    _style_header(ws1, 2)
    ws1.column_dimensions["A"].width = 45
    ws1.column_dimensions["B"].width = 50

    # ROUND_A_CONTEXT
    wsA = wb.create_sheet("ROUND_A_CONTEXT")
    a_headers = [
        "review_id",
        "input_text",
        "highlighted_span",
        "disposition",
        "confidence",
        "reason_category",
        "natural_context_ok",
        "suspected_ambiguity",
        "reviewer_notes",
    ]
    wsA.append(a_headers)
    _style_header(wsA, len(a_headers))
    if round_a_rows:
        for row in round_a_rows:
            wsA.append(row)
    dv_disp = DataValidation(
        type="list",
        formula1='"' + ",".join(ROUND_A_DISPOSITIONS) + '"',
        allow_blank=True,
    )
    dv_conf = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=True)
    wsA.add_data_validation(dv_disp)
    wsA.add_data_validation(dv_conf)
    if round_a_rows:
        dv_disp.add(f"D2:D{len(round_a_rows)+1}")
        dv_conf.add(f"E2:E{len(round_a_rows)+1}")
    for i, h in enumerate(a_headers, start=1):
        wsA.column_dimensions[get_column_letter(i)].width = 18 if i != 2 else 55

    # ROUND_B_CANDIDATES
    wsB = wb.create_sheet("ROUND_B_CANDIDATES")
    b_headers = [
        "review_id",
        "input_text",
        "highlighted_span",
        "candidate_index",
        "candidate_surface",
        "acceptability",
        "reviewer_entered_candidate",
        "notes",
    ]
    wsB.append(b_headers)
    _style_header(wsB, len(b_headers))
    if round_b_rows and "B" in role["rounds"]:
        for row in round_b_rows:
            wsB.append(row)
    elif is_adjudicator:
        wsB.append(["(Round B locked templates released after Round A hash-lock)", "", "", "", "", "", "", ""])
    else:
        wsB.append(["(Round B opens only after Round A lock — template columns ready)", "", "", "", "", "", "", ""])
    dv_acc = DataValidation(
        type="list",
        formula1='"' + ",".join(ROUND_B_ACCEPTABILITY) + '"',
        allow_blank=True,
    )
    wsB.add_data_validation(dv_acc)
    if round_b_rows and "B" in role["rounds"]:
        dv_acc.add(f"F2:F{len(round_b_rows)+1}")
    for i in range(1, 9):
        wsB.column_dimensions[get_column_letter(i)].width = 16 if i != 2 else 50

    # REVIEW_PROGRESS
    wsP = wb.create_sheet("REVIEW_PROGRESS")
    wsP.append(["metric", "formula_or_value"])
    n_a = len(round_a_rows or [])
    wsP.append(["round_a_rows", n_a])
    wsP.append(["round_a_completed", f'=COUNTA(ROUND_A_CONTEXT!D2:D{max(n_a,1)+1})'])
    wsP.append(["round_b_rows", len(round_b_rows or [])])
    wsP.append(["note", "Do not edit review_id. Complete declaration first."])
    _style_header(wsP, 2)

    # VALIDATION_ERRORS
    wsE = wb.create_sheet("VALIDATION_ERRORS")
    wsE.append(["error_code", "detail"])
    wsE.append(["(empty until steward validation)", ""])
    _style_header(wsE, 2)

    # SUBMISSION_CHECKLIST
    wsC = wb.create_sheet("SUBMISSION_CHECKLIST")
    for line in [
        ["checklist_item", "done_Y_N"],
        ["declaration_completed", ""],
        ["all_round_a_rows_labeled_or_abstained", ""],
        ["no_ai_autofill", ""],
        ["did_not_inspect_blind_mapping", ""],
        ["did_not_use_runtime_predictions", ""],
        ["saved_as_xlsx_without_macros", ""],
        ["returned_to_steward", ""],
    ]:
        wsC.append(line)
    _style_header(wsC, 2)

    path.parent.mkdir(parents=True, exist_ok=True)
    # Stabilize OOXML metadata so repeated builds hash identically.
    from datetime import datetime, timezone as tz

    fixed = datetime(2026, 7, 16, 0, 0, 0, tzinfo=tz.utc)
    wb.properties.creator = "MAI-07R3J-A"
    wb.properties.lastModifiedBy = "MAI-07R3J-A"
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.save(path)
    _stabilize_xlsx_zip(path)
    return _sha_file(path)


def write_documentation(out: Path) -> dict[str, str]:
    hashes: dict[str, str] = {}
    docs = {
        "README.md": f"""# MAI-07 V3 Independent Review Packet (R3J-A)

Status: **REVIEW_PACKET_READY** / overall **BLOCKED_PENDING_INDEPENDENT_HUMAN_REVIEW**

Phase: `{PHASE_ID}`

This packet retires V2 from further candidate selection and prepares an independently
sourced, blindly reviewed, professionally adjudicated V3 benchmark.

- Does **not** pass MAI-07 quality.
- `QUALITY_GATES_PASSED=false`
- `LINGUIST_APPROVED=false` until professional reviewer evidence exists
- `PRODUCTION_APPROVED=false`
- `MAI-08=NOT_STARTED`
- Next: `MAI-07R3J-B-ADJUDICATION-AND-V3-FREEZE`

## V2 status

`HISTORICAL_BENCHMARK_EXHAUSTED_FOR_MODEL_SELECTION` — immutable historical record only.
Do not mine V2 failures for corrective development. Do not delete V2.

## Firewall

Builders must not open V1/V2 case bodies, frozen prediction rows, or prior blind mappings as gold.
See `SOURCE_PROVENANCE_POLICY.md` and ADR_0010.

## Human next steps

See final report section and `START_HERE_EN_NP.md`.
""",
        "START_HERE_EN_NP.md": """# START HERE / यहाँबाट सुरु गर्नुहोस्

## English

1. Open **your** role-specific workbook only.
2. Complete **REVIEWER_DECLARATION** first.
3. Label **ROUND_A_CONTEXT** only (no candidates/ranks/scores).
4. Save and return Round A for hash-lock.
5. Round B opens only after Round A lock.
6. Do not edit blind mappings or import schemas.
7. Do not use AI autofill.
8. Product-owner review alone cannot set `LINGUIST_APPROVED=true`.

## नेपाली

1. आफ्नो भूमिकाको workbook मात्र खोल्नुहोस्।
2. पहिले **REVIEWER_DECLARATION** पूरा गर्नुहोस्।
3. **ROUND_A_CONTEXT** मात्र लेबल गर्नुहोस् (candidate/rank/score देखाइँदैन)।
4. Round A फाइल फर्काएर hash-lock गराउनुहोस्।
5. Round B Round A lock पछि मात्र खुल्छ।
6. Blind mapping वा import schema नचलाउनुहोस्।
7. AI ले review field नभर्नुहोस्।
8. Product owner को review मात्रले `LINGUIST_APPROVED=true` हुँदैन।
""",
        "REVIEW_INSTRUCTIONS_EN.md": """# V3 Review Instructions (English)

## Round A — Context / Disposition

Allowed dispositions:
ENGLISH_IDENTITY_REQUIRED, DEVANAGARI_TRANSLITERATION_REQUIRED, CONTEXT_DEPENDENT,
IDENTITY_FIRST_REVIEW_REQUIRED, TRANSLITERATION_OPTIONAL, NO_TRANSLITERATION_ALLOWED,
NAME_OR_ENTITY, ACRONYM_OR_IDENTIFIER, PROTECTED, ABSTAIN_CANNOT_DECIDE.

Capture confidence, reason category, natural context judgment, ambiguity, notes.
You must not see runtime candidates, ranks, scores, or failure history.

## Round B — Candidates

Shown only after Round A lock. Candidates are shuffled per reviewer.
Labels: ACCEPTABLE_PREFERRED, ACCEPTABLE_ALTERNATIVE, UNNATURAL_BUT_POSSIBLE,
INCORRECT, CANNOT_DECIDE.

Round B cannot overwrite Round A.

## Round C — Adjudication

Adjudicator sees disagreements only after independent rounds are locked.
""",
        "REVIEW_INSTRUCTIONS_NP.md": """# V3 समीक्षा निर्देशन (नेपाली)

## राउन्ड A

अनुमति प्राप्त disposition मात्र प्रयोग गर्नुहोस्। Runtime candidate, rank, score नहेर्नुहोस्।

## राउन्ड B

Round A lock पछि मात्र। Candidate क्रम reviewer अनुसार फरक हुन्छ। Round A अधिलेखन नगर्नुहोस्।

## राउन्ड C

असहमति मात्र adjudicator लाई।
""",
        "REVIEWER_ROLE_REQUIREMENTS.md": """# Reviewer Role Requirements

| Role | Independent labels | Linguist approval eligible |
|------|--------------------|----------------------------|
| Product Policy Reviewer | Policy intent | **No** |
| Nepali-Fluent Reviewer A | Language/naturalness | No (alone) |
| Professional Nepali Linguist B | Language/form/transliteration | **Only with** identity, qualification, declaration, decisions, adjudication |
| Nepal Accounting/Business Domain | Technical terminology | No |
| Independent Adjudicator | Disagreements only | No (alone) |

At least two independent reviews per scored item.
Every accounting/business item requires domain reviewer or adjudicator.
`LINGUIST_APPROVED=true` is prohibited without professional evidence package.
""",
        "SOURCE_PROVENANCE_POLICY.md": """# V3 Source Provenance Policy

Allowed provenance classes:
- NEW_HUMAN_AUTHORED
- LICENSED_PUBLIC_CORPUS
- OFFICIAL_PUBLIC_TEXT_WITH_PERMITTED_EVALUATION_USE
- INDEPENDENT_ENGINEERING_SCENARIO
- PROFESSIONAL_REVIEWER_SUBMITTED

Forbidden sources for V3 authoring:
- V1/V2 case bodies
- Frozen prediction rows / failed case IDs / acceptable-target sets
- Prior blind mappings as gold
- Consumed holdout bodies
- Runtime prediction outputs as gold
- Runtime lexicon-authored gold labels

All items: `prohibited_for_training=true`.
No PII / real account identifiers (synthetic tokens only).
If coverage cannot be met independently: `BLOCKED_SOURCE_DATA_REQUIRED`.
""",
        "V3_DATASET_DESIGN.md": """# V3 Dataset Design

Two pools (family-level hash split; reviewers cannot see assignment):

1. **POLICY_DEVELOPMENT** — may open after adjudication; never frozen-quality evidence.
2. **FROZEN_EVALUATION** — sealed after adjudication; opened only after new candidate + thresholds locked.

Split seed recorded before human review: see `V3_PACKET_MANIFEST.json`.

Purpose dimensions: English identity, Romanized Nepali, shared context, generation,
ranking/top-5, unambiguous top-1, multi-token, accounting terms, names/entities,
acronyms/ids, protected spans, ambiguity/review, unicode/cap safety.

Scope: MAI-07 only (not MAI-08 typo/code-mix expansion).
""",
        "V3_POPULATION_TAXONOMY.md": """# Future V3 Population Taxonomy (defined, not frozen)

- TRANSLITERATION_REQUIRED
- IDENTITY_REQUIRED
- CONTEXT_DEPENDENT_ENGLISH
- CONTEXT_DEPENDENT_NEPALI
- IDENTITY_FIRST_REVIEW_REQUIRED
- TRANSLITERATION_OPTIONAL
- NO_TRANSLITERATION_ALLOWED
- PROTECTED_IDENTITY
- HUMAN_REVIEW_REQUIRED
- INFORMATIONAL_EXCLUDED

Candidate roles:
- IDENTITY
- DEVANAGARI_TARGET
- OTHER_LATIN_REWRITE
- OTHER_SCRIPT
- INVALID_OR_UNSUPPORTED

Identity never receives target credit.
Populations must not be derived from model predictions.
""",
        "V3_METRIC_DEFINITIONS.md": """# Proposed V3 Metric Definitions (thresholds not locked from candidates)

Metrics (same population for numerator and denominator; no max(1,den)):
- target top-1 / recall@5 / MRR
- target-generation recall
- retention conditional on generation
- identity top-1
- false Devanagari
- complete counterfactual-group accuracy
- unresolved-review accuracy
- protected/raw mutations
- candidate caps
- deterministic output
- naturalness/human preference agreement

Required empty populations → INVALID_REQUIRED_POPULATION.
Optional empty → NOT_APPLICABLE.
Integer gates locked before any model evaluation.
Independent scorer required.
Do not set thresholds from observed candidate results.
""",
    }
    for name, body in docs.items():
        hashes[name] = _write_text(out / name, body)
    return hashes


def build_packet(repo: Path = REPO) -> dict[str, Any]:
    out = repo / "docs" / "mokxya-ai" / "reviews" / "mai07_v3"
    out.mkdir(parents=True, exist_ok=True)
    reviewers_dir = out / "reviewers"
    reviewers_dir.mkdir(parents=True, exist_ok=True)

    # Firewall on this builder's source
    builder_path = Path(__file__)
    viol = assert_source_code_firewall(builder_path)
    if viol:
        raise RuntimeError(f"firewall_import_violations:{viol}")

    items = build_independent_corpus()
    cov = coverage_report(items)
    if not cov["ok"]:
        return {
            "status": "BLOCKED_SOURCE_DATA_REQUIRED",
            "coverage": cov,
        }

    # Provenance completeness
    for it in items:
        assert it.provenance_class
        assert it.prohibited_for_training is True
        assert_path_allowed(out / "dummy_ok.json")  # ensure helper works

    role_assign = assign_items_to_roles(items)
    mapping = build_blind_mapping(items, role_assign)
    mapping_path = out / "V3_BLIND_MAPPING.json"
    mapping_sha = _write_json(mapping_path, mapping)

    by_id = {it.source_item_id: it for it in items}
    mapping_by_review = {r["review_id"]: r["candidate_order"] for r in mapping["rows"]}

    doc_hashes = write_documentation(out)

    schema = {
        "schema_version": "mai07_v3_review.1.0.0",
        "phase": PHASE_ID,
        "round_a_dispositions": list(ROUND_A_DISPOSITIONS),
        "round_b_acceptability": list(ROUND_B_ACCEPTABILITY),
        "confidence": ["HIGH", "MEDIUM", "LOW"],
        "notes": [
            "Round B cannot overwrite Round A",
            "CANNOT_DECIDE remains unresolved until adjudication",
            "Automatic bulk mapping rejected",
            "Product owner cannot satisfy linguist approval alone",
        ],
        "acceptance_gates": ACCEPTANCE_GATES,
        "prohibited_for_training": True,
    }
    schema_sha = _write_json(out / "V3_REVIEW_SCHEMA.json", schema)

    import_template = {
        "schema": "mai07_v3_review_import_v1",
        "status": "TEMPLATE_ONLY_NO_HUMAN_DECISIONS",
        "reviewer_declaration_required": True,
        "round_a": [],
        "round_b": [],
        "adjudication": [],
        "linguist_approved": False,
        "production_approved": False,
        "prohibited_for_training": True,
    }
    template_sha = _write_json(out / "V3_REVIEW_IMPORT_TEMPLATE.jsonl", import_template)

    # Control sample: duplicate pairs for consistency (synthetic duplicates of design tags)
    controls = {
        "schema_version": "1.0.0",
        "duplicate_control_pairs": 40,
        "note": "Steward inserts duplicate review rows at import time from mapping; not reviewer-authored",
        "prohibited_for_training": True,
    }
    control_sha = _write_json(out / "V3_CONTROL_SAMPLE_MANIFEST.json", controls)

    provenance_registry = {
        "schema_version": "1.0.0",
        "items": [
            {
                "source_item_id": it.source_item_id,
                "family_id": it.family_id,
                "provenance_class": it.provenance_class,
                "source_document": it.source_document,
                "source_licence": it.source_licence,
                "prohibited_for_training": True,
            }
            for it in items
        ],
    }
    prov_sha = _write_json(out / "V3_SOURCE_PROVENANCE_REGISTRY.json", provenance_registry)

    # Pool summary (not revealed in reviewer sheets)
    pool_counts = {"POLICY_DEVELOPMENT": 0, "FROZEN_EVALUATION": 0}
    for it in items:
        pool_counts[family_pool_assignment(it.family_id)] += 1

    file_hashes: dict[str, str] = {
        **{f"doc:{k}": v for k, v in doc_hashes.items()},
        "V3_REVIEW_SCHEMA.json": schema_sha,
        "V3_REVIEW_IMPORT_TEMPLATE.jsonl": template_sha,
        "V3_BLIND_MAPPING.json": mapping_sha,
        "V3_CONTROL_SAMPLE_MANIFEST.json": control_sha,
        "V3_SOURCE_PROVENANCE_REGISTRY.json": prov_sha,
    }

    # Per-reviewer artifacts
    for role in REVIEWER_ROLES:
        role_id = role["role_id"]
        assigned_ids = role_assign[role_id]
        role_items = [by_id[i] for i in assigned_ids]
        # Deterministic order by opaque review id
        role_items = sorted(role_items, key=lambda it: opaque_review_id(it.source_item_id, role_id))

        stub = role["file_stub"]
        round_a_rows = [
            [
                opaque_review_id(it.source_item_id, role_id),
                _csv_cell(it.input_text),
                _csv_cell(it.highlighted_span),
                "",
                "",
                "",
                "",
                "",
                "",
            ]
            for it in role_items
        ] if "A" in role["rounds"] else None

        round_b_rows = None
        if "B" in role["rounds"]:
            round_b_rows = []
            for it in role_items:
                rid = opaque_review_id(it.source_item_id, role_id)
                cands = mapping_by_review[rid]
                for idx, surf in enumerate(cands):
                    round_b_rows.append(
                        [
                            rid,
                            _csv_cell(it.input_text),
                            _csv_cell(it.highlighted_span),
                            idx,
                            _csv_cell(surf),
                            "",
                            "",
                            "",
                        ]
                    )

        if "A" in role["rounds"]:
            csv_a = reviewers_dir / f"V3_ROUND_A__{stub}.csv"
            file_hashes[csv_a.name] = write_round_a_csv(csv_a, role_items, role_id)
        if "B" in role["rounds"]:
            csv_b = reviewers_dir / f"V3_ROUND_B_TEMPLATE__{stub}.csv"
            file_hashes[csv_b.name] = write_round_b_csv(csv_b, role_items, role_id, mapping_by_review)

        xlsx = reviewers_dir / f"MokXya_MAI07_V3__{stub}.xlsx"
        file_hashes[xlsx.name] = build_workbook(
            xlsx,
            role=role,
            round_a_rows=round_a_rows,
            round_b_rows=round_b_rows if "B" in role["rounds"] else None,
            is_adjudicator=role_id == "INDEPENDENT_ADJUDICATOR",
        )

    # Adjudication + agreement templates (blank)
    adj = {
        "schema_version": "1.0.0",
        "status": "TEMPLATE_AWAITING_DISAGREEMENTS",
        "rows": [],
        "fields": [
            "review_id",
            "original_labels",
            "adjudicated_disposition",
            "rationale_category",
            "candidate_roles",
            "human_review_remains_required",
            "inclusion_decision",
        ],
    }
    file_hashes["V3_ADJUDICATION_TEMPLATE.json"] = _write_json(out / "V3_ADJUDICATION_TEMPLATE.json", adj)
    agr = {
        "schema_version": "1.0.0",
        "gates": ACCEPTANCE_GATES,
        "formulas": {
            "exact_agreement": "agree/n",
            "cohen_kappa": "(po-pe)/(1-pe)",
            "duplicate_consistency": "identical_duplicate_pairs/pairs",
        },
        "status": "TEMPLATE_PRE_REVIEW",
    }
    file_hashes["V3_AGREEMENT_TEMPLATE.json"] = _write_json(out / "V3_AGREEMENT_TEMPLATE.json", agr)

    # Import contract (spec only)
    import_contract = {
        "schema_version": "1.0.0",
        "fail_closed": [
            "wrong_file_hash",
            "missing_reviewer_declaration",
            "duplicate_review_ids",
            "missing_required_cells",
            "invalid_enums",
            "overwritten_round_a_evidence",
            "candidate_order_mismatch",
            "unauthorized_bulk_mapping",
            "reviewer_identity_collision",
            "incomplete_professional_review",
            "missing_domain_review",
            "unresolved_adjudication",
            "mapping_runtime_import_attempt",
        ],
        "immutable_submitted_reviews": True,
        "produces": "immutable_adjudication_record",
        "does_not_import_incomplete_reviews": True,
    }
    file_hashes["V3_REVIEW_IMPORT_CONTRACT.json"] = _write_json(
        out / "V3_REVIEW_IMPORT_CONTRACT.json", import_contract
    )

    hist = snapshot_historical_hashes(repo)
    manifest = {
        "schema_version": "1.0.0",
        "phase": PHASE_ID,
        "status": "REVIEW_PACKET_READY",
        "blocker": "BLOCKED_PENDING_INDEPENDENT_HUMAN_REVIEW",
        "QUALITY_GATES_PASSED": False,
        "LINGUIST_APPROVED": False,
        "PRODUCTION_APPROVED": False,
        "MAI_08": "NOT_STARTED",
        "next_phase": "MAI-07R3J-B-ADJUDICATION-AND-V3-FREEZE",
        "v2_governance": "HISTORICAL_BENCHMARK_EXHAUSTED_FOR_MODEL_SELECTION",
        "split_seed": SPLIT_SEED,
        "packet_seed": PACKET_SEED,
        "item_count": len(items),
        "pool_counts": pool_counts,
        "coverage": cov,
        "role_assignment_counts": {k: len(v) for k, v in role_assign.items()},
        "file_hashes": file_hashes,
        "historical_artifact_hashes": hist,
        "prohibited_for_training": True,
        "human_decisions_included": False,
        "packet_authority_utc": PACKET_AUTHORITY_UTC,
    }
    man_sha = _write_json(out / "V3_PACKET_MANIFEST.json", manifest)
    hash_manifest = {
        "V3_PACKET_MANIFEST.json": man_sha,
        **file_hashes,
        "historical": hist,
    }
    _write_json(out / "V3_HASH_MANIFEST.json", hash_manifest)

    return {
        "status": "REVIEW_PACKET_READY",
        "blocker": "BLOCKED_PENDING_INDEPENDENT_HUMAN_REVIEW",
        "manifest_sha256": man_sha,
        "item_count": len(items),
        "coverage": cov,
        "pool_counts": pool_counts,
        "out_dir": str(out.relative_to(repo)).replace("\\", "/"),
    }


def main() -> int:
    import argparse

    p = argparse.ArgumentParser(description="MAI-07R3J-A V3 review packet builder")
    p.add_argument("--build", action="store_true")
    args = p.parse_args()
    if args.build or True:
        result = build_packet()
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0 if result.get("status") == "REVIEW_PACKET_READY" else 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
