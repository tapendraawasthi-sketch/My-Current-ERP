"""MAI-07R3K AI-assisted cross-role consensus diagnostic tests."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.oip.modules.language_runtime.transliteration import (
    ENABLE_PROMOTION_OVERLAY,
    RUNTIME_VERSION,
)
from src.oip.modules.language_runtime.transliteration.application.mai07_r3k_cross_role_consensus_diagnostic import (
    DEFAULT_OUT,
    EXPECTED_ACCT_SEMANTIC,
    EXPECTED_FOUR_ROLE,
    EXPECTED_REM_SEMANTIC,
    EXPECTED_THREE_ROLE,
    EXPECTED_TOTAL_JUDGMENTS,
    EXPECTED_UNIQUE_CASES,
    FIXED_PROVENANCE,
    OFFICIAL_INBOX,
    Mai07R3KDiagnosticError,
    assert_official_inbox_empty,
    prove_deterministic_rerun,
    run_diagnostic,
    verify_input_semantic_hashes,
)
from src.oip.modules.language_runtime.transliteration.infrastructure import (
    resource_repository as xlrr,
)

ACTIVE = "1617425373bf525968b5af2a3b1cc8b8e5ad83e68457cfbbb47c73c78c84e930"
REPORT = DEFAULT_OUT / "reports" / "CONSENSUS_DIAGNOSTIC_REPORT.json"


def test_input_semantic_hashes():
    h = verify_input_semantic_hashes()
    assert h["accounting"] == EXPECTED_ACCT_SEMANTIC
    assert h["remaining"] == EXPECTED_REM_SEMANTIC


def test_population_counts():
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    assert report["unique_source_item_ids"] == EXPECTED_UNIQUE_CASES == 1111
    assert report["total_role_judgments"] == EXPECTED_TOTAL_JUDGMENTS == 3944
    assert report["four_role_cases"] == EXPECTED_FOUR_ROLE == 611
    assert report["three_role_cases"] == EXPECTED_THREE_ROLE == 500
    assert report["role_counts"]["PRODUCT_POLICY"] == 1111
    assert report["role_counts"]["NEPALI_FLUENT_A"] == 1111
    assert report["role_counts"]["PROFESSIONAL_LINGUIST_B"] == 1111
    assert report["role_counts"]["ACCOUNTING_DOMAIN"] == 611


def test_no_majority_as_gold_field():
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    assert report["governance"]["majority_voting_is_gold"] is False
    assert report["provenance"]["majority_voting_is_gold"] is False
    assert FIXED_PROVENANCE["majority_voting_is_gold"] is False
    decisions = [
        json.loads(ln)
        for ln in (DEFAULT_OUT / "canonical/CROSS_ROLE_DECISIONS.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
        if ln.strip()
    ]
    assert all(d.get("majority_as_gold") is False for d in decisions)


def test_agreement_and_contamination_documented():
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    a = report["agreement"]
    assert a["ai_output_similarity_metrics_non_independent"]["not_human_inter_rater_reliability"] is True
    assert "Cohen" not in json.dumps(a) or "not Cohen" in a["ai_output_similarity_metrics_non_independent"]["note"]
    # Source contamination: map vs heuristic buckets present
    buckets = a["agreement_by_decision_source_bucket"]
    assert buckets["ACCOUNTING_MAP_INHERITED_NO_HEURISTIC"]["n"] == 611
    assert buckets["HEURISTIC_V1_PRESENT"]["n"] == 500


def test_risk_queue_inclusion():
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    assert report["risk_queue_count"] > 0
    assert sum(report["risk_tier_counts"].values()) == report["risk_queue_count"]
    q = [
        json.loads(ln)
        for ln in (DEFAULT_OUT / "canonical/RISK_QUEUE.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
        if ln.strip()
    ]
    assert len(q) == report["risk_queue_count"]
    assert all(item["reason_codes"] for item in q)
    assert all(item["risk_tier"].startswith("TIER_") for item in q)


def test_blinded_packet_no_leakage():
    xlsx = DEFAULT_OUT / "targeted_review_packet/MokXya_MAI07_V3_R3K_TARGETED_BLIND_REVIEW.xlsx"
    assert xlsx.is_file()
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        assert wb.sheetnames == ["TARGETED_REVIEW"]
        forbidden = (
            "PRODUCT_POLICY",
            "NEPALI_FLUENT_A",
            "ACCOUNTING_DOMAIN",
            "HEURISTIC_V1",
            "ENGLISH_IDENTITY_REQUIRED",
        )
        for row in wb["TARGETED_REVIEW"].iter_rows(min_row=2, values_only=True):
            # disposition/confidence/notes blank
            assert not row[4] and not row[5] and not row[6]
            blob = " ".join("" if c is None else str(c) for c in row[:4])
            for tok in forbidden:
                # ENGLISH may appear in instructions context_note? check context
                if tok == "ENGLISH_IDENTITY_REQUIRED":
                    assert tok not in blob
                elif tok in blob:
                    pytest.fail(f"leakage {tok}")
    finally:
        wb.close()
    private = DEFAULT_OUT / "private_adjudication_import_only/PRIVATE_BLIND_MAPPING_ADJUDICATION_IMPORT_ONLY.jsonl"
    assert private.is_file()


def test_deterministic_rerun(tmp_path):
    proof = prove_deterministic_rerun(tmp_path / "det")
    assert proof["ok"] is True
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    assert proof["semantic_hash"] == report["semantic_hash"]


def test_official_inbox_isolation():
    assert_official_inbox_empty()
    assert not list(OFFICIAL_INBOX.rglob("*.xlsx")) if OFFICIAL_INBOX.exists() else True


def test_runtime_unchanged():
    assert RUNTIME_VERSION == "mai-07.1.3-r3f-sealnew"
    assert ENABLE_PROMOTION_OVERLAY is False
    r = xlrr.validate_resources()
    assert r["ok"] and r["content_hash"] == ACTIVE


def test_governance_flags():
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    g = report["governance"]
    assert g["QUALITY_GATES_PASSED"] is False
    assert g["LINGUIST_APPROVED"] is False
    assert g["PRODUCTION_APPROVED"] is False
    assert g["ROUND_A_LOCKED"] is False
    assert g["ROUND_B_READY"] is False
    assert g["MAI-08"] == "NOT_STARTED"
    assert report["provenance"]["independent_human_review"] is False
    assert report["provenance"]["agreement_is_independent_human_irr"] is False


def test_module_has_no_accounting_mutation_imports():
    src = Path(
        "src/oip/modules/language_runtime/transliteration/application/"
        "mai07_r3k_cross_role_consensus_diagnostic.py"
    )
    # relative to erp_bot when pytest cwd may be repo root
    candidates = [
        Path(__file__).resolve().parents[3]
        / "src/oip/modules/language_runtime/transliteration/application/"
        / "mai07_r3k_cross_role_consensus_diagnostic.py",
        Path("erp_bot") / src,
        src,
    ]
    text = None
    for p in candidates:
        if p.is_file():
            text = p.read_text(encoding="utf-8")
            break
    assert text is not None
    import_lines = [
        ln for ln in text.splitlines() if ln.strip().startswith(("import ", "from "))
    ]
    joined = "\n".join(import_lines)
    for bad in ("orbixPostingService", "ledger_tools", "khataConfirm", "mutation_gateway"):
        assert bad not in joined


def test_altered_authority_text_rejected(monkeypatch):
    from src.oip.modules.language_runtime.transliteration.application import (
        mai07_r3k_cross_role_consensus_diagnostic as mod,
    )

    rows = mod.load_jsonl(mod.ACCT_JSONL)
    mapping = mod.load_blind_mapping()
    rem = mod.load_jsonl(mod.REM_JSONL)
    # tamper one remaining row text while keeping same review_id mapping
    rem[0] = dict(rem[0])
    rem[0]["input_text"] = "TAMPERED_AUTHORITY_TEXT"
    with pytest.raises(Mai07R3KDiagnosticError, match="altered_authority_text_span"):
        mod.build_case_text_map(rows, rem, mapping)


def test_duplicate_role_decision_rejected():
    from src.oip.modules.language_runtime.transliteration.application import (
        mai07_r3k_cross_role_consensus_diagnostic as mod,
    )

    rem = mod.load_jsonl(mod.REM_JSONL)
    mapping = mod.load_blind_mapping()
    sources = mod.load_decision_sources()
    # duplicate first row
    rem2 = rem + [rem[0]]
    with pytest.raises(Mai07R3KDiagnosticError, match="duplicate_role_decision"):
        mod.build_judgments([], rem2, mapping, sources)
