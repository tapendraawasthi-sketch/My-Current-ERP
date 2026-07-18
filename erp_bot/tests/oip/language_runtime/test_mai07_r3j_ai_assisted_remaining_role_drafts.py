"""MAI-07R3J-AI-ASSISTED-REMAINING-ROLE-DRAFTS tests."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.oip.modules.language_runtime.transliteration import (
    ENABLE_PROMOTION_OVERLAY,
    RUNTIME_VERSION,
)
from src.oip.modules.language_runtime.transliteration.application.draft_mai07_r3j_ai_assisted_remaining_roles import (
    DEFAULT_OUT,
    FIXED_PROVENANCE,
    OFFICIAL_INBOX,
    REMAINING_ROLES,
    draft_all_remaining_roles,
    heuristic_decision,
    load_accounting_content_map,
)
from src.oip.modules.language_runtime.transliteration.infrastructure import (
    resource_repository as xlrr,
)

REPO = Path(__file__).resolve().parents[4]
ACTIVE_RESOURCE_HASH = "1617425373bf525968b5af2a3b1cc8b8e5ad83e68457cfbbb47c73c78c84e930"
SUMMARY = DEFAULT_OUT / "REMAINING_ROLE_DRAFTS_SUMMARY.json"


@pytest.fixture(scope="module")
def content_map():
    return load_accounting_content_map()


def test_accounting_map_has_611(content_map):
    assert len(content_map) == 611


def test_summary_exists_with_3333_rows():
    assert SUMMARY.is_file()
    data = json.loads(SUMMARY.read_text(encoding="utf-8"))
    assert data["ok"] is True
    assert data["total_rows"] == 3333
    assert data["total_workbooks"] == 30
    assert len(data["roles"]) == 3
    assert data["provenance"]["independent_human_review"] is False
    assert data["provenance"]["ai_autofill_used"] is True
    assert data["provenance"]["user_accepted"] is False
    assert data["governance"]["ROUND_A_LOCKED"] is False
    assert data["governance"]["ROUND_B_READY"] is False
    assert data["governance"]["QUALITY_GATES_PASSED"] is False


def test_each_role_1111_unique_and_zip(content_map):
    data = json.loads(SUMMARY.read_text(encoding="utf-8"))
    for role in data["roles"]:
        assert role["row_count"] == 1111
        assert role["unique_review_ids"] == 1111
        assert role["workbook_count"] == 10
        assert Path(role["zip_path"]).is_file()
        assert role["decision_source_counts"]["ACCOUNTING_VERIFIED_CONTENT_MAP"] == 611
        assert role["decision_source_counts"]["HEURISTIC_V1"] == 500


def test_authority_fields_unchanged_sample():
    role = "PRODUCT_POLICY"
    src = (
        REPO
        / "docs/mokxya-ai/reviews/mai07_v3/review_operations/reviewer_packages"
        / role
        / "round_a"
    )
    draft = DEFAULT_OUT / "product_policy" / "round_a_drafts"
    src_wb = sorted(src.glob("*.xlsx"))[0]
    dst_wb = draft / (src_wb.stem + "__AI_ASSISTED_DRAFT.xlsx")
    assert dst_wb.is_file()
    a = load_workbook(src_wb, read_only=True, data_only=True)
    b = load_workbook(dst_wb, read_only=True, data_only=True)
    try:
        for ra, rb in zip(
            a["ROUND_A_CONTEXT"].iter_rows(min_row=2, values_only=True),
            b["ROUND_A_CONTEXT"].iter_rows(min_row=2, values_only=True),
            strict=False,
        ):
            if not ra or not ra[0]:
                continue
            assert ra[0] == rb[0]
            assert ra[1] == rb[1]
            assert ra[2] == rb[2]
            assert rb[3]  # disposition filled
            assert rb[4] in {"HIGH", "MEDIUM", "LOW"}
    finally:
        a.close()
        b.close()


def test_declaration_and_round_b_untouched():
    draft = DEFAULT_OUT / "nepali_fluent_a" / "round_a_drafts"
    wb_path = sorted(draft.glob("*__AI_ASSISTED_DRAFT.xlsx"))[0]
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    try:
        for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            field = str(row[0])
            val = "" if row[1] is None else str(row[1]).strip()
            if field in {"role_id", "credential_status"}:
                continue
            assert val == "", f"declaration populated: {field}={val!r}"
        # Round B first data row should remain placeholder / blank answers
        rb = list(wb["ROUND_B_CANDIDATES"].iter_rows(min_row=2, max_row=3, values_only=True))
        assert rb
        # acceptability column empty on placeholder
        assert not (rb[0][5] if len(rb[0]) > 5 else None)
    finally:
        wb.close()


def test_protected_and_acronym_heuristics():
    d = heuristic_decision(
        "Do not mutate protected token SERIAL-SN-V3-0001-R059 inside the posting draft",
        "SERIAL-SN-V3-0001-R059",
    )
    assert d.disposition == "PROTECTED"
    d2 = heuristic_decision("Sync failed for GST queue", "GST")
    assert d2.disposition == "ACRONYM_OR_IDENTIFIER"


def test_content_map_preferred_over_heuristic(content_map):
    # Any accounting pair should resolve via map source
    (text, span), decision = next(iter(content_map.items()))
    from src.oip.modules.language_runtime.transliteration.application.draft_mai07_r3j_ai_assisted_remaining_roles import (
        decide,
    )

    got = decide(text, span, content_map)
    assert got.source == "ACCOUNTING_VERIFIED_CONTENT_MAP"
    assert got.disposition == decision.disposition


def test_deterministic_regeneration(tmp_path):
    a = draft_all_remaining_roles(out_root=tmp_path / "a")
    b = draft_all_remaining_roles(out_root=tmp_path / "b")
    assert a["semantic_hash"] == b["semantic_hash"]
    assert a["total_rows"] == b["total_rows"] == 3333
    for ra, rb in zip(a["roles"], b["roles"], strict=True):
        assert ra["disposition_counts"] == rb["disposition_counts"]
        assert ra["decision_source_counts"] == rb["decision_source_counts"]
        assert ra["row_count"] == rb["row_count"]


def test_official_inbox_untouched():
    if OFFICIAL_INBOX.exists():
        assert not list(OFFICIAL_INBOX.rglob("*AI_ASSISTED*"))
        assert not list(OFFICIAL_INBOX.rglob("*.xlsx"))


def test_runtime_unchanged():
    assert RUNTIME_VERSION == "mai-07.1.13-r3s-active"
    assert ENABLE_PROMOTION_OVERLAY is False
    report = xlrr.validate_resources()
    assert report["ok"] is True
    assert report["content_hash"] == ACTIVE_RESOURCE_HASH


def test_governance_flags_frozen():
    assert FIXED_PROVENANCE["eligible_for_frozen_quality_gold"] is False
    assert FIXED_PROVENANCE["prohibited_for_training"] is True
    assert FIXED_PROVENANCE["submission_ready_under_current_protocol"] is False
    assert set(REMAINING_ROLES) == {
        "PRODUCT_POLICY",
        "NEPALI_FLUENT_A",
        "PROFESSIONAL_LINGUIST_B",
    }
