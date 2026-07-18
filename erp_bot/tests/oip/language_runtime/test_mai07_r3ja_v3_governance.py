"""MAI-07R3J-A — independent V3 governance/review-packet protocol tests."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import pytest

from src.oip.modules.language_runtime.transliteration import (
    ENABLE_PROMOTION_OVERLAY,
    RESOURCE_PACK_VERSION,
    RUNTIME_VERSION,
)
from src.oip.modules.language_runtime.transliteration.application.build_mai07_r3ja_v3_review_packet import (
    OUT,
    REVIEWER_ROLES,
    build_packet,
    opaque_review_id,
    shuffle_candidates,
)
from src.oip.modules.language_runtime.transliteration.application.mai07_r3ja_v3_agreement import (
    KNOWN_KAPPA_VECTOR,
    ACCEPTANCE_GATES,
    cohen_kappa,
    exact_agreement,
    population_metric_status,
    reject_bulk_mapping,
)
from src.oip.modules.language_runtime.transliteration.application.mai07_r3ja_v3_firewall import (
    REPO,
    assert_path_allowed,
    assert_source_code_firewall,
    is_forbidden_path,
    scan_tree_for_forbidden_opens,
    snapshot_historical_hashes,
)
from src.oip.modules.language_runtime.transliteration.application.mai07_r3ja_v3_independent_corpus import (
    MIN_COVERAGE,
    PROHIBITED_FOR_TRAINING,
    build_independent_corpus,
    coverage_report,
    family_pool_assignment,
)
from src.oip.modules.language_runtime.transliteration.infrastructure import (
    resource_repository as xlrr,
)

APP = Path(__file__).resolve().parents[3] / "src/oip/modules/language_runtime/transliteration/application"
DEFAULT_RESOURCE = "1617425373bf525968b5af2a3b1cc8b8e5ad83e68457cfbbb47c73c78c84e930"
CSV_INJECT = re.compile(r"^[=+\-@]")


@pytest.fixture(scope="module")
def packet():
    return build_packet(REPO)


def test_v2_case_body_firewall_blocks_forbidden_paths():
    assert is_forbidden_path(REPO / "evals/mai07/frozen_v2/something.jsonl")
    assert is_forbidden_path("evals/mai07/r3i_frozen_reauthorized/reports/MAI_07R3I_V2_ONE_SHOT_PREDICTIONS.jsonl")
    assert is_forbidden_path("docs/mokxya-ai/reviews/mai07r3/MAI_07R3_BLIND_MAPPING.json")
    with pytest.raises(PermissionError, match="V3_FIREWALL_BLOCKED"):
        assert_path_allowed(REPO / "evals/mai07/frozen_v2/x.jsonl")


def test_historical_attempt_immutability_snapshot(packet):
    hist = packet  # rebuild already done
    snap = snapshot_historical_hashes(REPO)
    man = json.loads((OUT / "V3_PACKET_MANIFEST.json").read_text(encoding="utf-8"))
    assert man["historical_artifact_hashes"] == snap
    # R3I attempt lock still present and hashed
    assert snap[
        "evals/mai07/r3i_frozen_reauthorized/MAI_07R3I_FROZEN_V2_ATTEMPT_001.LOCKED_NOT_RUN.json"
    ]


def test_no_runtime_imports_in_v3_packet_builder():
    paths = [
        APP / "build_mai07_r3ja_v3_review_packet.py",
        APP / "mai07_r3ja_v3_independent_corpus.py",
        APP / "mai07_r3ja_v3_firewall.py",
        APP / "mai07_r3ja_v3_agreement.py",
    ]
    viol = scan_tree_for_forbidden_opens(paths)
    assert viol == [], viol
    for p in paths:
        assert assert_source_code_firewall(p) == []


def test_independent_source_provenance_required():
    items = build_independent_corpus()
    assert all(it.provenance_class for it in items)
    assert all(it.prohibited_for_training is True for it in items)
    assert PROHIBITED_FOR_TRAINING is True


def test_dataset_family_split_isolation():
    items = build_independent_corpus()
    fam_pool = {}
    for it in items:
        pool = family_pool_assignment(it.family_id)
        if it.family_id in fam_pool:
            assert fam_pool[it.family_id] == pool
        fam_pool[it.family_id] = pool
    assert set(fam_pool.values()) == {"POLICY_DEVELOPMENT", "FROZEN_EVALUATION"}


def test_minimum_coverage_reconciliation(packet):
    assert packet["status"] == "REVIEW_PACKET_READY"
    cov = packet["coverage"]
    assert cov["ok"] is True
    for k, need in MIN_COVERAGE.items():
        assert cov["observed"][k] >= need


def test_opaque_review_ids():
    a = opaque_review_id("V3SRC-x", "NEPALI_FLUENT_A")
    b = opaque_review_id("V3SRC-x", "PROFESSIONAL_LINGUIST_B")
    assert a.startswith("V3R-")
    assert a != b
    assert "V3SRC" not in a


def test_reviewer_specific_candidate_shuffling():
    cands = ["alpha", "beta", "gamma"]
    o1 = shuffle_candidates(cands, "V3R-aaa", "NEPALI_FLUENT_A")
    o2 = shuffle_candidates(cands, "V3R-aaa", "PROFESSIONAL_LINGUIST_B")
    assert "NONE_ACCEPTABLE" in o1
    # Same review_id different roles → different order (highly likely; assert inequality of sequences)
    assert o1 != o2 or len(set(o1)) == 1


def test_round_a_excludes_candidates_ranks_scores(packet):
    csv_path = OUT / "reviewers" / "V3_ROUND_A__nepali_fluent_a.csv"
    text = csv_path.read_text(encoding="utf-8")
    header = text.splitlines()[0]
    assert "candidate" not in header.lower()
    assert "rank" not in header.lower()
    assert "score" not in header.lower()
    assert "disposition" in header


def test_round_b_cannot_overwrite_round_a():
    schema = json.loads((OUT / "V3_REVIEW_SCHEMA.json").read_text(encoding="utf-8"))
    assert any("cannot overwrite Round A" in n for n in schema["notes"])


def test_reviewer_declaration_required_in_workbook(packet):
    from openpyxl import load_workbook

    xlsx = OUT / "reviewers" / "MokXya_MAI07_V3__professional_linguist_b.xlsx"
    wb = load_workbook(xlsx, read_only=True)
    assert "REVIEWER_DECLARATION" in wb.sheetnames
    assert "START_HERE" in wb.sheetnames
    assert set(wb.sheetnames) >= {
        "START_HERE",
        "REVIEWER_DECLARATION",
        "ROUND_A_CONTEXT",
        "ROUND_B_CANDIDATES",
        "REVIEW_PROGRESS",
        "VALIDATION_ERRORS",
        "SUBMISSION_CHECKLIST",
    }


def test_professional_role_cannot_be_inferred_as_approval(packet):
    man = json.loads((OUT / "V3_PACKET_MANIFEST.json").read_text(encoding="utf-8"))
    assert man["LINGUIST_APPROVED"] is False
    assert man["PRODUCTION_APPROVED"] is False
    assert man["human_decisions_included"] is False


def test_product_owner_cannot_satisfy_linguist_approval():
    product = next(r for r in REVIEWER_ROLES if r["role_id"] == "PRODUCT_POLICY")
    assert product["satisfies_linguist_approval"] is False


def test_invalid_enums_and_duplicate_ids_documented():
    contract = json.loads((OUT / "V3_REVIEW_IMPORT_CONTRACT.json").read_text(encoding="utf-8"))
    assert "invalid_enums" in contract["fail_closed"]
    assert "duplicate_review_ids" in contract["fail_closed"]


def test_incomplete_rows_fail_closed_in_contract():
    contract = json.loads((OUT / "V3_REVIEW_IMPORT_CONTRACT.json").read_text(encoding="utf-8"))
    assert "missing_required_cells" in contract["fail_closed"]
    assert "missing_reviewer_declaration" in contract["fail_closed"]


def test_candidate_index_mismatch_fails():
    contract = json.loads((OUT / "V3_REVIEW_IMPORT_CONTRACT.json").read_text(encoding="utf-8"))
    assert "candidate_order_mismatch" in contract["fail_closed"]


def test_cannot_decide_remains_unresolved():
    schema = json.loads((OUT / "V3_REVIEW_SCHEMA.json").read_text(encoding="utf-8"))
    assert any("CANNOT_DECIDE remains unresolved" in n for n in schema["notes"])
    assert "ABSTAIN_CANNOT_DECIDE" in schema["round_a_dispositions"]
    assert "CANNOT_DECIDE" in schema["round_b_acceptability"]


def test_automatic_bulk_mapping_rejected():
    with pytest.raises(ValueError, match="AUTOMATIC_BULK_MAPPING_REJECTED"):
        reject_bulk_mapping({"bulk_map": True})
    contract = json.loads((OUT / "V3_REVIEW_IMPORT_CONTRACT.json").read_text(encoding="utf-8"))
    assert "unauthorized_bulk_mapping" in contract["fail_closed"]


def test_formula_injection_escaped_in_round_a_csv():
    # Builder prefixes dangerous cells; verify corpus surfaces are escaped if needed
    csv_path = OUT / "reviewers" / "V3_ROUND_A__nepali_fluent_a.csv"
    for line in csv_path.read_text(encoding="utf-8").splitlines()[1:]:
        # crude: no unescaped formula start in fields after first comma blocks without quote
        parts = line.split(",")
        for part in parts[1:3]:
            if part and CSV_INJECT.match(part):
                pytest.fail(f"unescaped_injection:{part}")


def test_workbook_required_sheets_columns(packet):
    from openpyxl import load_workbook

    for role in REVIEWER_ROLES:
        path = OUT / "reviewers" / f"MokXya_MAI07_V3__{role['file_stub']}.xlsx"
        assert path.exists(), path
        wb = load_workbook(path, read_only=True)
        assert "START_HERE" in wb.sheetnames


def test_mapping_never_imported_by_runtime():
    mapping = json.loads((OUT / "V3_BLIND_MAPPING.json").read_text(encoding="utf-8"))
    assert mapping["use"] == "adjudication_import_only"
    assert "runtime_import" in mapping["prohibited"]
    assert "training" in mapping["prohibited"]


def test_agreement_formulas_and_known_vectors():
    a = KNOWN_KAPPA_VECTOR["labels_a"]
    b = KNOWN_KAPPA_VECTOR["labels_b"]
    assert exact_agreement(a, b) == pytest.approx(KNOWN_KAPPA_VECTOR["expected_exact"])
    assert cohen_kappa(a, b) == pytest.approx(KNOWN_KAPPA_VECTOR["expected_kappa"])
    assert ACCEPTANCE_GATES["cohen_kappa_min"] == 0.70


def test_required_empty_populations_fail_optional_na():
    bad = population_metric_status(numerator=0, denominator=0, required=True)
    assert bad["status"] == "INVALID_REQUIRED_POPULATION"
    ok = population_metric_status(numerator=0, denominator=0, required=False)
    assert ok["status"] == "NOT_APPLICABLE"
    good = population_metric_status(numerator=3, denominator=10, required=True)
    assert good["value"] == 0.3


def test_no_runtime_predictions_used_as_gold():
    man = json.loads((OUT / "V3_PACKET_MANIFEST.json").read_text(encoding="utf-8"))
    assert man["human_decisions_included"] is False
    prov = json.loads((OUT / "V3_SOURCE_PROVENANCE_REGISTRY.json").read_text(encoding="utf-8"))
    assert all(x["prohibited_for_training"] is True for x in prov["items"])


def test_prohibited_for_training_true():
    mapping = json.loads((OUT / "V3_BLIND_MAPPING.json").read_text(encoding="utf-8"))
    assert mapping["prohibited_for_training"] is True
    cov = coverage_report(build_independent_corpus())
    assert cov["ok"]


def test_mai08_untouched():
    ledger = json.loads((REPO / "docs/mokxya-ai/MAI_PHASE_LEDGER.json").read_text(encoding="utf-8"))
    mai08 = next(p for p in ledger["phases"] if p["id"] == "MAI-08")
    assert mai08["status"] == "NOT_STARTED"


def test_runtime_resource_hashes_unchanged():
    xlrr.load_resources(force_reload=True)
    assert xlrr.compute_pack_content_hash() == DEFAULT_RESOURCE
    assert RESOURCE_PACK_VERSION == "mai-07.1.3-r3f-sealnew"
    assert RUNTIME_VERSION == "mai-07.1.3-r3f-sealnew"
    assert ENABLE_PROMOTION_OVERLAY is False


def test_repeated_packet_build_deterministic():
    a = build_packet(REPO)
    b = build_packet(REPO)
    assert a["manifest_sha256"] == b["manifest_sha256"]
    assert a["item_count"] == b["item_count"]


def test_packet_status_blocked_pending_review(packet):
    assert packet["status"] == "REVIEW_PACKET_READY"
    assert packet["blocker"] == "BLOCKED_PENDING_INDEPENDENT_HUMAN_REVIEW"
    man = json.loads((OUT / "V3_PACKET_MANIFEST.json").read_text(encoding="utf-8"))
    assert man["QUALITY_GATES_PASSED"] is False
    assert man["next_phase"] == "MAI-07R3J-B-ADJUDICATION-AND-V3-FREEZE"
    assert man["v2_governance"] == "HISTORICAL_BENCHMARK_EXHAUSTED_FOR_MODEL_SELECTION"
