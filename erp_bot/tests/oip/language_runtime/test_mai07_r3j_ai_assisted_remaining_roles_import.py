"""MAI-07R3J-AI-ASSISTED-REMAINING-ROLES-IMPORT tests."""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.oip.modules.language_runtime.transliteration import (
    ENABLE_PROMOTION_OVERLAY,
    RUNTIME_VERSION,
)
from src.oip.modules.language_runtime.transliteration.application.import_mai07_r3j_ai_assisted_remaining_roles import (
    DEFAULT_EVIDENCE_ROOT,
    DRAFTS_ROOT,
    EXPECTED_TOTAL_ROWS,
    FIXED_PROVENANCE,
    OFFICIAL_INBOX,
    ROLES,
    assert_official_inbox_untouched,
    import_all_remaining_roles,
    import_role,
    prove_deterministic_reimport,
    sha256_file,
)
from src.oip.modules.language_runtime.transliteration.infrastructure import (
    resource_repository as xlrr,
)

ACTIVE_RESOURCE_HASH = "1617425373bf525968b5af2a3b1cc8b8e5ad83e68457cfbbb47c73c78c84e930"
SUMMARY = DEFAULT_EVIDENCE_ROOT / "reports" / "IMPORT_REPORT.json"


@pytest.fixture(scope="module")
def bundle(tmp_path_factory):
    evidence = tmp_path_factory.mktemp("remaining_verified")
    result = import_all_remaining_roles(
        drafts_root=DRAFTS_ROOT, evidence_root=evidence, write_evidence=True
    )
    assert result.ok, result.errors
    return result, evidence


def test_three_roles_3333_rows(bundle):
    result, _ = bundle
    assert result.total_rows == EXPECTED_TOTAL_ROWS == 3333
    assert len(result.roles) == 3
    for r in result.roles:
        assert r.ok
        assert len(r.records) == 1111
        assert len(r.workbook_hashes) == 10
        assert len({x.review_id for x in r.records}) == 1111


def test_governance_flags_hardcoded(bundle):
    result, _ = bundle
    for role in result.roles:
        for rec in role.records:
            assert rec.independent_human_review is False
            assert rec.ai_autofill_used is True
            assert rec.user_accepted is True
            assert rec.professional_linguist_adjudication is False
            assert rec.prohibited_for_training is True
            assert rec.eligible_for_frozen_quality_gold is False
            if role.role_id == "PROFESSIONAL_LINGUIST_B":
                assert rec.professional_linguist_b_is_ai_role_simulation is True
            else:
                assert rec.professional_linguist_b_is_ai_role_simulation is False
    assert FIXED_PROVENANCE["linguist_approved"] is False
    assert FIXED_PROVENANCE["professional_linguist_b_is_ai_role_simulation"] is True
    assert FIXED_PROVENANCE["official_round_a_lock_eligible"] is False


def test_repo_evidence_materialized():
    assert SUMMARY.is_file()
    report = json.loads(SUMMARY.read_text(encoding="utf-8"))
    assert report["ok"] is True
    assert report["total_rows"] == 3333
    assert report["governance"]["LINGUIST_APPROVED"] is False
    assert report["governance"]["ROUND_A_LOCKED"] is False
    assert report["governance"]["professional_linguist_b_is_ai_role_simulation"] is True
    assert report["provenance"]["user_confirmation"] == (
        "USER_ACCEPTED_AI_SUGGESTIONS_WITHOUT_CHANGES"
    )


def test_authority_unchanged_and_enums(bundle):
    result, evidence = bundle
    for role in result.roles:
        for rec in role.records[:5]:
            assert rec.disposition
            assert rec.confidence in {"HIGH", "MEDIUM", "LOW"}
        # spot-check evidence workbook exists
        wb_dir = evidence / role.role_id.lower() / "evidence" / "workbooks"
        assert len(list(wb_dir.glob("*VERIFIED.xlsx"))) == 10


def test_unknown_disposition_rejected(tmp_path):
    # Clone one role's drafts and corrupt a disposition
    src = DRAFTS_ROOT / "product_policy" / "round_a_drafts"
    dest = tmp_path / "product_policy" / "round_a_drafts"
    shutil.copytree(src, dest)
    wb_path = sorted(dest.glob("*DRAFT.xlsx"))[0]
    wb = load_workbook(wb_path)
    wb["ROUND_A_CONTEXT"]["D2"] = "NOT_A_DISPOSITION"
    wb.save(wb_path)
    wb.close()
    fake_root = tmp_path
    bad = import_role(
        "PRODUCT_POLICY",
        drafts_root=fake_root,
        evidence_root=tmp_path / "ev",
        write_evidence=False,
    )
    assert not bad.ok
    assert any("unknown_disposition" in e for e in bad.errors)


def test_changed_authority_rejected(tmp_path):
    src = DRAFTS_ROOT / "nepali_fluent_a" / "round_a_drafts"
    dest = tmp_path / "nepali_fluent_a" / "round_a_drafts"
    shutil.copytree(src, dest)
    wb_path = sorted(dest.glob("*DRAFT.xlsx"))[0]
    wb = load_workbook(wb_path)
    wb["ROUND_A_CONTEXT"]["B2"] = "TAMPERED"
    wb.save(wb_path)
    wb.close()
    bad = import_role(
        "NEPALI_FLUENT_A",
        drafts_root=tmp_path,
        evidence_root=tmp_path / "ev",
        write_evidence=False,
    )
    assert not bad.ok
    assert any("changed_input_text" in e for e in bad.errors)


def test_populated_declaration_rejected(tmp_path):
    src = DRAFTS_ROOT / "professional_linguist_b" / "round_a_drafts"
    dest = tmp_path / "professional_linguist_b" / "round_a_drafts"
    shutil.copytree(src, dest)
    wb_path = sorted(dest.glob("*DRAFT.xlsx"))[0]
    wb = load_workbook(wb_path)
    for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=2, max_col=2):
        if row[0].value == "reviewer_full_name":
            row[1].value = "Fake Linguist"
            break
    wb.save(wb_path)
    wb.close()
    bad = import_role(
        "PROFESSIONAL_LINGUIST_B",
        drafts_root=tmp_path,
        evidence_root=tmp_path / "ev",
        write_evidence=False,
    )
    assert not bad.ok
    assert any("populated_reviewer_declaration" in e for e in bad.errors)


def test_deterministic_reimport(tmp_path):
    proof = prove_deterministic_reimport(drafts_root=DRAFTS_ROOT, work_dir=tmp_path / "det")
    assert proof["ok"] is True
    assert proof["total_rows"] == 3333
    report = json.loads(SUMMARY.read_text(encoding="utf-8"))
    assert proof["semantic_hash"] == report["semantic_hash"]


def test_official_inbox_untouched():
    assert_official_inbox_untouched()
    assert set(ROLES) == {
        "PRODUCT_POLICY",
        "NEPALI_FLUENT_A",
        "PROFESSIONAL_LINGUIST_B",
    }


def test_runtime_unchanged():
    assert RUNTIME_VERSION == "mai-07.1.3-r3f-sealnew"
    assert ENABLE_PROMOTION_OVERLAY is False
    report = xlrr.validate_resources()
    assert report["ok"] is True
    assert report["content_hash"] == ACTIVE_RESOURCE_HASH
