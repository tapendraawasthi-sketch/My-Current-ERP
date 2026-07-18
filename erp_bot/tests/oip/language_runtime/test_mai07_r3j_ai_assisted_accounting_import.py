"""MAI-07R3J-AI-ASSISTED-ACCOUNTING-IMPORT — integrity and governance tests."""

from __future__ import annotations

import hashlib
import json
import shutil
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.oip.modules.language_runtime.transliteration import (
    ENABLE_PROMOTION_OVERLAY,
    RUNTIME_VERSION,
)
from src.oip.modules.language_runtime.transliteration.application.import_mai07_r3j_ai_assisted_accounting import (
    EXPECTED_PACKAGE_SHA256,
    EXPECTED_ROW_COUNT,
    EXPECTED_WORKBOOK_COUNT,
    FIXED_PROVENANCE,
    OFFICIAL_OPS_INBOX,
    Mai07AiAssistedAccountingImportError,
    assert_no_accounting_mutation_imports,
    assert_official_inbox_untouched,
    compute_semantic_hash,
    import_from_round_a_dir,
    import_package,
    prove_deterministic_reimport,
    sha256_file,
    verify_package_sha256,
)
from src.oip.modules.language_runtime.transliteration.infrastructure import (
    resource_repository as xlrr,
)

REPO = Path(__file__).resolve().parents[4]
ZIP_PATH = REPO / "MokXya_MAI07_V3_ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED_READY.zip"
EVIDENCE = REPO / "docs/mokxya-ai/reviews/mai07_v3_ai_assisted/accounting_domain"
REVIEW_STATUS = REPO / "docs/mokxya-ai/reviews/mai07_v3/review_operations/REVIEW_STATUS.json"
ACTIVE_RESOURCE_HASH = "1617425373bf525968b5af2a3b1cc8b8e5ad83e68457cfbbb47c73c78c84e930"


@pytest.fixture(scope="module")
def package_available():
    if not ZIP_PATH.is_file():
        pytest.skip("AI-assisted accounting package ZIP not present")
    return ZIP_PATH


@pytest.fixture(scope="module")
def import_result(package_available, tmp_path_factory):
    extract = tmp_path_factory.mktemp("ai_acct_extract")
    evidence = tmp_path_factory.mktemp("ai_acct_evidence")
    result = import_package(
        package_available,
        extract_dir=extract,
        evidence_root=evidence,
        write_evidence=True,
    )
    assert result.ok, result.errors
    return result, extract, evidence


def _clone_round_a(src_round_a: Path, dest: Path) -> Path:
    dest.mkdir(parents=True, exist_ok=True)
    shutil.copytree(src_round_a / "ACCOUNTING_DOMAIN" / "round_a", dest / "round_a")
    return dest / "round_a"


def test_package_sha256_exact(package_available):
    assert verify_package_sha256(package_available) == EXPECTED_PACKAGE_SHA256


def test_six_workbooks_611_rows_unique_ids(import_result):
    result, _, _ = import_result
    assert len(result.workbook_hashes) == EXPECTED_WORKBOOK_COUNT
    assert len(result.records) == EXPECTED_ROW_COUNT
    assert len({r.review_id for r in result.records}) == EXPECTED_ROW_COUNT


def test_source_manifest_reconciliation(import_result):
    result, extract, _ = import_result
    round_a = extract / "ACCOUNTING_DOMAIN" / "round_a"
    source = json.loads((round_a / "SOURCE_BATCH_MANIFEST.json").read_text(encoding="utf-8"))
    expected = {rid for b in source["batches"] for rid in b["review_ids"]}
    assert {r.review_id for r in result.records} == expected


def test_verified_sha256_enforcement(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    # tamper one workbook byte
    wb = next(round_a.glob("*VERIFIED.xlsx"))
    data = bytearray(wb.read_bytes())
    data[-1] ^= 0x01
    wb.write_bytes(bytes(data))
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("workbook_sha256_mismatch" in e for e in bad.errors)


def test_unknown_disposition_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    # rewrite workbook then fix manifest hash to isolate enum failure
    wb = load_workbook(wb_path)
    ws = wb["ROUND_A_CONTEXT"]
    ws["D2"] = "NOT_A_REAL_DISPOSITION"
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("unknown_disposition" in e for e in bad.errors)


def test_unknown_confidence_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    wb = load_workbook(wb_path)
    wb["ROUND_A_CONTEXT"]["E2"] = "ULTRA"
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("unknown_confidence" in e for e in bad.errors)


def test_missing_supporting_field_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    wb = load_workbook(wb_path)
    wb["ROUND_A_CONTEXT"]["F2"] = ""
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("missing_supporting_field" in e for e in bad.errors)


def test_duplicate_review_id_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    wb = load_workbook(wb_path)
    ws = wb["ROUND_A_CONTEXT"]
    # copy first data review_id onto second row
    ws["A3"] = ws["A2"].value
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any(
        "duplicate_review_ids" in e or "batch_id_mismatch" in e for e in bad.errors
    )


def test_changed_authority_field_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    wb = load_workbook(wb_path)
    wb["ROUND_A_CONTEXT"]["B2"] = "TAMPERED_INPUT_TEXT"
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("changed_input_text" in e for e in bad.errors)


def test_populated_declaration_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    wb = load_workbook(wb_path)
    # set reviewer_full_name
    for row in wb["REVIEWER_DECLARATION"].iter_rows(min_row=2, max_col=2):
        if row[0].value == "reviewer_full_name":
            row[1].value = "Fabricated Person"
            break
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("populated_reviewer_declaration" in e for e in bad.errors)


def test_populated_round_b_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    wb = load_workbook(wb_path)
    ws = wb["ROUND_B_CANDIDATES"]
    ws["A2"] = "V3R-0144779dffa0"
    ws["F2"] = "ACCEPTABLE_PREFERRED"
    wb.save(wb_path)
    wb.close()
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("round_b" in e for e in bad.errors)


def test_macro_enabled_rejection(import_result, tmp_path):
    _, extract, _ = import_result
    round_a = _clone_round_a(extract, tmp_path / "pkg")
    wb_path = sorted(round_a.glob("*VERIFIED.xlsx"))[0]
    # inject fake vbaProject.bin into xlsx zip
    tmp_xlsm = tmp_path / "macro.xlsx"
    with zipfile.ZipFile(wb_path, "r") as zin, zipfile.ZipFile(tmp_xlsm, "w") as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr("xl/vbaProject.bin", b"FAKE_VBA")
    wb_path.write_bytes(tmp_xlsm.read_bytes())
    _rewrite_manifest_hash(round_a, wb_path)
    bad = import_from_round_a_dir(round_a, write_evidence=False)
    assert not bad.ok
    assert any("macro" in e for e in bad.errors)


def test_deterministic_import_and_semantic_hash(package_available, tmp_path):
    proof = prove_deterministic_reimport(package_available, work_dir=tmp_path / "det")
    assert proof["ok"] is True
    assert proof["row_count"] == EXPECTED_ROW_COUNT
    assert len(proof["semantic_hash"]) == 64


def test_governance_flags_hardcoded(import_result):
    result, _, _ = import_result
    for r in result.records:
        assert r.independent_human_review is False
        assert r.ai_autofill_used is True
        assert r.professional_linguist_adjudication is False
        assert r.prohibited_for_training is True
        assert r.eligible_for_frozen_quality_gold is False
        assert r.review_method == "AI_ASSISTED_HUMAN_VERIFIED"
    assert FIXED_PROVENANCE["linguist_approved"] is False
    assert FIXED_PROVENANCE["production_approved"] is False
    assert FIXED_PROVENANCE["official_round_a_lock_eligible"] is False
    assert FIXED_PROVENANCE["round_b_authorized"] is False
    assert FIXED_PROVENANCE["frozen_v3_quality_gate_authorized"] is False


def test_official_round_a_inbox_untouched():
    assert_official_inbox_untouched()
    if OFFICIAL_OPS_INBOX.exists():
        assert not list(OFFICIAL_OPS_INBOX.glob("*.xlsx"))
        assert not list(OFFICIAL_OPS_INBOX.glob("*.xlsm"))
    # evidence must not live under official inbox
    assert "round_a_inbox" not in str(EVIDENCE).replace("\\", "/")


def test_runtime_and_resources_unchanged():
    assert RUNTIME_VERSION == "mai-07.1.3-r3f-sealnew"
    assert ENABLE_PROMOTION_OVERLAY is False
    report = xlrr.validate_resources()
    assert report["ok"] is True
    assert report["content_hash"] == ACTIVE_RESOURCE_HASH
    # review ops status must remain waiting (not locked by this import)
    status = json.loads(REVIEW_STATUS.read_text(encoding="utf-8"))
    assert status.get("ROUND_A_LOCKED") is not True
    assert status.get("ROUND_B_READY") is not True
    state = status.get("workflow_state") or status.get("state") or ""
    assert "ROUND_B" not in str(state) or "WAITING" in str(status)


def test_no_accounting_mutation_path():
    assert_no_accounting_mutation_imports()


def test_canonical_repo_evidence_if_materialized():
    """If the governed evidence path exists, it must match importer invariants."""
    jsonl = EVIDENCE / "canonical" / "ACCOUNTING_DOMAIN_ROUND_A_AI_ASSISTED_HUMAN_VERIFIED.jsonl"
    if not jsonl.is_file():
        pytest.skip("canonical evidence not yet materialized in repo")
    lines = [ln for ln in jsonl.read_text(encoding="utf-8").splitlines() if ln.strip()]
    assert len(lines) == EXPECTED_ROW_COUNT
    report = json.loads((EVIDENCE / "reports" / "IMPORT_REPORT.json").read_text(encoding="utf-8"))
    assert report["ok"] is True
    assert report["governance"]["QUALITY_GATES_PASSED"] is False
    assert report["governance"]["LINGUIST_APPROVED"] is False
    assert report["governance"]["PRODUCTION_APPROVED"] is False
    assert report["governance"]["ROUND_A_LOCKED"] is False
    assert report["governance"]["ROUND_B_READY"] is False
    assert report["row_count"] == EXPECTED_ROW_COUNT


def test_semantic_hash_stable_for_same_records(import_result):
    result, _, _ = import_result
    h1 = compute_semantic_hash(result.records)
    h2 = compute_semantic_hash(list(reversed(result.records)))
    assert h1 == h2 == result.semantic_hash


def _rewrite_manifest_hash(round_a: Path, workbook_path: Path) -> None:
    """Update verified_sha256 in package manifest after intentional mutation."""
    man_path = round_a / "AI_ASSISTED_HUMAN_VERIFIED_MANIFEST.json"
    man = json.loads(man_path.read_text(encoding="utf-8"))
    digest = sha256_file(workbook_path)
    for f in man["files"]:
        if f["filename"] == workbook_path.name:
            f["verified_sha256"] = digest
            f["size_bytes"] = workbook_path.stat().st_size
    man_path.write_text(json.dumps(man, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
